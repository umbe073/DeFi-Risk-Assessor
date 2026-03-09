#!/usr/bin/env python3
import ast
import pandas as pd
import json
import os
import numpy as np
import time
import sys
import threading
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
from working_progress_bar import update_progress_phase, finish_progress_bar, working_progress_bar

SCRIPT_DIR = os.path.abspath(os.path.dirname(__file__))
PROJECT_ROOT = SCRIPT_DIR
DEFAULT_DATA_DIR = os.path.join(PROJECT_ROOT, 'data')
# Backward-compatible fallback for repos keeping data files at project root.
DATA_DIR = DEFAULT_DATA_DIR if os.path.isdir(DEFAULT_DATA_DIR) else PROJECT_ROOT

print("[DEBUG] Using DATA_DIR:", DATA_DIR)
xlsx_file = os.path.join(DATA_DIR, 'DeFi Tokens Risk Assessment Results.xlsx')
print("[DEBUG] Excel file path:", xlsx_file)
json_file = os.path.join(DATA_DIR, 'risk_report.json')

# Load symbol and name mapping from JSON
with open(os.path.join(DATA_DIR, 'cmc_symbol_map.json'), 'r') as f:
    CMC_MAP = json.load(f)
RED_FLAGS = [
    'unverified_contract',
    'low_liquidity',
    'high_concentration',
    'is_proxy_contract',
    'eu_unlicensed_stablecoin',
    'eu_regulatory_issues',
    'mica_non_compliant',
    'mica_no_whitepaper'
]
COMPONENTS = ['industry_impact','tech_innovation','whitepaper_quality','roadmap_adherence','business_model','team_expertise','management_strategy','global_reach','code_security','dev_activity','aml_data','compliance_data','market_dynamics','marketing_demand','esg_impact']

# --- Simple Console Progress Bar (No AppleEvents) ---
class ConsoleProgressBar:
    def __init__(self, total_items, description="Processing"):
        self.total_items = total_items
        self.description = description
        self.completed = 0
        self.lock = threading.Lock()
        self.start_time = time.time()
        self.last_update = 0
        self.update_interval = 0.5  # Update every 0.5 seconds
        
    def update(self, completed=None, message=""):
        with self.lock:
            if completed is not None:
                self.completed = completed
            else:
                self.completed += 1
                
            now = time.time()
            if now - self.last_update < self.update_interval and self.completed < self.total_items:
                return  # Skip update if too soon and not final
            self.last_update = now

            # Handle empty workloads gracefully (e.g. empty risk_report.json).
            bar_length = 30
            if self.total_items <= 0:
                percent = 100
                eta_str = "ETA: 0.0s"
                filled_length = bar_length
            else:
                percent = int((self.completed / self.total_items) * 100)
                elapsed = now - self.start_time
                # Calculate ETA
                if self.completed > 0:
                    eta = (elapsed / self.completed) * (self.total_items - self.completed)
                    eta_str = f"ETA: {eta:.1f}s"
                else:
                    eta_str = "ETA: --"
                # Create progress bar
                filled_length = int(bar_length * self.completed // self.total_items)
            bar = '█' * filled_length + '-' * (bar_length - filled_length)
            
            # Clear line and print progress
            sys.stdout.write(f'\r{self.description}: [{bar}] {percent}% ({self.completed}/{self.total_items}) {eta_str} {message}')
            sys.stdout.flush()
            
            if self.completed >= self.total_items:
                print()  # New line when complete
    
    def finish(self, message="Complete!"):
        self.update(self.total_items, message)
        print()

# Global progress bar instance
progress_bar = None

def launch_progress_bar():
    """Initialize the console progress bar"""
    global progress_bar
    progress_bar = ConsoleProgressBar(1, "Updating Excel report")

def update_progress_bar(percent, message):
    """Update the console progress bar"""
    global progress_bar
    if progress_bar:
        total_items = max(progress_bar.total_items, 1)
        completed = int((percent / 100) * total_items)
        progress_bar.update(completed, message)

def close_progress_bar():
    """Finish the console progress bar"""
    global progress_bar
    if progress_bar:
        progress_bar.finish("Excel update complete!")
        progress_bar = None


def main():
    print("Analysis Complete")  # Only print at the start of the Excel update script
    # Load the spreadsheet
    df = pd.read_excel(xlsx_file)
    # Load the JSON report
    with open(json_file) as report_file:
        raw_data = json.load(report_file)
    if isinstance(raw_data, list):
        data = raw_data
    elif isinstance(raw_data, dict):
        data = [raw_data]
    else:
        data = []

    # Show 'Generating final reports...' in the web progress bar
    update_progress_phase(2, "Generating final reports...")

    def ensure_column(dataframe, column_name, insert_after=None, default_value=''):
        """Insert a missing column in a deterministic location."""
        if column_name in dataframe.columns:
            return dataframe
        cols = list(dataframe.columns)
        insert_idx = len(cols)
        if insert_after and insert_after in cols:
            insert_idx = cols.index(insert_after) + 1
        cols.insert(insert_idx, column_name)
        dataframe[column_name] = default_value
        return dataframe[cols]

    # Ensure core columns exist in the expected order.
    df = ensure_column(df, 'Token Address', insert_after='Token Name')
    df = ensure_column(df, 'Symbol', insert_after='Token Address')
    df = ensure_column(df, 'Is Stablecoin', insert_after='Symbol')
    df = ensure_column(df, 'EU Compliance Status', insert_after='Is Stablecoin')

    # Ensure all red flag columns exist and are ordered together in the middle
    # Find the index after which to insert red flag columns (after 'EU Compliance Status')
    if 'EU Compliance Status' in df.columns:
        base_cols = list(df.columns)
        eu_idx = base_cols.index('EU Compliance Status')
        # Remove any existing red flag columns
        for flag in RED_FLAGS:
            col = f'Red Flag: {flag}'
            if col in base_cols:
                base_cols.remove(col)
        # Insert red flag columns after 'EU Compliance Status'
        for i, flag in enumerate(RED_FLAGS):
            base_cols.insert(eu_idx + 1 + i, f'Red Flag: {flag}')
        # Now, ensure all other columns (Chain, Risk Score, etc.) remain in their original left positions
        # by moving only the red flag columns, not the rest
        # The left columns should be:
        left_cols = [
            'Token Name', 'Token Address', 'Symbol', 'Is Stablecoin', 'EU Compliance Status',
            'Chain', 'Risk Score', 'Risk Category', 'Market Cap', 'Volume 24h', 'Holders', 'Liquidity'
        ]
        # Remove any of these from base_cols if present (to avoid duplicates)
        for col in left_cols:
            if col in base_cols:
                base_cols.remove(col)
        # Prepend left_cols to the reordered base_cols
        new_cols = left_cols + base_cols
        df = df.reindex(columns=new_cols)

    # Helper: get row index by token address
    def find_row(token_addr):
        token_column = df.get('Token Address')
        if token_column is None:
            return None
        normalized_addresses = token_column.fillna('').astype(str).str.lower()
        matches = df.index[normalized_addresses == token_addr.lower()].tolist()
        return matches[0] if matches else None

    # At the start of main execution (before updating Excel)
    launch_progress_bar()
    progress_bar.total_items = len(data)

    for idx, entry in enumerate(data):
        if not isinstance(entry, dict):
            update_progress_bar(((idx + 1) / max(len(data), 1)) * 100, f"Skipping invalid entry {idx + 1}")
            continue
        token_value = str(entry.get('token') or '').strip()
        if not token_value:
            update_progress_bar(((idx + 1) / max(len(data), 1)) * 100, f"Skipping entry {idx + 1} (missing token)")
            continue

        token_addr = token_value.lower()
        # Get symbol and name from CMC_MAP if available
        symbol = CMC_MAP.get(token_addr, {}).get('symbol', entry.get('symbol', ''))
        token_name = CMC_MAP.get(token_addr, {}).get('name', symbol or token_addr)
        row_idx = find_row(token_addr)
        key_metrics = entry.get('key_metrics') if isinstance(entry.get('key_metrics'), dict) else {}
        component_scores = entry.get('component_scores') if isinstance(entry.get('component_scores'), dict) else {}
        # Prepare update dict
        update = {
            'Token Name': token_name,
            'Token Address': token_value,
            'Symbol': symbol,
            'Is Stablecoin': 'Yes' if entry.get('is_stablecoin', False) else 'No',
            'EU Compliance Status': entry.get('eu_compliance_status', 'Unknown'),
            'Chain': entry.get('chain', ''),
            'Risk Score': entry.get('risk_score', 0),
            'Risk Category': entry.get('risk_category', 'Unknown'),
            'Market Cap': key_metrics.get('market_cap', 0),
            'Volume 24h': key_metrics.get('volume_24h', 0),
            'Holders': key_metrics.get('holders', 0),
            'Liquidity': key_metrics.get('liquidity', 0),
        }
        # Red flags as columns (ensure accuracy)
        red_flags_list = entry.get('red_flags', [])
        if isinstance(red_flags_list, str):
            # If red_flags is a string, try to parse it
            try:
                # Handle both list string format and comma-separated format
                if red_flags_list.startswith('[') and red_flags_list.endswith(']'):
                    parsed_flags = ast.literal_eval(red_flags_list) if red_flags_list else []
                    red_flags_list = [str(flag).strip() for flag in parsed_flags if str(flag).strip()] if isinstance(parsed_flags, list) else []
                else:
                    # Handle comma-separated format
                    red_flags_list = [flag.strip() for flag in red_flags_list.split(',') if flag.strip()]
            except (ValueError, SyntaxError):
                red_flags_list = []
        elif not isinstance(red_flags_list, list):
            red_flags_list = []
            
        for flag in RED_FLAGS:
            update[f'Red Flag: {flag}'] = 'Yes' if flag in red_flags_list else 'No'
        # Component scores
        for comp in COMPONENTS:
            col = comp.replace('_', ' ').title()
            if comp == 'esg_impact':
                col = 'Esg Impact'
            update[col] = component_scores.get(comp, '')
        # Update or append
        if row_idx is not None:
            for k, v in update.items():
                if k in df.columns:
                    # Handle numeric columns properly
                    if np.issubdtype(df[k].dtype, np.number):
                        # Convert empty strings to 0 for numeric columns
                        if v == '' or v is None:
                            df.at[row_idx, k] = 0
                        else:
                            # Try to convert to numeric, fallback to 0
                            try:
                                df.at[row_idx, k] = float(v) if v != '' else 0
                            except (ValueError, TypeError):
                                df.at[row_idx, k] = 0
                    else:
                        df.at[row_idx, k] = v
        else:
            # Append as new row, filling all columns
            row = {col: update.get(col, '') for col in df.columns}
            df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)

        update_progress_bar(
            ((idx + 1) / max(len(data), 1)) * 100,
            f"Processed {idx + 1}/{len(data)} token entries"
        )

    # Remove any rows with NaN Token Name (these are extra rows)
    df = df.dropna(subset=['Token Name'])

    # Replace NaN values appropriately for each column type
    for col in df.columns:
        if np.issubdtype(df[col].dtype, np.number):
            df[col] = df[col].fillna(0)  # Fill numeric columns with 0
        else:
            df[col] = df[col].fillna('')  # Fill text columns with empty string

    # Save the updated file
    df.to_excel(xlsx_file, index=False)
    print("[DEBUG] Excel file updated at:", xlsx_file)

    # Update or add timestamp at the bottom of the first sheet
    wb = load_workbook(xlsx_file)
    ws = wb.active
    
    # Style the EU Compliance Status column with lighter yellow background
    try:
        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")  # Bright yellow
        
        # Find the EU Compliance Status column
        eu_compliance_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "EU Compliance Status":
                eu_compliance_col = col
                break
        
        # Apply yellow background only to specific values
        if eu_compliance_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=eu_compliance_col)
                if cell.value in [
                    'Non-Compliant (Unlicensed Stablecoin)',
                    'Non-Compliant (Regulatory Issues)'
                ]:
                    cell.fill = yellow_fill
                else:
                    cell.fill = PatternFill(fill_type=None)  # Default/no fill
    except Exception as e:
        print(f"Warning: Could not apply styling to EU Compliance Status column: {e}")
    
    # Remove all but the last 'Last Updated:' row
    last_updated_rows = []
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Last Updated:":
            last_updated_rows.append(row)
    # If more than one, delete all but the last
    if len(last_updated_rows) > 1:
        for row_idx in reversed(last_updated_rows[:-1]):
            ws.delete_rows(row_idx, 1)

    # Look for existing "Last Updated:" row (after cleanup)
    timestamp_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Last Updated:":
            timestamp_row = row
            break
    
    if timestamp_row is None:
        # Add new timestamp row if it doesn't exist
        timestamp_row = ws.max_row + 2  # Leave one blank row for better spacing
        ws.cell(row=timestamp_row, column=1, value="Last Updated:")
        ws.cell(row=timestamp_row, column=1).font = Font(bold=True)
    else:
        # If timestamp row exists, ensure there's a blank row before it
        if timestamp_row > 2 and ws.cell(row=timestamp_row-1, column=1).value is not None:
            ws.insert_rows(timestamp_row, 1)
            timestamp_row += 1
    
    # Update the timestamp (clear any existing data in the row first)
    for col in range(2, ws.max_column + 1):
        ws.cell(row=timestamp_row, column=col).value = None
    ws.cell(row=timestamp_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    wb.save(xlsx_file)

    # At the end of the script (after all processing is done)
    close_progress_bar()
    if working_progress_bar:
        working_progress_bar.completed_phases = working_progress_bar.total_phases
    finish_progress_bar("Risk Assessment Completed!")

    # Wait for progress bar to close before showing notification
    import time
    time.sleep(3)  # Wait for the webpage to close

    # Show completion dialog with disclaimer (after webpage closes)
    try:
        import subprocess
        completion_script = f'''
        tell application "System Events"
            display dialog "✅ Update Completed!

📊 Risk assessment has been completed successfully!

📁 Results saved to:
• {xlsx_file}
• {json_file}
• {os.path.join(PROJECT_ROOT, "logs", "risk_assessment_summary.txt")}

📝 Check the 'data/' directory for detailed reports and the 'logs/' directory for execution logs.

⚠️ DISCLAIMER:
This tool provides automated risk assessment based on available data and should be used as part of a comprehensive due diligence process. Results are for informational purposes only and do not constitute financial advice. Always conduct your own research and consult with qualified professionals before making investment decisions.

Market data provided by CoinGecko (https://www.coingecko.com)" with title "DeFi Risk Assessment - Complete" buttons {{"OK"}} default button "OK"
        end tell
        '''
        subprocess.run(['osascript', '-e', completion_script], check=True)
    except Exception as e:
        print(f"Could not show completion dialog: {e}")
        print("✅ Risk assessment completed successfully!")
        print(f"📊 Check the 'data/' directory for results")
        print(f"📝 Check the 'logs/' directory for detailed logs")

    # Remove the flag file so the next run will wait again
    flag_path = os.path.join(DATA_DIR, "risk_report.done")
    if os.path.exists(flag_path):
        os.remove(flag_path)

if __name__ == "__main__":
    main() 