#!/usr/bin/env python3
import pandas as pd
import json
import os
import numpy as np
import time
import sys
import threading
import requests
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
# Import working progress bar
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

try:
    from working_progress_bar import (
        update_progress_phase,
        complete_phase_progress,
        finish_progress_bar,
        working_progress_bar
    )
    print("Using working progress bar for Excel update")
except ImportError as e:
    print(f"Warning: Could not import working_progress_bar: {e}")
    print("Using console progress bar fallback for Excel update")
    
    # Fallback implementation if import fails
    def update_progress_phase(phase, message):
        """Update progress phase"""
        print(f"Phase {phase}: {message}")

    def complete_phase_progress(message):
        """Complete phase progress"""
        print(f"Phase complete: {message}")

    def finish_progress_bar(message):
        """Finish progress bar"""
        print(f"Complete: {message}")

    # Global progress bar instance
    working_progress_bar = None

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
DATA_DIR = os.path.join(PROJECT_ROOT, 'data')

# Create data directory if it doesn't exist
os.makedirs(DATA_DIR, exist_ok=True)

print("[DEBUG] Using DATA_DIR:", DATA_DIR)
xlsx_file = os.path.join(DATA_DIR, 'DeFi Tokens Risk Assessment Results.xlsx')
print("[DEBUG] Excel file path:", xlsx_file)
json_file = os.path.join(DATA_DIR, 'risk_report.json')

# Load symbol and name mapping from JSON
cmc_map_path = os.path.join(DATA_DIR, 'cmc_symbol_map.json')
if os.path.exists(cmc_map_path):
    with open(cmc_map_path, 'r') as f:
        CMC_MAP = json.load(f)
else:
    print("[WARNING] cmc_symbol_map.json not found, using empty mapping")
    CMC_MAP = {}
RED_FLAGS = [
    'unverified_contract',
    'low_liquidity',
    'high_concentration',
    'is_proxy_contract',
    'eu_unlicensed_stablecoin',
    'eu_regulatory_issues',
    'mica_non_compliant',
    'mica_no_whitepaper',
    'owner_change_last_24h'
]
COMPONENTS = ['industry_impact','tech_innovation','whitepaper_quality','roadmap_adherence','business_model','team_expertise','management_strategy','global_reach','code_security','dev_activity','aml_data','compliance_data','market_dynamics','marketing_demand','esg_impact','social_data']

# --- Simple Console Progress Bar (No AppleEvents) ---

# --- Robust HTTP request with retries ---

def robust_request(method, url, retries=3, backoff=2, **kwargs):
    """Make an HTTP request with retries and exponential backoff."""
    for attempt in range(retries):
        try:
            response = requests.request(method, url, timeout=20, **kwargs)
            response.raise_for_status()
            return response
        except requests.RequestException as e:
            if attempt < retries - 1:
                time.sleep(backoff ** attempt)
            else:
                print(f"Request failed after {retries} attempts: {e}")
                return None

# --- Ethplorer bulk token info ---
def fetch_ethplorer_bulk(addresses, ethplorer_key):
    """Fetch token info for multiple addresses using Ethplorer's bulk endpoint."""
    url = f"https://api.ethplorer.io/bulkMonitor?apiKey={ethplorer_key}"
    try:
        resp = robust_request("POST", url, json={"addresses": addresses})
        if resp and resp.status_code == 200:
            return resp.json()
        else:
            print(f"Ethplorer bulkMonitor error: {resp.text if resp else 'No response'}")
            return None
    except Exception as e:
        print(f"Ethplorer bulkMonitor exception: {e}")
        return None

# --- Coinpaprika market data ---
def fetch_coinpaprika_market(symbol):
    """Fetch market data for a token symbol from Coinpaprika."""
    url = f"https://api.coinpaprika.com/v1/tickers/{symbol.lower()}-usd"
    try:
        resp = robust_request("GET", url)
        if resp and resp.status_code == 200:
            return resp.json()
        else:
            print(f"Coinpaprika market error: {resp.text if resp else 'No response'}")
            return None
    except Exception as e:
        print(f"Coinpaprika market exception: {e}")
        return None

# --- Dune Analytics query results ---
def fetch_dune_query(query_id, dune_key):
    """Fetch query results from Dune Analytics."""
    url = f"https://api.dune.com/api/v1/query/{query_id}/results"
    headers = {"x-dune-api-key": dune_key}
    try:
        resp = robust_request("GET", url, headers=headers)
        if resp and resp.status_code == 200:
            return resp.json()
        else:
            print(f"Dune query error: {resp.text if resp else 'No response'}")
            return None
    except Exception as e:
        print(f"Dune query exception: {e}")
        return None
class ConsoleProgressBar:
    def __init__(self, total_items, description="Processing"):
        self.total_items = total_items
        self.description = description
        self.completed = 0
        self.lock = threading.Lock()
        self.start_time = time.time()
        self.last_update = 0
        self.update_interval = 0.5  # Update every 0.5 seconds
        
# --- Etherscan token info for multiple addresses (parallelized) ---
def fetch_etherscan_tokeninfo(addresses, etherscan_key):
    """Fetch token info for multiple addresses from Etherscan (parallelized)."""
    def fetch_one(addr):
        url = "https://api.etherscan.io/api"
        params = {
            "module": "token",
            "action": "tokeninfo",
            "contractaddress": addr,
            "apikey": etherscan_key
        }
        try:
            resp = robust_request("GET", url, params=params)
            if resp and resp.status_code == 200:
                return addr, resp.json()
            else:
                return addr, None
        except Exception as e:
            print(f"Etherscan tokeninfo error for {addr}: {e}")
            return addr, None

    results = {}
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_addr = {executor.submit(fetch_one, addr): addr for addr in addresses}
        for future in as_completed(future_to_addr):
            addr, data = future.result()
            results[addr] = data
    return results

    def update(self, completed=None, message=""):
        with self.lock:
            if completed is not None:
                self.completed = completed
            else:
                self.completed += 1
                
            now = time.time()
            if now - self.last_update < self.update_interval and self.completed < self.total_items:
                return  # Skip update if too soon and not final
            self.last_update = now
            
            percent = int((self.completed / self.total_items) * 100)
            elapsed = now - self.start_time
            
            # Calculate ETA
            if self.completed > 0:
                eta = (elapsed / self.completed) * (self.total_items - self.completed)
                eta_str = f"ETA: {eta:.1f}s"
            else:
                eta_str = "ETA: --"
            
            # Create progress bar
            bar_length = 30
            filled_length = int(bar_length * self.completed // self.total_items)
            bar = '█' * filled_length + '-' * (bar_length - filled_length)
            
            # Clear line and print progress
            sys.stdout.write(f'\r{self.description}: [{bar}] {percent}% ({self.completed}/{self.total_items}) {eta_str} {message}')
            sys.stdout.flush()
            
            if self.completed >= self.total_items:
                print()  # New line when complete
    
    def finish(self, message="Complete!"):
        self.update(self.total_items, message)
        print()

# Global progress bar instance
progress_bar = None

def launch_progress_bar():
    """Initialize the console progress bar"""
    global progress_bar
    # This will be set in main function
    pass

def update_progress_bar(percent, message):
    """Update the console progress bar"""
    global progress_bar
    if progress_bar:
        completed = int((percent / 100) * progress_bar.total_items)
        progress_bar.update(completed, message)

def close_progress_bar():
    """Finish the console progress bar"""
    global progress_bar
    if progress_bar:
        progress_bar.finish("Excel update complete!")
        progress_bar = None


def main():
    print("Analysis Complete")  # Only print at the start of the Excel update script
    
    # Show progress messages for Excel report generation
    try:
        complete_phase_progress("Analysis complete")
        update_progress_phase(2, "Generating final reports...")
    except Exception as e:
        print(f"[ProgressBar] Error updating progress bar: {e}")
    
    # Check if files exist
    if not os.path.exists(json_file):
        print(f"[ERROR] JSON file not found: {json_file}")
        return
    
    if not os.path.exists(xlsx_file):
        print(f"[ERROR] Excel file not found: {xlsx_file}")
        print("Creating new Excel file...")
        # Create a basic Excel file with required columns
        component_columns = [comp.replace('_', ' ').title() for comp in COMPONENTS]
        component_columns = ['Esg Impact' if col == 'Esg Impact' else col for col in component_columns]
        
        df = pd.DataFrame(columns=[
            'Token Name', 'Token Address', 'Symbol', 'Is Stablecoin', 'EU Compliance Status',
            'Chain', 'Risk Score', 'Risk Category', 'Market Cap', 'Volume 24h', 'Holders', 'Liquidity'
        ] + [f'Red Flag: {flag}' for flag in RED_FLAGS] + component_columns)
        df.to_excel(xlsx_file, index=False)
    else:
        # Load the spreadsheet
        df = pd.read_excel(xlsx_file)
    
    # Load the JSON report
    try:
        with open(json_file, 'r') as f:
            data = json.load(f)
    except Exception as e:
        print(f"[ERROR] Failed to load JSON file: {e}")
        return

    # Show 'Generating final reports...' in the web progress bar
    update_progress_phase(2, "Generating final reports...")

    # Ensure 'Symbol' column exists after 'Token Address'
    if 'Symbol' not in df.columns:
        addr_idx = df.columns.get_loc('Token Address')
        cols = list(df.columns)
        cols.insert(addr_idx + 1, 'Symbol')
        df['Symbol'] = ''
        df = df[cols]
    
    # Ensure 'Is Stablecoin' column exists after 'Symbol'
    if 'Is Stablecoin' not in df.columns:
        symbol_idx = df.columns.get_loc('Symbol')
        cols = list(df.columns)
        cols.insert(symbol_idx + 1, 'Is Stablecoin')
        df['Is Stablecoin'] = ''
        df = df[cols]
    
    # Ensure 'EU Compliance Status' column exists after 'Is Stablecoin'
    if 'EU Compliance Status' not in df.columns:
        stablecoin_idx = df.columns.get_loc('Is Stablecoin')
        cols = list(df.columns)
        cols.insert(stablecoin_idx + 1, 'EU Compliance Status')
        df['EU Compliance Status'] = ''
        df = df[cols]

    # Ensure all red flag columns exist and are ordered together in the middle
    # Find the index after which to insert red flag columns (after 'EU Compliance Status')
    if 'EU Compliance Status' in df.columns:
        # Add component score columns if they don't exist
        component_columns = [comp.replace('_', ' ').title() for comp in COMPONENTS]
        component_columns = ['Esg Impact' if col == 'Esg Impact' else col for col in component_columns]
        
        for col in component_columns:
            if col not in df.columns:
                df[col] = ''
        base_cols = list(df.columns)
        eu_idx = base_cols.index('EU Compliance Status')
        # Remove any existing red flag columns
        for flag in RED_FLAGS:
            col = f'Red Flag: {flag}'
            if col in base_cols:
                base_cols.remove(col)
        # Insert red flag columns after 'EU Compliance Status'
        for i, flag in enumerate(RED_FLAGS):
            base_cols.insert(eu_idx + 1 + i, f'Red Flag: {flag}')
        # Now, ensure all other columns (Chain, Risk Score, etc.) remain in their original left positions
        # by moving only the red flag columns, not the rest
        # The left columns should be:
        left_cols = [
            'Token Name', 'Token Address', 'Symbol', 'Is Stablecoin', 'EU Compliance Status',
            'Chain', 'Risk Score', 'Risk Category', 'Market Cap', 'Volume 24h', 'Holders', 'Liquidity'
        ]
        # Remove any of these from base_cols if present (to avoid duplicates)
        for col in left_cols:
            if col in base_cols:
                base_cols.remove(col)
        # Prepend left_cols to the reordered base_cols
        new_cols = left_cols + base_cols
        df = df.reindex(columns=new_cols)

    # Helper: get row index by token address
    def find_row(token_addr):
        matches = df.index[df['Token Address'].str.lower() == token_addr.lower()].tolist()
        return matches[0] if matches else None

    # At the start of main execution (before updating Excel)
    launch_progress_bar()

    for entry in data:
        token_addr = entry['token'].lower()
        # Get symbol and name from CMC_MAP if available
        symbol = CMC_MAP.get(token_addr, {}).get('symbol', entry.get('symbol', ''))
        token_name = CMC_MAP.get(token_addr, {}).get('name', symbol or token_addr)
        row_idx = find_row(token_addr)
        # Prepare update dict
        update = {
            'Token Name': token_name,
            'Token Address': entry['token'],
            'Symbol': symbol,
            'Is Stablecoin': 'Yes' if entry.get('is_stablecoin', False) else 'No',
            'EU Compliance Status': entry.get('eu_compliance_status', 'Unknown'),
            'Chain': entry['chain'],
            'Risk Score': entry['risk_score'],
            'Risk Category': entry['risk_category'],
            'Market Cap': entry['key_metrics']['market_cap'],
            'Volume 24h': entry['key_metrics']['volume_24h'],
            'Holders': entry['key_metrics']['holders'],
            'Liquidity': entry['key_metrics']['liquidity'],
        }
        # Red flags as columns (ensure accuracy)
        red_flags_list = entry.get('red_flags', [])
        if isinstance(red_flags_list, str):
            # If red_flags is a string, try to parse it
            try:
                # Handle both list string format and comma-separated format
                if red_flags_list.startswith('[') and red_flags_list.endswith(']'):
                    red_flags_list = eval(red_flags_list) if red_flags_list else []
                else:
                    # Handle comma-separated format
                    red_flags_list = [flag.strip() for flag in red_flags_list.split(',') if flag.strip()]
            except:
                red_flags_list = []
        elif not isinstance(red_flags_list, list):
            red_flags_list = []
            
        for flag in RED_FLAGS:
            update[f'Red Flag: {flag}'] = 'Yes' if flag in red_flags_list else 'No'
        # Component scores
        for comp in COMPONENTS:
            col = comp.replace('_', ' ').title()
            if comp == 'esg_impact':
                col = 'Esg Impact'
            update[col] = entry['component_scores'].get(comp, '')
        # Update or append
        if row_idx is not None:
            for k, v in update.items():
                if k in df.columns:
                    # Handle numeric columns properly
                    if np.issubdtype(df[k].dtype, np.number):
                        # Convert empty strings to 0 for numeric columns
                        if v == '' or v is None:
                            df.at[row_idx, k] = 0
                        else:
                            # Try to convert to numeric, fallback to 0
                            try:
                                df.at[row_idx, k] = float(v) if v != '' else 0
                            except (ValueError, TypeError):
                                df.at[row_idx, k] = 0
                    else:
                        df.at[row_idx, k] = v
        else:
            # Append as new row, filling all columns
            row = {col: update.get(col, '') for col in df.columns}
            df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)

    # Remove any rows with NaN Token Name or Token Address (these are extra rows)
    df = df.dropna(subset=['Token Name', 'Token Address'])
    
    # Remove any rows that are timestamps or have non-address data in Token Address column
    df = df[df['Token Address'].str.match(r'^0x[a-fA-F0-9]{40}$', na=False)]

    # Replace NaN values appropriately for each column type
    for col in df.columns:
        if np.issubdtype(df[col].dtype, np.number):
            df[col] = df[col].fillna(0)  # Fill numeric columns with 0
        else:
            df[col] = df[col].fillna('')  # Fill text columns with empty string

    # Save the updated file
    df.to_excel(xlsx_file, index=False)
    print("[DEBUG] Excel file updated at:", xlsx_file)

    # Update or add timestamp at the bottom of the first sheet
    wb = load_workbook(xlsx_file)
    ws = wb.active
    
    # Style the EU Compliance Status column with lighter yellow background
    try:
        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")  # Bright yellow
        
        # Find the EU Compliance Status column
        eu_compliance_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "EU Compliance Status":
                eu_compliance_col = col
                break
        
        # Apply yellow background only to specific values
        if eu_compliance_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=eu_compliance_col)
                if cell.value in [
                    'Non-Compliant (Unlicensed Stablecoin)',
                    'Non-Compliant (Regulatory Issues)'
                ]:
                    cell.fill = yellow_fill
                else:
                    cell.fill = PatternFill(fill_type=None)  # Default/no fill
    except Exception as e:
        print(f"Warning: Could not apply styling to EU Compliance Status column: {e}")
    
    # Remove all but the last 'Last Updated:' row
    last_updated_rows = []
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Last Updated:":
            last_updated_rows.append(row)
    # If more than one, delete all but the last
    if len(last_updated_rows) > 1:
        for row_idx in reversed(last_updated_rows[:-1]):
            ws.delete_rows(row_idx, 1)

    # Look for existing "Last Updated:" row (after cleanup)
    timestamp_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Last Updated:":
            timestamp_row = row
            break
    
    if timestamp_row is None:
        # Add new timestamp row if it doesn't exist
        timestamp_row = ws.max_row + 2  # Leave one blank row for better spacing
        ws.cell(row=timestamp_row, column=1, value="Last Updated:")
        ws.cell(row=timestamp_row, column=1).font = Font(bold=True)
    else:
        # If timestamp row exists, ensure there's a blank row before it
        if timestamp_row > 2 and ws.cell(row=timestamp_row-1, column=1).value is not None:
            ws.insert_rows(timestamp_row, 1)
            timestamp_row += 1
    
    # Update the timestamp (clear any existing data in the row first)
    for col in range(2, ws.max_column + 1):
        ws.cell(row=timestamp_row, column=col).value = None
    ws.cell(row=timestamp_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    wb.save(xlsx_file)

    # Show final progress message for Excel report generation
    try:
        complete_phase_progress("Reports generated successfully")
    except Exception as e:
        print(f"[ProgressBar] Error updating progress bar: {e}")

    # At the end of the script (after all processing is done)
    if working_progress_bar:
        working_progress_bar.completed_phases = working_progress_bar.total_phases
    # Don't call finish_progress_bar here as it's already called in the main script

    # Wait for progress bar to close before showing notification
    import time
    time.sleep(3)  # Wait for the webpage to close

    # Show completion message
    print("✅ Risk assessment completed successfully!")
    print(f"📊 Results saved to:")
    print(f"   • {xlsx_file}")
    print(f"   • {json_file}")
    
    logs_dir = os.path.join(PROJECT_ROOT, "logs")
    if os.path.exists(logs_dir):
        print(f"   • {os.path.join(logs_dir, 'risk_assessment_summary.txt')}")
    
    print(f"📝 Check the 'data/' directory for detailed reports")
    if os.path.exists(logs_dir):
        print(f"📝 Check the 'logs/' directory for detailed logs")
    
    print("\n⚠️ DISCLAIMER:")
    print("This tool provides automated risk assessment based on available data and should be used")
    print("as part of a comprehensive due diligence process. Results are for informational purposes")
    print("only and do not constitute financial advice. Always conduct your own research and consult")
    print("with qualified professionals before making investment decisions.")
    print("\nMarket data provided by CoinGecko (https://www.coingecko.com)")

    # Show completion notification (only console output, no dialog)
    print("✅ Excel report updated successfully!")

    # Remove the flag file so the next run will wait again
    flag_path = os.path.join(DATA_DIR, "risk_report.done")
    if os.path.exists(flag_path):
        os.remove(flag_path)

if __name__ == "__main__":
    main() 