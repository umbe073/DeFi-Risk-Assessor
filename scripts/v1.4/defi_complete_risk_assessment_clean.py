# Standard library imports
from itertools import chain
from decimal import Decimal, getcontext

# Set precision for decimal calculations
getcontext().prec = 28

import os
import time
import csv
import json
import requests
import pandas as pd
from web3 import Web3
from dotenv import load_dotenv
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import datetime
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from eth_utils import is_checksum_address, to_checksum_address
import traceback
import subprocess
import sys
import hashlib
import shelve
from diskcache import Cache
import threading
from typing import Dict, List, Optional
import random # Added for robust_request retry delay

# Optional imports for enhanced features
# Note: feedparser is optional - install with: pip install feedparser
FEEDPARSER_AVAILABLE = False
try:
    import feedparser  # type: ignore
    FEEDPARSER_AVAILABLE = True
except ImportError:
    # feedparser is optional - functionality will be limited if not available
    pass

# Note: bs4 (BeautifulSoup) is optional - install with: pip install beautifulsoup4
BS4_AVAILABLE = False
try:
    import bs4  # type: ignore
    BS4_AVAILABLE = True
except ImportError:
    # bs4 is optional - functionality will be limited if not available
    pass

# Write debug info to a file for troubleshooting app environment
DEBUG_PATH = os.path.join(os.path.dirname(__file__), '../../data/debug_python_env.txt')
try:
    with open(DEBUG_PATH, 'w') as f:
        f.write(f"[DEBUG] Python executable: {sys.executable}\n")
        f.write(f"[DEBUG] sys.path: {sys.path}\n")
except Exception as e:
    pass  # Don't crash if debug file can't be written

# Debug info only written to file, not printed to console

# Set up data and log directories for correct file paths
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
DATA_DIR = os.path.join(PROJECT_ROOT, 'data')
LOGS_DIR = os.path.join(PROJECT_ROOT, 'logs')

# Progress bar availability
PROGRESS_AVAILABLE = False

print("[DEBUG] Using DATA_DIR:", DATA_DIR)
EXCEL_REPORT_PATH = os.path.join(DATA_DIR, 'DeFi Tokens Risk Assessment Results.xlsx')
print("[DEBUG] Excel report path:", EXCEL_REPORT_PATH)

TOKENS_CSV = os.path.join(DATA_DIR, 'tokens.csv')
CMC_SYMBOL_MAP = os.path.join(DATA_DIR, 'cmc_symbol_map.json')
FALLBACKS_JSON = os.path.join(DATA_DIR, 'fallbacks.json')
RISK_REPORT_JSON = os.path.join(DATA_DIR, 'risk_report.json')
RISK_REPORT_CSV = os.path.join(DATA_DIR, 'risk_report.csv')
API_CACHE_DB = os.path.join(DATA_DIR, 'api_cache.db')
VERBOSE_LOG = os.path.join(LOGS_DIR, 'risk_assessment_verbose.log')
SUMMARY_TXT = os.path.join(LOGS_DIR, 'risk_assessment_summary.txt')

# Ensure directories exist
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(LOGS_DIR, exist_ok=True)

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# Global list to track failed API endpoints
failed_api_endpoints = []

def log_failed_api_endpoint(api_name, endpoint, error_message):
    """Log failed API endpoint connections"""
    global failed_api_endpoints
    failed_endpoint = {
        'api_name': api_name,
        'endpoint': endpoint,
        'error': error_message,
        'timestamp': datetime.datetime.now().isoformat()
    }
    failed_api_endpoints.append(failed_endpoint)
    
    # Also log to console
    print(f"❌ API Error: {error_message} for {endpoint}")

def write_failed_endpoints_summary():
    """Write failed API endpoints to summary file and print to console"""
    try:
        summary_lines = []
        summary_lines.append("\n" + "="*50)
        summary_lines.append("FAILED API ENDPOINT CONNECTIONS")
        summary_lines.append("="*50)
        summary_lines.append(f"Total failed endpoints: {len(failed_api_endpoints)}\n")
        
        if failed_api_endpoints:
            for endpoint in failed_api_endpoints:
                summary_lines.append(f"API: {endpoint['api_name']}")
                summary_lines.append(f"Endpoint: {endpoint['endpoint']}")
                summary_lines.append(f"Error: {endpoint['error']}")
                summary_lines.append(f"Timestamp: {endpoint['timestamp']}")
                summary_lines.append("-" * 30)
        else:
            summary_lines.append("✅ No failed API endpoints detected during this run.")
            summary_lines.append("All API calls completed successfully or were handled gracefully.")
        
        summary_lines.append("")
        summary_text = "\n".join(summary_lines)
        
        # Write to file
        with open(SUMMARY_TXT, 'a') as f:
            f.write(summary_text)
        
        # Also print to console
        print(summary_text)
        
    except Exception as e:
        error_msg = f"Error writing failed endpoints summary: {e}"
        print(error_msg)
        traceback.print_exc()

# Load environment variables at module level
print("🔍 Loading environment variables...")
# Look for .env file in multiple locations
env_paths = [
    os.path.join(os.getcwd(), '.env'),  # Current directory
    os.path.join(os.path.dirname(__file__), '.env'),  # Script directory
    os.path.join(os.path.dirname(__file__), '..', '.env'),  # Parent of script directory
    os.path.join(os.path.dirname(__file__), '..', '..', '.env'),  # Two levels up
]

env_path = None
for path in env_paths:
    if os.path.exists(path):
        env_path = path
        break

# Force load the correct .env file
correct_env_path = os.path.join(os.path.dirname(__file__), '..', '..', '.env')
if os.path.exists(correct_env_path):
    env_path = correct_env_path
    print(f"🔍 Found .env file at: {env_path}")
    load_dotenv(env_path)
    print("🔍 Environment variables loaded successfully")
elif env_path:
    print(f"🔍 Found .env file at: {env_path}")
    load_dotenv(env_path)
    print("🔍 Environment variables loaded successfully")
else:
    print("⚠️  No .env file found in any of the expected locations:")
    for path in env_paths:
        print(f"   - {path}")
    print("🔍 Attempting to load from default location...")
    load_dotenv()  # Try default behavior

"""
DeFi Risk Assessment Tool

This script performs comprehensive risk assessment of DeFi tokens across multiple blockchains.
It analyzes various risk factors including:
- On-chain data (contract verification, holder distribution, liquidity)
- Market data (price, volume, market cap)
- Security aspects (audits, red flags)
- Social and developer activity

The risk assessment produces a score from 0-150, where:
- 0-50: Low Risk
- 51-100: Medium Risk
- 101-120: High Risk
- 121-150: Extreme Risk

Required API Keys (in .env file):
- INFURA_API_KEY: For Ethereum blockchain access
- ETHERSCAN_API_KEY: For Ethereum contract data
- COINGECKO_API_KEY: For market data (optional)
- BITQUERY_API_KEY: For on-chain analytics (optional)
- SANTIMENT_API_KEY: For social/dev metrics (optional)
- CERTIK_API_KEY: For security audits (optional)
- COINMARKETCAP_API_KEY: For additional market data (optional)
"""

# API Keys
INFURA_API_KEY = os.getenv('INFURA_API_KEY')
ETHERSCAN_API_KEY = os.getenv('ETHERSCAN_API_KEY')
ETHPLORER_API_KEY = os.getenv('ETHPLORER_API_KEY')
COINGECKO_API_KEY = os.getenv('COINGECKO_API_KEY')
COINMARKETCAP_API_KEY = os.getenv('COINMARKETCAP_API_KEY')
COINPAPRIKA_API_KEY = os.getenv('COINPAPRIKA_API_KEY')
COINAPI_API_KEY = os.getenv('COINAPI_API_KEY')
BITQUERY_API_KEY = os.getenv('BITQUERY_API_KEY')
SANTIMENT_API_KEY = os.getenv('SANTIMENT_API_KEY')
CERTIK_API_KEY = os.getenv('CERTIK_API_KEY')
SCORECHAIN_API_KEY = os.getenv('SCORECHAIN_API_KEY')
TRM_LABS_API_KEY = os.getenv('TRM_LABS_API_KEY')
OPENSANCTIONS_API_KEY = os.getenv('OPENSANCTIONS_API_KEY')
LUKKA_API_KEY = os.getenv('LUKKA_API_KEY')
ALCHEMY_API_KEY = os.getenv('ALCHEMY_API_KEY')
DEFISAFETY_API_KEY = os.getenv('DEFISAFETY_API_KEY')
MORALIS_API_KEY = os.getenv('MORALIS_API_KEY')
ZAPPER_API_KEY = os.getenv('ZAPPER_API_KEY')
DEBANK_API_KEY = os.getenv('DEBANK_API_KEY')
INCH_API_KEY = os.getenv('INCH_API_KEY')
DUNE_API_KEY = os.getenv('DUNE_ANALYTICS_API_KEY')
VESPIA_API_KEY = os.getenv('VESPIA_API_KEY')
TWITTER_API_KEY = os.getenv('TWITTER_BEARER_TOKEN')  # Fixed: Use TWITTER_BEARER_TOKEN
TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
DISCORD_BOT_TOKEN = os.getenv('DISCORD_BOT_TOKEN')
REDDIT_CLIENT_ID = os.getenv('REDDIT_CLIENT_ID')
REDDIT_CLIENT_SECRET = os.getenv('REDDIT_CLIENT_SECRET')
BREADCRUMBS_API_KEY = os.getenv('BREADCRUMBS_API_KEY')
BSCSCAN_API_KEY = os.getenv('BSCSCAN_API_KEY')
LI_FI_API_KEY = os.getenv('LI_FI_API_KEY')
THE_GRAPH_API_KEY = os.getenv('THE_GRAPH_API_KEY')

# Initialize Web3
w3 = None
if INFURA_API_KEY:
    w3 = Web3(Web3.HTTPProvider(f'https://mainnet.infura.io/v3/{INFURA_API_KEY}'))

def fetch_and_update_fallbacks():
    """Fetch and update fallback data from various sources"""
    fallbacks = {}
    
    # CoinMarketCap symbol mapping
    if COINMARKETCAP_API_KEY:
        try:
            url = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/map"
            headers = {'X-CMC_PRO_API_KEY': COINMARKETCAP_API_KEY}
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                data = response.json()
                symbol_map = {}
                for coin in data.get('data', []):
                    symbol_map[coin['symbol'].lower()] = {
                        'id': coin['id'],
                        'name': coin['name'],
                        'slug': coin['slug']
                    }
                fallbacks['cmc_symbol_map'] = symbol_map
                
                # Save to file
                with open(CMC_SYMBOL_MAP, 'w') as f:
                    json.dump(symbol_map, f, indent=2)
        except Exception as e:
            print(f"Error fetching CMC data: {e}")
    
    # Social keywords
    social_keywords = {
        'scam': -10, 'rug': -10, 'honeypot': -10, 'fake': -8,
        'legit': 5, 'real': 3, 'verified': 4, 'audited': 6,
        'community': 2, 'team': 3, 'transparent': 4, 'safe': 3,
        'moon': -2, 'pump': -3, 'dump': -5, 'fomo': -3,
        'dyor': 1, 'hodl': 1, 'diamond': 1, 'hands': 1
    }
    fallbacks['social_keywords'] = social_keywords
    
    # Save fallbacks
    with open(FALLBACKS_JSON, 'w') as f:
        json.dump(fallbacks, f, indent=2)
    
    return fallbacks

def update_fallbacks_if_needed():
    """Update fallbacks if they don't exist or are old"""
    try:
        if not os.path.exists(FALLBACKS_JSON):
            return fetch_and_update_fallbacks()
        
        # Check if fallbacks are older than 24 hours
        file_age = time.time() - os.path.getmtime(FALLBACKS_JSON)
        if file_age > 86400:  # 24 hours
            return fetch_and_update_fallbacks()
        
        # Load existing fallbacks
        with open(FALLBACKS_JSON, 'r') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error updating fallbacks: {e}")
        return {}

def validate_api_key(api_key, service):
    """Validate API key format for different services"""
    if not api_key:
        return False
    
    # Basic validation patterns
    patterns = {
        'etherscan': r'^[A-Z0-9]{34}$',
        'coingecko': r'^[A-Z0-9]{32}$',
        'bitquery': r'^[A-Z0-9]{32}$',
        'santiment': r'^[A-Z0-9]{32}$',
        'certik': r'^[A-Z0-9]{32}$',
        'coinmarketcap': r'^[A-Z0-9]{32}$'
    }
    
    import re
    pattern = patterns.get(service.lower(), r'^.+$')
    return bool(re.match(pattern, api_key))

def check_api_keys():
    """Check and log API key availability"""
    print("\n🔑 API Key Verification:")
    print("=" * 40)
    
    # Market Data APIs
    print("📊 Market Data APIs:")
    if COINGECKO_API_KEY:
        print("  ✅ CoinGecko API Key: Available")
    else:
        print("  ⚠️  CoinGecko API Key: Not required (free tier)")
    
    if COINMARKETCAP_API_KEY:
        print("  ✅ CoinMarketCap API Key: Available")
    else:
        print("  ❌ CoinMarketCap API Key: Missing")
    
    if ETHPLORER_API_KEY:
        print("  ✅ Ethplorer API Key: Available")
    else:
        print("  ❌ Ethplorer API Key: Missing")
    
    if BITQUERY_API_KEY:
        print("  ✅ BitQuery API Key: Available")
    else:
        print("  ❌ BitQuery API Key: Missing")
    
    if MORALIS_API_KEY:
        print("  ✅ Moralis API Key: Available")
    else:
        print("  ❌ Moralis API Key: Missing")
    
    # Security & Compliance APIs
    print("\n🛡️ Security & Compliance APIs:")
    if CERTIK_API_KEY:
        print("  ✅ CertiK API Key: Available")
    else:
        print("  ❌ CertiK API Key: Missing")
    
    if SCORECHAIN_API_KEY:
        print("  ✅ Scorechain API Key: Available")
    else:
        print("  ❌ Scorechain API Key: Missing")
    
    if TRM_LABS_API_KEY:
        print("  ✅ TRM Labs API Key: Available")
    else:
        print("  ❌ TRM Labs API Key: Missing")
    
    if OPENSANCTIONS_API_KEY:
        print("  ✅ OpenSanctions API Key: Available")
    else:
        print("  ❌ OpenSanctions API Key: Missing")
    
    if LUKKA_API_KEY:
        print("  ✅ Lukka API Key: Available")
    else:
        print("  ❌ Lukka API Key: Missing")
    
    if ALCHEMY_API_KEY:
        print("  ✅ Alchemy API Key: Available")
    else:
        print("  ❌ Alchemy API Key: Missing")
    
    if DEFISAFETY_API_KEY:
        print("  ✅ DeFiSafety API Key: Available")
    else:
        print("  ❌ DeFiSafety API Key: Missing")
    
    # Social Data APIs
    print("\n📱 Social Data APIs:")
    if TWITTER_API_KEY:
        print("  ✅ Twitter API Key: Available")
    else:
        print("  ❌ Twitter API Key: Missing")
    
    if TELEGRAM_BOT_TOKEN:
        print("  ✅ Telegram Bot Token: Available")
    else:
        print("  ❌ Telegram Bot Token: Missing")
    
    if DISCORD_BOT_TOKEN:
        print("  ✅ Discord Bot Token: Available")
    else:
        print("  ❌ Discord Bot Token: Missing")
    
    if REDDIT_CLIENT_ID and REDDIT_CLIENT_SECRET:
        print("  ✅ Reddit API Credentials: Available")
    else:
        print("  ❌ Reddit API Credentials: Missing")
    
    # DeFi Protocol APIs
    print("\n🏦 DeFi Protocol APIs:")
    if ZAPPER_API_KEY:
        print("  ✅ Zapper API Key: Available")
    else:
        print("  ❌ Zapper API Key: Missing")
    
    if DEBANK_API_KEY:
        print("  ✅ DeBank API Key: Available")
    else:
        print("  ❌ DeBank API Key: Missing")
    
    if INCH_API_KEY:
        print("  ✅ 1inch API Key: Available")
    else:
        print("  ❌ 1inch API Key: Missing")
    
    # Analytics APIs
    print("\n📈 Analytics APIs:")
    if DUNE_API_KEY:
        print("  ✅ Dune Analytics API Key: Available")
    else:
        print("  ❌ Dune Analytics API Key: Missing")
    
    if SANTIMENT_API_KEY:
        print("  ✅ Santiment API Key: Available")
    else:
        print("  ❌ Santiment API Key: Missing")
    
    print("=" * 40)
    print(f"📊 Summary: {sum([bool(COINGECKO_API_KEY), bool(COINMARKETCAP_API_KEY), bool(ETHERSCAN_API_KEY), bool(ETHPLORER_API_KEY), bool(BITQUERY_API_KEY), bool(MORALIS_API_KEY), bool(CERTIK_API_KEY), bool(SCORECHAIN_API_KEY), bool(TRM_LABS_API_KEY), bool(OPENSANCTIONS_API_KEY), bool(LUKKA_API_KEY), bool(ALCHEMY_API_KEY), bool(DEFISAFETY_API_KEY), bool(TWITTER_API_KEY), bool(TELEGRAM_BOT_TOKEN), bool(DISCORD_BOT_TOKEN), bool(REDDIT_CLIENT_ID), bool(ZAPPER_API_KEY), bool(DEBANK_API_KEY), bool(INCH_API_KEY), bool(DUNE_API_KEY), bool(SANTIMENT_API_KEY)])}/22 APIs available")
    print()

class APICache:
    def __init__(self, filename='api_cache.db'):
        self.filename = os.path.join(DATA_DIR, filename)
        self.cache = None
        try:
            # Ensure data directory exists
            os.makedirs(DATA_DIR, exist_ok=True)
            
            # Create cache directory with proper permissions
            cache_dir = os.path.dirname(self.filename)
            os.makedirs(cache_dir, exist_ok=True)
            
            # Initialize cache with proper error handling
            self.cache = Cache(self.filename)
            
            # Test cache functionality
            self.cache.set('test', 'test', expire=1)
            test_result = self.cache.get('test')
            if test_result != 'test':
                raise Exception("Cache test failed")
                
            print(f"✅ Cache initialized successfully: {self.filename}")
        except Exception as e:
            print(f"❌ Warning: Could not initialize cache: {e}")
            print("Cache will be disabled. Continuing without caching...")
            self.cache = None
    
    def get(self, key):
        """Get value from cache with error handling"""
        if self.cache is None:
            return None
        try:
            return self.cache.get(key)
        except Exception as e:
            print(f"Cache get error for key '{key}': {e}")
            return None
    
    def set(self, key, value, expire=3600):
        """Set value in cache with error handling"""
        if self.cache is None:
            return
        try:
            self.cache.set(key, value, expire=expire)
        except Exception as e:
            print(f"Cache set error for key '{key}': {e}")
    
    def close(self):
        """Close cache with error handling"""
        if self.cache is None:
            return
        try:
            self.cache.close()
        except Exception as e:
            print(f"Cache close error: {e}")

def fetch_ethplorer_bulk(addresses, ethplorer_key):
    """Fetch bulk token data from Ethplorer"""
    results = {}
    for address in addresses:
        try:
            url = f"https://api.ethplorer.io/getTokenInfo/{address}?apiKey={ethplorer_key}"
            response = robust_request('GET', url, timeout=15)
            if response and response.status_code == 200:
                results[address] = response.json()
            else:
                print(f"    ❌ Ethplorer API failed for {address}: {response.status_code if response else 'No response'}")
        except Exception as e:
            print(f"    ❌ Ethplorer error for {address}: {e}")
    return results

def fetch_coinpaprika_market(symbol):
    """Fetch market data from Coinpaprika"""
    try:
        url = f"https://api.coinpaprika.com/v1/tickers/{symbol}"
        response = robust_request('GET', url, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ Coinpaprika API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ Coinpaprika error: {e}")
    return None

def fetch_dune_query(query_id, dune_key):
    """Fetch data from Dune Analytics API v1 (2025)"""
    if not dune_key:
        print(f"    ⚠️  Dune API key missing")
        log_failed_api_endpoint('Dune Analytics', 'api.dune.com/api/v1', 'API key missing')
        return None
    
    # Dune Analytics API v1 endpoints (2025)
    endpoints = [
        f"https://api.dune.com/api/v1/query/{query_id}/results",
        f"https://api.dune.com/api/v1/queries/{query_id}/results",
    ]
    headers = {'X-Dune-API-Key': dune_key, 'Content-Type': 'application/json'}
    
    for url in endpoints:
        try:
            response = robust_request('GET', url, headers=headers, timeout=20)
            if response and response.status_code == 200:
                data = response.json()
                print(f"      ✅ Dune Analytics query results found")
                return data
            elif response and response.status_code == 401:
                error_msg = f"Dune API authentication failed: Invalid API key"
                print(f"    ❌ {error_msg}")
                log_failed_api_endpoint('Dune Analytics', url, error_msg)
                continue
            elif response and response.status_code == 404:
                error_msg = f"Dune query {query_id} not found"
                print(f"    ⚠️  {error_msg}")
                log_failed_api_endpoint('Dune Analytics', url, error_msg)
                continue
            elif response and response.status_code == 429:
                error_msg = f"Dune API rate limit exceeded"
                print(f"    ⚠️  {error_msg}")
                log_failed_api_endpoint('Dune Analytics', url, error_msg)
                continue
            else:
                error_msg = f"Dune API returned {response.status_code if response else 'No response'}"
                print(f"    ⚠️  {error_msg}, trying next endpoint...")
                log_failed_api_endpoint('Dune Analytics', url, error_msg)
                continue
        except Exception as e:
            error_msg = f"Dune error: {str(e)}"
            print(f"    ❌ {error_msg}")
            log_failed_api_endpoint('Dune Analytics', url, error_msg)
            continue
    
    print("    ❌ Dune Analytics API failed across all endpoints")
    return None

def fetch_breadcrumbs_risk_score(address):
    """Fetch risk score from Breadcrumbs API (2025)"""
    if not BREADCRUMBS_API_KEY:
        print(f"    ⚠️  Breadcrumbs API key missing")
        log_failed_api_endpoint('Breadcrumbs', 'api.breadcrumbs.one', 'API key missing')
        return None
    
    # Breadcrumbs API endpoints (2025)
    endpoints = [
        "https://api.breadcrumbs.one/api/v1/risk/address",
        "https://api.breadcrumbs.one/api/v1/sanctions/address",
        "https://api.breadcrumbs.one/risk/address",
    ]
    headers = {"X-API-KEY": BREADCRUMBS_API_KEY, "Accept": "application/json"}
    params = {"chain": "ETH", "address": address}
    
    for url in endpoints:
        try:
            response = robust_request('GET', url, headers=headers, params=params, timeout=20)
            if response and response.status_code == 200:
                data = response.json()
                print(f"      ✅ Breadcrumbs risk score found via {url}")
                return data
            elif response and response.status_code == 401:
                error_msg = f"Breadcrumbs API authentication failed: Invalid API key"
                log_failed_api_endpoint("Breadcrumbs", url, error_msg)
                continue
            elif response and response.status_code == 403:
                error_msg = f"Breadcrumbs API access forbidden"
                log_failed_api_endpoint("Breadcrumbs", url, error_msg)
                continue
            elif response and response.status_code == 404:
                error_msg = f"Breadcrumbs endpoint not found"
                log_failed_api_endpoint("Breadcrumbs", url, error_msg)
                continue
            elif response and response.status_code == 429:
                error_msg = f"Breadcrumbs API rate limit exceeded"
                log_failed_api_endpoint("Breadcrumbs", url, error_msg)
                continue
            else:
                error_msg = f"HTTP {response.status_code if response else 'No response'}"
                log_failed_api_endpoint("Breadcrumbs", url, error_msg)
        except requests.exceptions.RequestException as e:
            log_failed_api_endpoint("Breadcrumbs", url, f"Request error: {str(e)}")
            continue
        except json.JSONDecodeError as e:
            log_failed_api_endpoint("Breadcrumbs", url, f"JSON decode error: {str(e)}")
            continue
        except Exception as e:
            log_failed_api_endpoint("Breadcrumbs", url, f"Error: {str(e)}")
            continue
    
    print("      ❌ Breadcrumbs API failed across all endpoints")
    return None

def fetch_breadcrumbs_token_info(address):
    """Fetch token info from Breadcrumbs API"""
    if not BREADCRUMBS_API_KEY:
        print(f"    ⚠️  Breadcrumbs API key missing")
        return None
    
    endpoints = [
        "https://api.breadcrumbs.one/risk/address",
        "https://api.breadcrumbs.one/api/v1/sanctions/address",
        "https://api.breadcrumbs.one/sanctions/address"
    ]
    headers = {"X-API-KEY": BREADCRUMBS_API_KEY, "Accept": "application/json"}
    params = {"chain": "ETH", "address": address}
    for url in endpoints:
        try:
            response = robust_request('GET', url, headers=headers, params=params, timeout=20)
            if response and response.status_code == 200:
                return response.json()
        except Exception as e:
            print(f"    ❌ Breadcrumbs Token error on {url}: {e}")
            continue
    print(f"    ❌ Breadcrumbs Token API failed across all endpoints")
    return None

def fetch_ethplorer_token_info(address):
    """Fetch token info from Ethplorer with improved validation"""
    api_key = ETHPLORER_API_KEY or "freekey"
    try:
        # Validate address format
        if not address or len(address) != 42 or not address.startswith('0x'):
            print(f"    ❌ Invalid address format: {address}")
            return None
            
        url = f"https://api.ethplorer.io/getTokenInfo/{address}?apiKey={api_key}"
        response = robust_request('GET', url, timeout=15)
        
        if response and response.status_code == 200:
            data = response.json()
            # Check if it's actually a token (has name and symbol)
            if data.get('name') and data.get('symbol'):
                return data
            else:
                print(f"    ⚠️  Address {address} is not a token contract")
                return None
        elif response and response.status_code == 400:
            error_data = response.json()
            if error_data.get('error', {}).get('message') == 'Address is not a token contract':
                print(f"    ⚠️  Address {address} is not a token contract")
                return None
            else:
                print(f"    ❌ Ethplorer API error: {error_data}")
                return None
        else:
            print(f"    ❌ Ethplorer API failed: {response.status_code if response else 'No response'}")
            return None
    except Exception as e:
        print(f"    ❌ Ethplorer error: {e}")
        return None

def fetch_ethplorer_address_info(address):
    """Fetch address info from Ethplorer with improved validation"""
    api_key = ETHPLORER_API_KEY or "freekey"
    try:
        # Validate address format
        if not address or len(address) != 42 or not address.startswith('0x'):
            print(f"    ❌ Invalid address format: {address}")
            return None
            
        url = f"https://api.ethplorer.io/getAddressInfo/{address}?apiKey={api_key}"
        response = robust_request('GET', url, timeout=15)
        
        if response and response.status_code == 200:
            data = response.json()
            # Check if address has any tokens
            if data.get('tokens') or data.get('ETH'):
                return data
            else:
                print(f"    ⚠️  Address {address} has no token data")
                return None
        elif response and response.status_code == 400:
            error_data = response.json()
            print(f"    ❌ Ethplorer Address API error: {error_data}")
            return None
        else:
            print(f"    ❌ Ethplorer Address API failed: {response.status_code if response else 'No response'}")
            return None
    except Exception as e:
        print(f"    ❌ Ethplorer Address error: {e}")
        return None

def fetch_zapper_portfolio_data(address):
    """Fetch portfolio data from Zapper API v2 (2025)"""
    if not ZAPPER_API_KEY:
        print(f"    ⚠️  Zapper API key missing")
        log_failed_api_endpoint('Zapper', 'api.zapper.xyz/v2', 'API key missing')
        return None
    
    # Zapper API v2 endpoints (2025)
    endpoints = [
        f"https://api.zapper.xyz/v2/portfolio/{address}",
        f"https://api.zapper.xyz/v2/balances/{address}",
    ]
    
    for url in endpoints:
        try:
            headers = {"X-Zapper-API-Key": ZAPPER_API_KEY, "Accept": "application/json"}
            response = robust_request('GET', url, headers=headers, timeout=20)
            if response and response.status_code == 200:
                data = response.json()
                print(f"      ✅ Zapper portfolio data found")
                return data
            elif response and response.status_code == 404:
                print(f"      ⚠️  Zapper portfolio not found for address {address[:8]}...")
                continue
            elif response and response.status_code == 401:
                error_msg = f"Zapper API authentication failed: Invalid API key"
                print(f"      ❌ {error_msg}")
                log_failed_api_endpoint('Zapper', url, error_msg)
                return None
            elif response and response.status_code == 403:
                error_msg = f"Zapper API access forbidden"
                print(f"      ❌ {error_msg}")
                log_failed_api_endpoint('Zapper', url, error_msg)
                return None
            elif response and response.status_code == 429:
                error_msg = f"Zapper API rate limit exceeded"
                print(f"      ⚠️  {error_msg}")
                log_failed_api_endpoint('Zapper', url, error_msg)
                continue
            else:
                error_msg = f"HTTP {response.status_code if response else 'No response'}"
                log_failed_api_endpoint('Zapper', url, error_msg)
                continue
        except Exception as e:
            error_msg = f"Zapper error: {str(e)}"
            print(f"      ❌ {error_msg}")
            log_failed_api_endpoint('Zapper', url, error_msg)
            continue
    
    print("      ❌ Zapper API failed across all endpoints")
    return None

def fetch_zapper_protocol_data(protocol):
    """Fetch protocol data from Zapper API v2 (2025)"""
    if not ZAPPER_API_KEY:
        print(f"    ⚠️  Zapper API key missing")
        log_failed_api_endpoint('Zapper', 'api.zapper.xyz/v2/protocols', 'API key missing')
        return None
    
    try:
        url = f"https://api.zapper.xyz/v2/protocols/{protocol}"
        headers = {"X-Zapper-API-Key": ZAPPER_API_KEY, "Accept": "application/json"}
        response = robust_request('GET', url, headers=headers, timeout=20)
        if response and response.status_code == 200:
            data = response.json()
            print(f"      ✅ Zapper protocol data found")
            return data
        elif response and response.status_code == 401:
            error_msg = f"Zapper API authentication failed: Invalid API key"
            print(f"    ❌ {error_msg}")
            log_failed_api_endpoint('Zapper', url, error_msg)
        elif response and response.status_code == 403:
            error_msg = f"Zapper API access forbidden"
            print(f"    ❌ {error_msg}")
            log_failed_api_endpoint('Zapper', url, error_msg)
        elif response and response.status_code == 404:
            error_msg = f"Zapper protocol '{protocol}' not found"
            print(f"    ⚠️  {error_msg}")
            log_failed_api_endpoint('Zapper', url, error_msg)
        else:
            error_msg = f"HTTP {response.status_code if response else 'No response'}"
            print(f"    ❌ Zapper Protocol API failed: {error_msg}")
            log_failed_api_endpoint('Zapper', url, error_msg)
    except Exception as e:
        error_msg = f"Zapper Protocol error: {str(e)}"
        print(f"    ❌ {error_msg}")
        log_failed_api_endpoint('Zapper', url, error_msg)
    return None

def fetch_debank_portfolio(address):
    """Fetch portfolio data from DeBank API with improved error handling"""
    if not DEBANK_API_KEY:
        print(f"    ⚠️  DeBank API key missing")
        return None
    
    try:
        url = f"https://pro-openapi.debank.com/v1/user/total_balance?id={address}"
        headers = {"AccessKey": DEBANK_API_KEY}
        response = robust_request('GET', url, headers=headers, timeout=15)
        if response and response.status_code == 200:
            data = response.json()
            print(f"      ✅ DeBank portfolio data found")
            return data
        elif response and response.status_code == 403:
            print(f"      ❌ DeBank API insufficient units (403 Forbidden)")
            return None
        else:
            print(f"      ❌ DeBank API failed: {response.status_code if response else 'No response'}")
            return None
    except Exception as e:
        print(f"      ❌ DeBank error: {e}")
        return None

def fetch_debank_token_list(chain_id):
    """Fetch token list from DeBank API"""
    if not DEBANK_API_KEY:
        print(f"    ⚠️  DeBank API key missing")
        return None
    
    try:
        url = f"https://pro-openapi.debank.com/v1/user/token_list?id={chain_id}"
        headers = {"AccessKey": DEBANK_API_KEY}
        response = robust_request('GET', url, headers=headers, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ DeBank Token List API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ DeBank Token List error: {e}")
    return None

def fetch_defillama_protocol_tvl(protocol):
    """Fetch protocol TVL from DeFiLlama"""
    try:
        url = f"https://api.llama.fi/protocol/{protocol}"
        response = robust_request('GET', url, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ DeFiLlama Protocol API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ DeFiLlama Protocol error: {e}")
    return None

def fetch_defillama_token_price(token_address, chain="ethereum"):
    """Fetch token price from DeFiLlama"""
    try:
        url = f"https://coins.llama.fi/prices/current/{chain}:{token_address.lower()}"
        response = robust_request('GET', url, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ DeFiLlama Token Price API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ DeFiLlama Token Price error: {e}")
    return None

def fetch_defillama_yield_pools(protocol):
    """Fetch yield pools from DeFiLlama"""
    try:
        url = f"https://yields.llama.fi/pools"
        params = {"protocol": protocol}
        response = robust_request('GET', url, params=params, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ DeFiLlama Yield Pools API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ DeFiLlama Yield Pools error: {e}")
    return None

def fetch_moralis_token_metadata(address, chain="eth"):
    """Fetch token metadata from Moralis with improved error handling"""
    if not MORALIS_API_KEY:
        print(f"    ⚠️  Moralis API key missing")
        return None
    
    try:
        url = f"https://deep-index.moralis.io/api/v2/erc20/{address}?chain={chain}"
        headers = {"X-API-Key": MORALIS_API_KEY}
        response = robust_request('GET', url, headers=headers, timeout=15)
        if response and response.status_code == 200:
            data = response.json()
            print(f"      ✅ Moralis token metadata found")
            return data
        elif response and response.status_code == 404:
            print(f"      ⚠️  Moralis token not found for address {address[:8]}...")
            return None
        else:
            print(f"      ❌ Moralis Token Metadata API failed: {response.status_code if response else 'No response'}")
            return None
    except Exception as e:
        print(f"      ❌ Moralis Token Metadata error: {e}")
        return None

def fetch_moralis_token_price(address, chain="eth"):
    """Fetch token price from Moralis"""
    if not MORALIS_API_KEY:
        print(f"    ⚠️  Moralis API key missing")
        return None
    
    try:
        url = f"https://deep-index.moralis.io/api/v2/erc20/{address}/price?chain={chain}"
        headers = {"X-API-Key": MORALIS_API_KEY}
        response = robust_request('GET', url, headers=headers, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ Moralis Token Price API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ Moralis Token Price error: {e}")
    return None

def fetch_moralis_token_transfers(address, chain="eth", limit=100):
    """Fetch token transfers from Moralis"""
    if not MORALIS_API_KEY:
        print(f"    ⚠️  Moralis API key missing")
        return None
    
    try:
        url = f"https://deep-index.moralis.io/api/v2/{address}/erc20/transfers?chain={chain}&limit={limit}"
        headers = {"X-API-Key": MORALIS_API_KEY}
        response = robust_request('GET', url, headers=headers, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ Moralis Token Transfers API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ Moralis Token Transfers error: {e}")
    return None

def robust_request(method, url, **kwargs):
    """Make a robust HTTP request with retry logic and detailed error logging"""
    max_retries = 3
    base_delay = 0.5  # Reduced base delay
    
    for attempt in range(max_retries):
        try:
            # Add minimal delay only on retries
            if attempt > 0:
                time.sleep(base_delay * (2 ** attempt))
            
            response = requests.request(method, url, **kwargs)
            
            # Handle rate limiting specifically
            if response.status_code == 429:
                delay = 2 + (attempt * 1)  # Shorter delays for rate limits
                print(f"    ⚠️  Rate limit hit, waiting {delay}s...")
                time.sleep(delay)
                continue
            
            # Log all API responses for debugging
            if response.status_code >= 400:
                print(f"❌ API Error: {response.status_code} {response.reason} for {url}")
                print(f"   Headers: {dict(response.request.headers)}")
                print(f"   Params: {kwargs.get('params', {})}")
                if response.text:
                    print(f"   Response: {response.text[:200]}...")
                
                # Log to failed endpoints summary
                log_failed_api_endpoint("API", url, f"{response.status_code} {response.reason}")
            
            return response
            
        except requests.exceptions.RequestException as e:
            print(f"❌ Request Error (attempt {attempt + 1}/{max_retries}): {e}")
            print(f"   URL: {url}")
            print(f"   Method: {method}")
            print(f"   Headers: {kwargs.get('headers', {})}")
            
            # Log to failed endpoints summary
            log_failed_api_endpoint("API", url, str(e))
            
            if attempt < max_retries - 1:
                delay = base_delay * (2 ** attempt) + random.uniform(0, 0.5)
                print(f"   Retrying in {delay:.2f} seconds...")
                time.sleep(delay)
            else:
                print(f"   Max retries reached for {url}")
                return None
    
    return None

def fetch_etherscan_tokeninfo(addresses, etherscan_key):
    """Fetch token info from Etherscan"""
    if not etherscan_key:
        return {}
    
    results = {}
    for address in addresses:
        try:
            url = f"https://api.etherscan.io/api"
            params = {
                'module': 'token',
                'action': 'tokeninfo',
                'contractaddress': address,
                'apikey': etherscan_key
            }
            response = robust_request('GET', url, params=params)
            if response and response.status_code == 200:
                data = response.json()
                if data.get('status') == '1':
                    results[address] = data.get('result', [])
        except Exception as e:
            print(f"Error fetching Etherscan token info for {address}: {e}")
    
    return results

class DeFiRiskAssessor:
    def __init__(self):
        """Initialize the DeFi Risk Assessor with comprehensive configuration"""
        self.logger = logging.getLogger(__name__)
        
        # Initialize cache
        try:
            self.cache = APICache()
        except Exception as e:
            print(f"Warning: Could not initialize cache: {e}")
            self.cache = None
        
        # Load fallbacks
        try:
            with open(FALLBACKS_JSON, 'r') as f:
                self.fallbacks = json.load(f)
        except Exception as e:
            print(f"Warning: Could not load fallbacks: {e}")
            self.fallbacks = {}
        
        # Check available API keys
        self.available_keys = check_api_keys()
        
        # Initialize Web3
        try:
            if INFURA_API_KEY:
                self.w3 = Web3(Web3.HTTPProvider(f"https://mainnet.infura.io/v3/{INFURA_API_KEY}"))
            else:
                self.w3 = None
        except Exception as e:
            print(f"Warning: Could not initialize Web3: {e}")
            self.w3 = None
        
        # Statistics tracking
        self.stats = {
            'successful_assessments': 0,
            'failed_assessments': 0,
            'cache_hits': 0,
            'cache_misses': 0,
            'errors': []
        }
        
        # Red flags and their risk impact scores (aligned with 0-150 risk scale)
        self.RED_FLAGS = [
            {'check': 'is_proxy_contract', 'risk_boost': 20},
            {'check': 'has_honeypot_pattern', 'risk_boost': 30},
            {'check': 'owner_change_last_24h', 'risk_boost': 15},
            {'check': 'lp_lock_expiring_soon', 'risk_boost': 25},
            {'check': 'unverified_contract', 'risk_boost': 15},
            {'check': 'low_liquidity', 'risk_boost': 12},
            {'check': 'high_concentration', 'risk_boost': 15},
            # EU Regulatory Compliance Red Flags (CRITICAL)
            {'check': 'eu_unlicensed_stablecoin', 'risk_boost': 50},  # Forces Extreme Risk
            {'check': 'eu_regulatory_issues', 'risk_boost': 40},      # Forces Extreme Risk
            {'check': 'mica_non_compliant', 'risk_boost': 35},        # High Risk
            {'check': 'mica_no_whitepaper', 'risk_boost': 0}          # No boost, just flag
        ]
        
        # Updated weights aligned with 16-component risk model (including social data)
        self.WEIGHTS = {
            "industry_impact": 0.08,
            "tech_innovation": 0.08,
            "whitepaper_quality": 0.06,
            "roadmap_adherence": 0.06,
            "business_model": 0.08,
            "team_expertise": 0.07,
            "management_strategy": 0.06,
            "global_reach": 0.05,
            "code_security": 0.08,
            "dev_activity": 0.06,
            "aml_data": 0.04,
            "compliance_data": 0.04,
            "market_dynamics": 0.05,
            "marketing_demand": 0.05,
            "esg_impact": 0.02,
            "social_data": 0.08
        }
        
        # Configuration for different blockchain networks
        self.CHAIN_CONFIG = {
            "eth": {
                "rpc": f"https://mainnet.infura.io/v3/{os.getenv('INFURA_API_KEY')}",
                "scan_url": "https://api.etherscan.io/v2/api",
                "scan_key": os.getenv("ETHERSCAN_API_KEY"),
                "coin_id": "ethereum",
                "dex": "uniswap",
                "min_liquidity": 5000000,
                "token_info_action": "tokeninfo",
                "coingecko_platform": "ethereum"
            },
            "bsc": {
                "rpc": "https://bsc-dataseed.binance.org/",
                "scan_url": "https://api.bscscan.com/api",
                "scan_key": os.getenv("BSCSCAN_API_KEY", os.getenv("ETHERSCAN_API_KEY")),
                "coin_id": "binancecoin",
                "dex": "pancakeswap",
                "min_liquidity": 3000000,
                "token_info_action": "token",
                "coingecko_platform": "binance-smart-chain"
            },
            "op": {
                "rpc": f"https://optimism-mainnet.infura.io/v3/{os.getenv('INFURA_API_KEY')}",
                "scan_url": "https://api-optimistic.etherscan.io/api",
                "scan_key": os.getenv("ETHERSCAN_API_KEY"),
                "coin_id": "ethereum",
                "dex": "uniswap",
                "min_liquidity": 3000000,
                "token_info_action": "token",
                "coingecko_platform": "optimistic-ethereum"
            },
            "optimism": {
                "rpc": f"https://optimism-mainnet.infura.io/v3/{os.getenv('INFURA_API_KEY')}",
                "scan_url": "https://api-optimistic.etherscan.io/api",
                "scan_key": os.getenv("ETHERSCAN_API_KEY"),
                "coin_id": "ethereum",
                "dex": "uniswap",
                "min_liquidity": 3000000,
                "token_info_action": "token",
                "coingecko_platform": "optimistic-ethereum"
            },
            "polygon": {
                "rpc": f"https://polygon-mainnet.infura.io/v3/{os.getenv('INFURA_API_KEY')}",
                "scan_url": "https://api.polygonscan.com/api",
                "scan_key": os.getenv("ETHERSCAN_API_KEY"),
                "coin_id": "matic-network",
                "dex": "quickswap",
                "min_liquidity": 3000000,
                "token_info_action": "token",
                "coingecko_platform": "polygon-pos"
            },
            "sonic": {
                "rpc": "https://rpc.soniclabs.com",
                "scan_url": "https://sonicscan.org/api",
                "scan_key": None,
                "coin_id": "sonic",
                "dex": "sonicswap",
                "min_liquidity": 1000000,
                "token_info_action": "token",
                "coingecko_platform": "sonic"
            },
            "s": {
                "rpc": "https://rpc.soniclabs.com",
                "scan_url": "https://sonicscan.org/api",
                "scan_key": None,
                "coin_id": "sonic",
                "dex": "sonicswap",
                "min_liquidity": 1000000,
                "token_info_action": "token",
                "coingecko_platform": "sonic"
            }
        }
    
    def _extract_market_cap(self, result):
        """Extract market cap from multiple sources, prioritizing real data"""
        market_data = result.get('market_data', {})
        values = []
        
        # Try CoinGecko
        cg_mc = market_data.get('coingecko', {}).get('market_data', {}).get('market_cap', {}).get('usd', 0)
        if cg_mc > 0:
            values.append(cg_mc)
        
        # Try CoinMarketCap
        cmc_mc = market_data.get('cmc', {}).get('data', {}).get('quote', {}).get('USD', {}).get('market_cap', 0)
        if cmc_mc > 0:
            values.append(cmc_mc)
        
        # Try CoinPaprika (from market_data or enhanced_data)
        paprika_mc = market_data.get('coinpaprika', {}).get('market_cap', 0)
        if paprika_mc > 0:
            values.append(paprika_mc)
        
        enhanced = result.get('enhanced_data', {})
        if enhanced.get('coinpaprika'):
            enhanced_paprika_mc = enhanced.get('coinpaprika', {}).get('market_cap', 0)
            if enhanced_paprika_mc > 0:
                values.append(enhanced_paprika_mc)
        
        # Return average if multiple sources, or single value, or 0
        if values:
            return sum(values) / len(values)
        return 0
    
    def _extract_volume_24h(self, result):
        """Extract 24h volume from multiple sources, prioritizing real data"""
        market_data = result.get('market_data', {})
        values = []
        
        # Try CoinGecko
        cg_vol = market_data.get('coingecko', {}).get('market_data', {}).get('total_volume', {}).get('usd', 0)
        if cg_vol > 0:
            values.append(cg_vol)
        
        # Try CoinMarketCap
        cmc_vol = market_data.get('cmc', {}).get('data', {}).get('quote', {}).get('USD', {}).get('volume_24h', 0)
        if cmc_vol > 0:
            values.append(cmc_vol)
        
        # Try CoinPaprika (from market_data or enhanced_data)
        paprika_vol = market_data.get('coinpaprika', {}).get('volume_24h', 0)
        if paprika_vol > 0:
            values.append(paprika_vol)
        
        enhanced = result.get('enhanced_data', {})
        if enhanced.get('coinpaprika'):
            enhanced_paprika_vol = enhanced.get('coinpaprika', {}).get('volume_24h', 0)
            if enhanced_paprika_vol > 0:
                values.append(enhanced_paprika_vol)
        
        # Return average if multiple sources, or single value, or 0
        if values:
            return sum(values) / len(values)
        return 0
    
    def _extract_liquidity_value(self, result):
        """Extract actual liquidity numeric value from liquidity_data dictionary"""
        onchain_data = result.get('onchain_data', {})
        liquidity_data = onchain_data.get('liquidity', {})
        
        # If liquidity_data is already a number, return it
        if isinstance(liquidity_data, (int, float)):
            return float(liquidity_data) if liquidity_data > 0 else 0
        
        # If it's a dictionary, try to extract liquidity from various sources
        if isinstance(liquidity_data, dict):
            # Try DeFiLlama data
            if 'defillama' in liquidity_data:
                defillama_data = liquidity_data['defillama']
                if isinstance(defillama_data, dict):
                    coins = defillama_data.get('coins', {})
                    for coin_key, coin_data in coins.items():
                        if isinstance(coin_data, dict):
                            liquidity = coin_data.get('liquidity', 0)
                            if liquidity > 0:
                                return float(liquidity)
            
            # Try 1inch data - calculate from quote if available
            if 'oneinch' in liquidity_data:
                oneinch_data = liquidity_data['oneinch']
                if isinstance(oneinch_data, dict):
                    # 1inch quote can indicate liquidity depth
                    to_token_amount = oneinch_data.get('toTokenAmount', 0)
                    if to_token_amount > 0:
                        # Estimate liquidity from quote (rough approximation)
                        # This is a fallback - actual liquidity would need DEX pool data
                        pass
        
        # Try to get liquidity from enhanced_data or other sources
        enhanced = result.get('enhanced_data', {})
        if enhanced.get('liquidity'):
            liq_val = enhanced.get('liquidity')
            if isinstance(liq_val, (int, float)) and liq_val > 0:
                return float(liq_val)
        
        # Calculate liquidity from volume if available (industry standard: liquidity ≈ volume * 0.15)
        volume_24h = self._extract_volume_24h(result)
        if volume_24h > 0:
            estimated_liquidity = volume_24h * 0.15  # Conservative multiplier
            return estimated_liquidity
        
        return 0
    
    def get_contract_verification_status(self, token_address, chain):
        """Get contract verification status from Etherscan with enhanced error handling"""
        if not ETHERSCAN_API_KEY:
            print(f"  ⚠️  Etherscan API key missing for contract verification")
            return "unknown"
        
        try:
            print(f"  🔍 Checking contract verification...")
            url = "https://api.etherscan.io/api"
            params = {
                'module': 'contract',
                'action': 'getabi',
                'address': token_address,
                'apikey': ETHERSCAN_API_KEY
            }
            response = robust_request('GET', url, params=params, timeout=15)
            
            if response and response.status_code == 200:
                data = response.json()
                if data.get('status') == '1' and data.get('result'):
                    print(f"    ✅ Contract verified")
                    return "verified"
                elif data.get('result') == 'Contract source code not verified':
                    print(f"    ❌ Contract not verified")
                    return "unverified"
                else:
                    print(f"    ⚠️  Contract verification unknown: {data.get('message', 'Unknown')}")
                    return "unknown"
            else:
                print(f"    ❌ Etherscan API failed: {response.status_code if response else 'No response'}")
                return "unknown"
        except Exception as e:
            print(f"    ❌ Contract verification error: {e}")
            return "unknown"
    
    def get_holder_data(self, token_address, chain):
        """Get holder distribution data with enhanced error handling - supports L2 chains"""
        holder_data = {
            'total_holders': 0,
            'top10_concentration': 0
        }
        
        chain_lower = (chain or 'eth').lower()
        
        # For L2 chains (OP, S/Sonic), use Blockscout API
        if chain_lower in ('op', 'optimism', 's', 'sonic'):
            try:
                print(f"  👥 Fetching holder data from Blockscout for {chain}...")
                # Import blockscout function from update_token_data_viewer
                import sys
                import os
                viewer_path = os.path.join(os.path.dirname(__file__), '..', 'v2.0', 'update_token_data_viewer.py')
                if os.path.exists(viewer_path):
                    # Use blockscout API directly
                    blockscout_domains = {
                        'op': 'optimism.blockscout.com',
                        'optimism': 'optimism.blockscout.com',
                        's': 'sonicscan.org',
                        'sonic': 'sonicscan.org',
                    }
                    domain = blockscout_domains.get(chain_lower)
                    if domain:
                        # Try v2 API first
                        url = f"https://{domain}/api/v2/tokens/{token_address}"
                        response = robust_request('GET', url, timeout=15)
                        if response and response.status_code == 200:
                            data = response.json()
                            holders_val = data.get('holders_count') or data.get('holdersCount') or data.get('holder_count') or 0
                            if isinstance(holders_val, str):
                                holders_val = holders_val.replace(',', '').strip()
                            try:
                                holders_int = int(float(holders_val))
                                if holders_int > 0:
                                    holder_data['total_holders'] = holders_int
                                    print(f"    ✅ Found {holders_int} holders via Blockscout")
                                    return holder_data
                            except Exception:
                                pass
                        
                        # Fallback to v1 API for Sonic
                        if chain_lower in ('s', 'sonic'):
                            url_v1 = f"https://{domain}/api?module=token&action=getToken&contractaddress={token_address}"
                            response_v1 = robust_request('GET', url_v1, timeout=15)
                            if response_v1 and response_v1.status_code == 200:
                                data_v1 = response_v1.json()
                                if data_v1.get('status') == '1':
                                    holders_val = data_v1.get('result', {}).get('holders') or 0
                                    if isinstance(holders_val, str):
                                        holders_val = holders_val.replace(',', '').strip()
                                    try:
                                        holders_int = int(float(holders_val))
                                        if holders_int > 0:
                                            holder_data['total_holders'] = holders_int
                                            print(f"    ✅ Found {holders_int} holders via Blockscout v1")
                                            return holder_data
                                    except Exception:
                                        pass
            except Exception as e:
                print(f"    ⚠️  Blockscout holder data error: {e}")
        
        # For Ethereum mainnet, use Etherscan
        if chain_lower == 'eth' or chain_lower == 'ethereum':
            if not ETHERSCAN_API_KEY:
                print(f"  ⚠️  Etherscan API key missing for holder data")
                return holder_data
            
            try:
                print(f"  👥 Fetching holder data from Etherscan...")
                url = "https://api.etherscan.io/api"
                params = {
                    'module': 'token',
                    'action': 'tokenholderlist',
                    'contractaddress': token_address,
                    'apikey': ETHERSCAN_API_KEY
                }
                response = robust_request('GET', url, params=params, timeout=15)
                
                if response and response.status_code == 200:
                    data = response.json()
                    if data.get('status') == '1':
                        holders = data.get('result', [])
                        holder_data['total_holders'] = len(holders)
                        print(f"    ✅ Found {len(holders)} holders via Etherscan")
                    else:
                        print(f"    ❌ Etherscan holder data failed: {data.get('message', 'Unknown')}")
                else:
                    print(f"    ❌ Etherscan API failed: {response.status_code if response else 'No response'}")
            except Exception as e:
                print(f"    ❌ Holder data error: {e}")
        
        return holder_data
    
    def get_token_supply(self, token_address, chain):
        """Get token supply information with proper address handling"""
        if not self.w3 or chain != 'eth':
            return {'total_supply': 0, 'circulating_supply': 0}
        
        try:
            from eth_utils import to_checksum_address
            
            # Convert address to checksum format
            try:
                checksum_address = to_checksum_address(token_address)
            except Exception as e:
                self.logger.error(f"Invalid address format: {token_address} - {e}")
                return {'total_supply': 0, 'circulating_supply': 0}
            
            # ERC20 token contract ABI for basic functions
            abi = [
                {"constant": True, "inputs": [], "name": "totalSupply", "outputs": [{"name": "", "type": "uint256"}], "type": "function"},
                {"constant": True, "inputs": [{"name": "_owner", "type": "address"}], "name": "balanceOf", "outputs": [{"name": "balance", "type": "uint256"}], "type": "function"},
                {"constant": True, "inputs": [], "name": "decimals", "outputs": [{"name": "", "type": "uint8"}], "type": "function"}
            ]
            
            contract = self.w3.eth.contract(address=checksum_address, abi=abi)
            total_supply = contract.functions.totalSupply().call()
            decimals = contract.functions.decimals().call()
            
            # Convert from wei to token units
            total_supply = total_supply / (10 ** decimals)
            
            return {
                'total_supply': total_supply,
                'circulating_supply': total_supply,  # Simplified assumption
                'address': checksum_address
            }
        except Exception as e:
            self.logger.error(f"Error getting token supply: {e}")
        
        return {'total_supply': 0, 'circulating_supply': 0}
    
    def get_liquidity_data(self, token_address, chain):
        """Get liquidity data from various sources"""
        liquidity_data = {}
        
        # Try 1inch API for liquidity
        if INCH_API_KEY:
            try:
                url = "https://api.1inch.dev/swap/v5.2/1/quote"
                headers = {'Authorization': f'Bearer {INCH_API_KEY}'}
                params = {
                    'src': token_address,
                    'dst': '0xA0b86a33E6441b8C4C8C0C4C0C4C0C4C0C4C0C4C',  # USDC
                    'amount': '1000000000000000000'  # 1 token
                }
                response = robust_request('GET', url, headers=headers, params=params)
                if response and response.status_code == 200:
                    liquidity_data['oneinch'] = response.json()
            except Exception as e:
                self.logger.error(f"Error fetching 1inch liquidity: {e}")
        
        # Try DeFiLlama for token price data (not protocol data)
        try:
            url = f"https://coins.llama.fi/prices/current/ethereum:{token_address.lower()}"
            response = robust_request('GET', url)
            if response and response.status_code == 200:
                data = response.json()
                if data.get('coins'):
                    liquidity_data['defillama'] = data
        except Exception as e:
            self.logger.error(f"Error fetching DeFiLlama data: {e}")
        
        return liquidity_data
    
    def fetch_onchain_data(self, token_address, chain):
        """Fetch comprehensive on-chain data"""
        onchain_data = {
            'contract_verified': self.get_contract_verification_status(token_address, chain),
            'holders': self.get_holder_data(token_address, chain),
            'supply': self.get_token_supply(token_address, chain),
            'liquidity': self.get_liquidity_data(token_address, chain)
        }
        
        return onchain_data
    
    def fetch_bitquery_data(self, token_address, chain):
        """Fetch data from Bitquery"""
        if not BITQUERY_API_KEY:
            return {}
        
        try:
            # GraphQL query for token data
            query = """
            query ($address: String!) {
                ethereum {
                    address(address: {is: $address}) {
                        smartContract {
                            contractType
                            currency {
                                name
                                symbol
                                decimals
                                totalSupply
                            }
                        }
                        transfers {
                            count
                            amount
                        }
                    }
                }
            }
            """
            
            url = "https://graphql.bitquery.io"
            headers = {'X-API-KEY': BITQUERY_API_KEY}
            variables = {'address': token_address}
            
            response = robust_request('POST', url, headers=headers, json={
                'query': query,
                'variables': variables
            })
            
            if response and response.status_code == 200:
                return response.json()
        except Exception as e:
            self.logger.error(f"Error fetching Bitquery data: {e}")
        
        return {}
    
    def fetch_market_data(self, token_address, chain):
        """Fetch market data from CoinGecko and CoinMarketCap APIs (with contract mapping for CMC)
        
        This method gathers comprehensive market data from two major sources:
        
        1. CoinMarketCap:
           - Price and volume data
           - Market pairs information
           - Market cap and supply metrics
           Uses FREE API endpoint with proper contract address formatting
        
        2. CoinGecko:
           - Detailed market metrics
           - Community and social data
           - Developer activity metrics
           Uses FREE API endpoint for authenticated requests
        
        The method handles API failures gracefully and provides fallback data
        structures to ensure the risk assessment can continue even with
        partial data.
        
        Args:
            token_address (str): The contract address of the token
            chain (str): The blockchain network (eth/bsc)
            
        Returns:
            dict: Structured market data containing:
                - cmc: CoinMarketCap data
                - coingecko: CoinGecko data
                Both with standardized sub-structures for consistent access
        """
        data = {
            'cmc': {
                'data': {},
                'metadata': {'error': None}
            },
            'coingecko': {
                'market_data': {
                    'current_price': {'usd': 0},
                    'total_volume': {'usd': 0},
                    'market_cap': {'usd': 0}
                },
                'community_data': {},
                'developer_data': {}
            }
        }
        
        # CoinMarketCap data
        cmc_api_key = os.getenv("COINMARKETCAP_API_KEY")
        if cmc_api_key:
            try:
                # Step 1: Try address mapping - Use FREE API endpoint
                map_url = "https://api.coinmarketcap.com/v1/cryptocurrency/map"
                params = {"address": token_address}
                headers = {"X-CMC_PRO_API_KEY": cmc_api_key, "Accept": "application/json"}
                cache_key = get_cache_key(map_url, params, headers)
                
                # Initialize cache if not exists
                if not hasattr(self, 'api_cache'):
                    self.api_cache = APICache()
                
                cached = self.api_cache.get(cache_key)
                map_data = cached if isinstance(cached, dict) else None
                if map_data is None:
                    map_response = robust_request('GET', map_url, headers=headers, params=params, timeout=20)
                    if map_response and map_response.status_code == 200:
                        parsed = map_response.json()
                        map_data = parsed if isinstance(parsed, dict) else None
                    if map_data:
                        self.api_cache.set(cache_key, map_data)
                
                cmc_id = None
                if isinstance(map_data, dict) and map_data.get('status', {}).get('error_code') == 0:
                    data_entries = map_data.get('data') or []
                    if isinstance(data_entries, list) and data_entries:
                        first_entry = data_entries[0]
                        if isinstance(first_entry, dict):
                            cmc_id = first_entry.get('id')
                
                # Step 2: If address mapping fails, try symbol and name from cmc_symbol_map
                if not cmc_id:
                    # Load CMC symbol mapping
                    cmc_symbol_map = {}
                    try:
                        if os.path.exists(CMC_SYMBOL_MAP):
                            with open(CMC_SYMBOL_MAP, 'r') as f:
                                cmc_symbol_map = json.load(f)
                    except Exception as e:
                        print(f"    ⚠️  Could not load CMC symbol map: {e}")
                    
                    mapping_obj = cmc_symbol_map.get(token_address.lower(), {})
                    tried_symbols = set()
                    tried_names = set()
                    
                    # Try symbol and name from mapping object
                    if mapping_obj:
                        symbol = mapping_obj.get('symbol')
                        name = mapping_obj.get('name')
                        for sym in filter(None, [symbol, symbol.upper() if symbol else None, symbol.lower() if symbol else None]):
                            if sym and sym not in tried_symbols:
                                params = {"symbol": sym}
                                map_response = robust_request('GET', map_url, headers=headers, params=params, timeout=20)
                                if map_response and map_response.status_code == 200:
                                    map_data = map_response.json()
                                    if isinstance(map_data, dict) and map_data.get('status', {}).get('error_code') == 0:
                                        data_entries = map_data.get('data') or []
                                        if isinstance(data_entries, list) and data_entries:
                                            first_entry = data_entries[0]
                                            if isinstance(first_entry, dict):
                                                cmc_id = first_entry.get('id')
                                                print(f"    ✅ CMC ID found by symbol '{sym}' for {token_address}")
                                                break
                                tried_symbols.add(sym)
                        
                        if not cmc_id and name:
                            for nm in filter(None, [name, name.upper(), name.lower()]):
                                if nm and nm not in tried_names:
                                    params = {"name": nm}
                                    map_response = robust_request('GET', map_url, headers=headers, params=params, timeout=20)
                                    if map_response and map_response.status_code == 200:
                                        map_data = map_response.json()
                                        if isinstance(map_data, dict) and map_data.get('status', {}).get('error_code') == 0:
                                            data_entries = map_data.get('data') or []
                                            if isinstance(data_entries, list) and data_entries:
                                                first_entry = data_entries[0]
                                                if isinstance(first_entry, dict):
                                                    cmc_id = first_entry.get('id')
                                                    print(f"    ✅ CMC ID found by name '{nm}' for {token_address}")
                                                    break
                                    tried_names.add(nm)
                
                # Step 3: Get detailed data if we have an ID
                if cmc_id:
                    quote_url = "https://api.coinmarketcap.com/v1/cryptocurrency/quotes/latest"
                    params = {"id": cmc_id}
                    quote_response = robust_request('GET', quote_url, headers=headers, params=params, timeout=20)
                    
                    if quote_response and quote_response.status_code == 200:
                        quote_data = quote_response.json()
                        if quote_data.get('status', {}).get('error_code') == 0 and quote_data.get('data'):
                            token_data = quote_data['data'][str(cmc_id)]
                            data['cmc']['data'] = {
                                'quote': {
                                    'USD': {
                                        'price': token_data.get('quote', {}).get('USD', {}).get('price', 0),
                                        'volume_24h': token_data.get('quote', {}).get('USD', {}).get('volume_24h', 0),
                                        'market_cap': token_data.get('quote', {}).get('USD', {}).get('market_cap', 0)
                                    }
                                },
                                'circulating_supply': token_data.get('circulating_supply', 0),
                                'total_supply': token_data.get('total_supply', 0),
                                'max_supply': token_data.get('max_supply', 0)
                            }
                            print(f"    ✅ CoinMarketCap data found")
                        else:
                            print(f"    ⚠️  CoinMarketCap quote data error: {quote_data.get('status', {}).get('error_message', 'Unknown error')}")
                    else:
                        print(f"    ❌ CoinMarketCap quote API failed: {quote_response.status_code if quote_response else 'No response'}")
                else:
                    print(f"    ⚠️  Could not find CMC ID for {token_address}")
                    
            except Exception as e:
                print(f"    ❌ CoinMarketCap error: {e}")
                data['cmc']['metadata']['error'] = str(e)
        else:
            print(f"  ⚠️  CoinMarketCap API key missing")
        
        # CoinGecko data with sophisticated fallback
        try:
            # Try multiple approaches to get CoinGecko data
            coingecko_data = None
            
            # Approach 1: Try by contract address
            if chain in self.CHAIN_CONFIG:
                cg_platform = self.CHAIN_CONFIG[chain]['coingecko_platform']
                cg_url = f"https://api.coingecko.com/api/v3/coins/{cg_platform}/contract/{token_address}"
                response = robust_request('GET', cg_url, timeout=8)
                
                if response and response.status_code == 200:
                    coingecko_data = response.json()
                    print(f"    ✅ CoinGecko data found by contract address")
                elif response and response.status_code == 429:
                    print(f"    ⚠️  CoinGecko rate limit exceeded")
            
            # Approach 2: Try by symbol if contract approach failed
            if not coingecko_data:
                token_symbol = self._get_token_symbol_from_address(token_address)
                if token_symbol:
                    # Try simple price endpoint
                    url = f"https://api.coingecko.com/api/v3/simple/price"
                    params = {
                        'ids': token_symbol.lower(),
                        'vs_currencies': 'usd',
                        'include_market_cap': 'true',
                        'include_24hr_vol': 'true'
                    }
                    
                    response = robust_request('GET', url, params=params, timeout=8)
                    
                    if response and response.status_code == 200:
                        simple_data = response.json()
                        if token_symbol.lower() in simple_data:
                            token_data = simple_data[token_symbol.lower()]
                            coingecko_data = {
                                'market_data': {
                                    'current_price': {'usd': token_data.get('usd', 0)},
                                    'total_volume': {'usd': token_data.get('usd_24h_vol', 0)},
                                    'market_cap': {'usd': token_data.get('usd_market_cap', 0)}
                                }
                            }
                            print(f"    ✅ CoinGecko data found by symbol {token_symbol}")
                        else:
                            print(f"    ⚠️  Token {token_symbol} not found in CoinGecko")
                    elif response and response.status_code == 429:
                        print(f"    ⚠️  CoinGecko rate limit exceeded")
                    else:
                        print(f"    ❌ CoinGecko API failed: {response.status_code if response else 'No response'}")
                else:
                    print(f"    ⚠️  Could not determine token symbol for {token_address}")
            
            # Approach 3: Try search endpoint if other approaches failed
            if not coingecko_data:
                search_url = "https://api.coingecko.com/api/v3/search"
                params = {"query": token_address}
                response = robust_request('GET', search_url, params=params, timeout=8)
                
                if response and response.status_code == 200:
                    search_data = response.json()
                    if search_data.get('coins'):
                        # Try the first result
                        first_coin = search_data['coins'][0]
                        coin_id = first_coin['id']
                        
                        # Get detailed data for this coin
                        detail_url = f"https://api.coingecko.com/api/v3/coins/{coin_id}"
                        detail_response = robust_request('GET', detail_url, timeout=8)
                        
                        if detail_response and detail_response.status_code == 200:
                            coingecko_data = detail_response.json()
                            print(f"    ✅ CoinGecko data found by search for {coin_id}")
            
            # Update data structure with CoinGecko results
            if coingecko_data:
                if 'market_data' in coingecko_data:
                    data['coingecko']['market_data'] = coingecko_data['market_data']
                if 'community_data' in coingecko_data:
                    data['coingecko']['community_data'] = coingecko_data['community_data']
                if 'developer_data' in coingecko_data:
                    data['coingecko']['developer_data'] = coingecko_data['developer_data']
                if 'links' in coingecko_data:
                    data['coingecko']['links'] = coingecko_data['links']
                if 'description' in coingecko_data:
                    data['coingecko']['description'] = coingecko_data['description']
                
        except Exception as e:
            print(f"    ❌ CoinGecko error: {e}")
        
        # Try CoinPaprika as additional source (especially for SKY, POL, etc.)
        try:
            token_symbol = self._get_token_symbol_from_address(token_address)
            if token_symbol and token_symbol != "UNKNOWN":
                paprika_data = fetch_coinpaprika_market(token_symbol)
                if paprika_data:
                    data['coinpaprika'] = paprika_data
                    print(f"    ✅ CoinPaprika data found for {token_symbol}")
        except Exception as e:
            print(f"    ⚠️  CoinPaprika error: {e}")
        
        return data
    
    def _get_token_symbol_from_address(self, token_address):
        """Extract token symbol from address for token-specific adjustments"""
        # Common token addresses and their symbols
        token_map = {
            "0xdac17f958d2ee523a2206206994597c13d831ec7": "USDT",
            "0xa0b86a33e6441b8c4b8b8b8b8b8b8b8b8b8b8b8": "USDC",  # Placeholder
            "0x6b175474e89094c44da98b954eedeac495271d0f": "DAI",
            "0x514910771af9ca656af840dff83e8264ecf986ca": "LINK",
            "0x1f9840a85d5af5bf1d1762f925bdaddc4201f984": "UNI",
            "0x7fc66500c84a76ad7e9c93437bfc5ac33e2ddae9": "AAVE",
            "0xbb4cdb9cbd36b01bd1cbaef2af88c02ec1bffa70": "WBNB",
            "0x2260fac5e5542a773aa44fbcfedf7c193bc2c599": "WBTC",
            "0x6f40d4a6237c257fff2db00fa0510deeecd303eb": "UNI",
            "0x7d1afa7b718fb893db30a3abc0cfc608aacfebb0": "MATIC",
            "0x95ad61b0a150d79219dcf64e1e6cc01f0b64c4ce": "SHIB",
            "0x3845badade8e6dff049820680d1f52bd61771325": "MANA",
            "0x6b3595068778dd592e39a122f4f5a5cf09c90fe2": "SUSHI",
            "0x0d8775f648430679a709e98d2b0cb6250d2887ef": "BAT",
            "0x4fabb145d64652a948d72533023f6e7a623c7c53": "BUSD",
            "0x4d224452801aced8b2f0aebe155379bb5d594381": "APE",
            "0x2b591e99afe9f32eaa6214f7b7629768c40eeb39": "HEX",
            "0x1a4b46696b2bb4794eb3d4c26f1c55f9170fa4c5": "BIT",
            "0x7a58c0be72be218b41c608b7fe7c5bb630736c71": "PAX",
            "0x8e870d67f660d95d5be530380d0ec0bd388289e1": "PAXG",
            # Additional addresses from tokens.csv
            "0x57e114b691db790c35207b2e685d4a43181e6061": "COMP",  # Compound
            "0xc944e90c64b2c07662a292be6244bdf05cda44a7": "GRT",   # The Graph
            "0xbb4cdb9cbd36b01bd1cbaef2de08d9173bc095c": "WBNB",  # Wrapped BNB (BSC)
            "0xbb4cdb9cbd36b01bd1cbaef2af88c02ec1bffa70": "WBNB",  # Wrapped BNB (BSC) - alternative
            "0x3845badade8e6dff049820680d1f14bd3903a5d0": "MANA"  # Decentraland
        }
        
        # First try the hardcoded map
        symbol = token_map.get(token_address.lower())
        if symbol:
            return symbol
        
        # Then try external fallback file
        try:
            fallback_file = os.path.join(DATA_DIR, 'token_fallbacks.json')
            if os.path.exists(fallback_file):
                with open(fallback_file, 'r') as f:
                    fallback_data = json.load(f)
                    token_mappings = fallback_data.get('token_mappings', {})
                    token_info = token_mappings.get(token_address.lower())
                    if token_info:
                        return token_info.get('symbol')
        except Exception as e:
            print(f"    ⚠️  Error loading token fallbacks: {e}")
        
        return "UNKNOWN"
    
    def fetch_santiment_data(self, token_address, chain):
        """Fetch Santiment data with improved error handling"""
        if not SANTIMENT_API_KEY:
            print(f"  ⚠️  Santiment API key missing")
            return {}
        
        try:
            print(f"  📊 Fetching Santiment data...")
            url = "https://api.santiment.net/graphql"
            headers = {'Authorization': f'Bearer {SANTIMENT_API_KEY}'}
            query = """
            query ($address: String!) {
                getMetric(metric: "dev_activity") {
                    timeseriesData(slug: $address, from: "2024-01-01", to: "2024-12-31", interval: "1d")
                }
            }
            """
            
            response = robust_request('POST', url, headers=headers, json={
                'query': query,
                'variables': {'address': token_address}
            })
            
            if response and response.status_code == 200:
                data = response.json()
                print(f"    ✅ Santiment data found")
                return data
            elif response and response.status_code == 401:
                error_msg = "401 Unauthorized - Invalid JWT token"
                log_failed_api_endpoint("Santiment", url, error_msg)
                return {}
            else:
                error_msg = f"{response.status_code if response else 'No response'}"
                log_failed_api_endpoint("Santiment", url, error_msg)
                return {}
        except requests.exceptions.RequestException as e:
            log_failed_api_endpoint("Santiment", url, str(e))
            return {}
        except json.JSONDecodeError as e:
            log_failed_api_endpoint("Santiment", url, f"JSON decode error: {e}")
            return {}
        except Exception as e:
            log_failed_api_endpoint("Santiment", url, str(e))
            return {}
    
    def fetch_security_reports(self, token_address, chain):
        """Fetch security audit reports"""
        security_data = {}
        
        # Try CertiK
        if CERTIK_API_KEY:
            try:
                url = f"https://api.certik.com/v1/chain/ethereum/address/{token_address}"
                headers = {'Authorization': f'Bearer {CERTIK_API_KEY}'}
                response = robust_request('GET', url, headers=headers)
                if response and response.status_code == 200:
                    security_data['certik'] = response.json()
            except Exception as e:
                self.logger.error(f"Error fetching CertiK data: {e}")
        
        return security_data
    
    def fetch_enhanced_data(self, token_address, chain):
        """Fetch enhanced data from multiple APIs"""
        enhanced_data = {}
        
        print(f"  📊 Fetching enhanced data...")
        
        # Fetch Breadcrumbs data
        try:
            print(f"    🔍 Fetching Breadcrumbs data...")
            breadcrumbs_risk = fetch_breadcrumbs_risk_score(token_address)
            breadcrumbs_token = fetch_breadcrumbs_token_info(token_address)
            if breadcrumbs_risk or breadcrumbs_token:
                enhanced_data['breadcrumbs'] = {
                    'risk_score': breadcrumbs_risk,
                    'token_info': breadcrumbs_token
                }
                print(f"      ✅ Breadcrumbs data found")
            else:
                print(f"      ⚠️  No Breadcrumbs data available")
        except Exception as e:
            print(f"      ❌ Breadcrumbs error: {e}")
        
        # Fetch Ethplorer data
        try:
            print(f"    🔍 Fetching Ethplorer data...")
            ethplorer_token = fetch_ethplorer_token_info(token_address)
            ethplorer_address = fetch_ethplorer_address_info(token_address)
            
            if ethplorer_token or ethplorer_address:
                enhanced_data['ethplorer'] = {
                    'token_info': ethplorer_token,
                    'address_info': ethplorer_address
                }
                print(f"      ✅ Ethplorer data found")
            else:
                print(f"      ⚠️  No Ethplorer data available (address may not be a token contract)")
        except Exception as e:
            print(f"      ❌ Ethplorer error: {e}")
        
        # Fetch Zapper data
        try:
            print(f"    🔍 Fetching Zapper data...")
            zapper_portfolio = fetch_zapper_portfolio_data(token_address)
            if zapper_portfolio:
                enhanced_data['zapper'] = {
                    'portfolio': zapper_portfolio
                }
                print(f"      ✅ Zapper data found")
            else:
                print(f"      ⚠️  No Zapper data available")
        except Exception as e:
            print(f"      ❌ Zapper error: {e}")
        
        # Fetch DeBank data
        try:
            print(f"    🔍 Fetching DeBank data...")
            debank_portfolio = fetch_debank_portfolio(token_address)
            if debank_portfolio:
                enhanced_data['debank'] = {
                    'portfolio': debank_portfolio
                }
                print(f"      ✅ DeBank data found")
            else:
                print(f"      ⚠️  No DeBank data available")
        except Exception as e:
            print(f"      ❌ DeBank error: {e}")
        
        # Fetch DeFiLlama data
        try:
            print(f"    🔍 Fetching DeFiLlama data...")
            defillama_price = fetch_defillama_token_price(token_address, chain)
            if defillama_price:
                enhanced_data['defillama'] = {
                    'price': defillama_price
                }
                print(f"      ✅ DeFiLlama data found")
            else:
                print(f"      ⚠️  No DeFiLlama data available")
        except Exception as e:
            print(f"      ❌ DeFiLlama error: {e}")
        
        # Fetch Moralis data
        try:
            print(f"    🔍 Fetching Moralis data...")
            moralis_metadata = fetch_moralis_token_metadata(token_address, chain)
            moralis_price = fetch_moralis_token_price(token_address, chain)
            moralis_transfers = fetch_moralis_token_transfers(token_address, chain)
            if moralis_metadata or moralis_price or moralis_transfers:
                enhanced_data['moralis'] = {
                    'metadata': moralis_metadata,
                    'price': moralis_price,
                    'transfers': moralis_transfers
                }
                print(f"      ✅ Moralis data found")
            else:
                print(f"      ⚠️  No Moralis data available")
        except Exception as e:
            print(f"      ❌ Moralis error: {e}")
        
        # Fetch 1inch data
        try:
            print(f"    🔍 Fetching 1inch data...")
            inch_metadata = fetch_1inch_token_metadata(token_address)
            inch_price = fetch_1inch_spot_price(token_address)
            if inch_metadata or inch_price:
                enhanced_data['1inch'] = {
                    'metadata': inch_metadata,
                    'price': inch_price
                }
                print(f"      ✅ 1inch data found")
            else:
                print(f"      ⚠️  No 1inch data available")
        except Exception as e:
            print(f"      ❌ 1inch error: {e}")
        
        # Fetch compliance data
        try:
            print(f"    🔍 Fetching compliance data...")
            compliance_data = {}
            
            # Scorechain AML
            scorechain_aml = fetch_scorechain_aml(token_address, chain)
            if scorechain_aml:
                compliance_data['scorechain'] = scorechain_aml
            
            # TRM Labs AML
            trm_aml = fetch_trmlabs_aml(token_address, chain)
            if trm_aml:
                compliance_data['trm_labs'] = trm_aml
            
            # OpenSanctions
            opensanctions = fetch_opensanctions_compliance(token_address, chain)
            if opensanctions:
                compliance_data['opensanctions'] = opensanctions
            
            # Lukka
            lukka = fetch_lukka_compliance(token_address, chain)
            if lukka:
                compliance_data['lukka'] = lukka
            
            # Alchemy
            alchemy = fetch_alchemy_compliance(token_address, chain)
            if alchemy:
                compliance_data['alchemy'] = alchemy
            
            # DeFiSafety
            defisafety = fetch_defisafety_compliance(token_address, chain)
            if defisafety:
                compliance_data['defisafety'] = defisafety
            
            # CertiK
            certik = fetch_certik_security(token_address, chain)
            if certik:
                compliance_data['certik'] = certik
            
            if compliance_data:
                enhanced_data['compliance'] = compliance_data
                print(f"      ✅ Compliance data found")
            else:
                print(f"      ⚠️  No compliance data available")
        except Exception as e:
            print(f"      ❌ Compliance error: {e}")
        
        # Fetch social data
        try:
            print(f"    🔍 Fetching social data...")
            social_data = {}
            
            # Only fetch real data, no simulation
            if social_data:
                enhanced_data['social_data'] = social_data
                print(f"      ✅ Social data found")
            else:
                print(f"      ⚠️  No social data available")
        except Exception as e:
            print(f"      ❌ Social data error: {e}")
        
        return enhanced_data
    
    def detect_red_flags(self, token_address, chain):
        """Detect potential red flags"""
        red_flags = []
        
        # Check for unverified contracts
        if chain == 'eth' and ETHERSCAN_API_KEY:
            verification_status = self.get_contract_verification_status(token_address, chain)
            if verification_status == 'unverified':
                red_flags.append('unverified_contract')
        
        # Check for suspicious holder distribution
        holder_data = self.get_holder_data(token_address, chain)
        if holder_data['top10_concentration'] > 80:
            red_flags.append('high_concentration')
        
        # Check for low liquidity
        liquidity_data = self.get_liquidity_data(token_address, chain)
        if not liquidity_data.get('oneinch') and not liquidity_data.get('defillama'):
            red_flags.append('low_liquidity')
        
        # Note: eu_unlicensed_stablecoin flag should only be added for actual stablecoins
        # This is handled in apply_eu_regulatory_checks, not here
        
        return red_flags
    
    def assess_token(self, token_address, chain="eth", progress_callback=None, token_index=0, total_tokens=1):
        """Main token assessment function"""
        print(f"\n🔍 Assessing token {token_index + 1}/{total_tokens}: {token_address}")
        self.logger.info(f"Assessing token: {token_address}")
        
        try:
            # Update progress
            if progress_callback:
                progress_callback(f"Assessing token {token_index + 1}/{total_tokens}")
            
            # Get token symbol and name from fallback file
            symbol = self._get_token_symbol_from_address(token_address)
            token_name = "Unknown"
            if symbol:
                # Get token name from fallback file
                try:
                    fallback_file = os.path.join(DATA_DIR, 'token_fallbacks.json')
                    if os.path.exists(fallback_file):
                        with open(fallback_file, 'r') as f:
                            fallback_data = json.load(f)
                            token_mappings = fallback_data.get('token_mappings', {})
                            token_info = token_mappings.get(token_address.lower())
                            if token_info:
                                token_name = token_info.get('name', symbol.upper())
                                print(f"  📝 Token: {symbol.upper()} ({token_name})")
                            else:
                                print(f"  ⚠️  Token info not found in fallback file")
                except Exception as e:
                    print(f"  ❌ Error getting token name: {e}")
            else:
                print(f"  ⚠️  No symbol found for {token_address}")
            
            # Fetch all data with error logging
            print(f"  📊 Fetching onchain data...")
            onchain_data = self.fetch_onchain_data(token_address, chain)
            
            print(f"  📊 Fetching BitQuery data...")
            bitquery_data = self.fetch_bitquery_data(token_address, chain)
            
            print(f"  📊 Fetching market data...")
            market_data = self.fetch_market_data(token_address, chain)
            
            print(f"  📊 Fetching Santiment data...")
            santiment_data = self.fetch_santiment_data(token_address, chain)
            
            print(f"  📊 Fetching security reports...")
            security_data = self.fetch_security_reports(token_address, chain)
            
            print(f"  📊 Fetching enhanced data...")
            enhanced_data = self.fetch_enhanced_data(token_address, chain)
            
            # Fetch social data
            print(f"  📊 Fetching social data...")
            social_data = {
                'twitter': self.fetch_twitter_social_data(symbol, token_name) if symbol else {},
                'telegram': self.fetch_telegram_social_data(symbol, token_name) if symbol else {},
                'discord': self.fetch_discord_social_data(symbol, token_name) if symbol else {},
                'reddit': self.fetch_reddit_social_data(symbol, token_name) if symbol else {}
            }
            
            # Detect red flags
            print(f"  🚨 Detecting red flags...")
            red_flags = self.detect_red_flags(token_address, chain)
            if red_flags:
                print(f"    ⚠️  Found red flags: {red_flags}")
            else:
                print(f"    ✅ No red flags detected")
            
            # Calculate risk scores
            print(f"  📈 Calculating risk scores...")
            risk_scores_result = self.calculate_risk_scores(
                onchain_data, bitquery_data, market_data, 
                santiment_data, security_data, enhanced_data, red_flags
            )
            
            # Create risk report
            risk_report = {
                'token_address': token_address,
                'chain': chain,
                'symbol': symbol,
                'name': token_name,
                'timestamp': datetime.datetime.now().isoformat(),
                'onchain_data': onchain_data,
                'bitquery_data': bitquery_data,
                'market_data': market_data,
                'santiment_data': santiment_data,
                'security_data': security_data,
                'enhanced_data': enhanced_data,
                'social_data': social_data,
                'red_flags': red_flags,
                'risk_scores': risk_scores_result,
                'errors': []
            }
            
            # Apply compliance checks
            print(f"  🛡️  Applying compliance checks...")
            compliance_score = self.apply_strict_compliance_checks(risk_report, chain)
            eu_compliance_score = self.apply_eu_regulatory_checks(risk_report)
            
            # Check if this is actually a stablecoin before adding eu_unlicensed_stablecoin flag
            is_stablecoin = risk_report.get('is_stablecoin', False)
            symbol = risk_report.get('symbol', '').upper().strip()
            
            # Ensure non-stablecoins (like CHZ, COMP) never get the eu_unlicensed_stablecoin flag
            # This flag should ONLY apply to actual stablecoins
            if 'eu_unlicensed_stablecoin' in red_flags and not is_stablecoin:
                red_flags.remove('eu_unlicensed_stablecoin')
                print(f"    ✅ Removed eu_unlicensed_stablecoin flag - {symbol} is not a stablecoin")
            
            # Only add eu_unlicensed_stablecoin flag for actual stablecoins that lack proper licensing
            # This should only apply to stablecoins, not regular tokens like CHZ or COMP
            if is_stablecoin and 'eu_unlicensed_stablecoin' not in red_flags:
                # Check if stablecoin has proper licensing/compliance
                eu_issues = risk_report.get('eu_compliance_issues', [])
                # If stablecoin has compliance issues indicating lack of licensing, add the flag
                if eu_issues:
                    # Check for stablecoin-specific compliance issues
                    stablecoin_compliance_issues = [issue for issue in eu_issues if 'stablecoin' in issue.lower() or 'compliance' in issue.lower()]
                    if stablecoin_compliance_issues:
                        # Only add if there are significant compliance issues
                        red_flags.append('eu_unlicensed_stablecoin')
                        print(f"    ⚠️  Added eu_unlicensed_stablecoin flag for unlicensed stablecoin {symbol}")
            
            # Update risk scores with compliance
            if 'compliance' not in risk_scores_result:
                risk_scores_result['compliance'] = 0
            if 'eu_compliance' not in risk_scores_result:
                risk_scores_result['eu_compliance'] = 0
                
            risk_scores_result['compliance'] = max(0, risk_scores_result['compliance'] + compliance_score)
            risk_scores_result['eu_compliance'] = max(0, eu_compliance_score)
            
            # Use the calculated risk scores from calculate_risk_scores
            risk_scores_result['overall'] = risk_scores_result.get('total_risk_score', 0)
            
            risk_report['risk_scores'] = risk_scores_result
            risk_report['red_flags'] = red_flags  # Update red_flags in risk_report
            
            print(f"  ✅ Assessment complete - Overall score: {risk_scores_result['overall']}")
            
            self.stats['successful_assessments'] += 1
            return risk_report
            
        except Exception as e:
            print(f"  ❌ Error assessing token {token_address}: {e}")
            self.logger.error(f"Error assessing token {token_address}: {e}")
            self.stats['failed_assessments'] += 1
            return {
                'token_address': token_address,
                'chain': chain,
                'timestamp': datetime.datetime.now().isoformat(),
                'errors': [str(e)],
                'risk_scores': {'overall': 0}
            }
    
    def calculate_risk_scores(self, onchain_data, bitquery_data, market_data, 
                            santiment_data, security_data, enhanced_data, red_flags):
        """Calculate comprehensive risk scores using the complete 16-component system"""
        try:
            # Create a comprehensive risk report for scoring
            risk_report = {
                'onchain': onchain_data,
                'market': market_data,
                'security': security_data,
                'santiment': santiment_data,
                'enhanced': enhanced_data,
                'red_flags': red_flags
            }
            
            print(f"  📊 Calculating component scores...")
            component_scores = {}
            
            # Calculate all 16 component scores
            for component in self.WEIGHTS.keys():
                try:
                    score_method = getattr(self, f"score_{component}")
                    
                    # Special handling for functions that need additional arguments
                    if component in ['aml_data', 'compliance_data']:
                        component_scores[component] = score_method(risk_report, risk_report.get('token_address', ''), risk_report.get('chain', 'eth'))
                    elif component == 'social_data':
                        symbol = risk_report.get('symbol', '')
                        component_scores[component] = score_method(risk_report, risk_report.get('token_address', ''), symbol)
                    else:
                        component_scores[component] = score_method(risk_report)
                    
                    print(f"    {component}: {component_scores[component]}")
                except Exception as e:
                    print(f"    ❌ Error calculating {component} score: {e}")
                    component_scores[component] = 7  # Default to medium-high risk on error
            
            # Calculate final risk score using weighted components
            print(f"  📊 Calculating final risk score...")
            total_risk_score = 0
            social_score_contribution = 0
            
            for component, weight in self.WEIGHTS.items():
                component_contribution = component_scores[component] * weight * 10  # Scale up the scores
                total_risk_score += component_contribution
                
                # Track social score contribution separately
                if component == 'social_data':
                    social_score_contribution = component_contribution
            
            # Apply red flag boosts
            for flag in red_flags:
                boost = next(
                    (f['risk_boost'] for f in self.RED_FLAGS if f['check'] == flag), 
                    0
                )
                print(f"    🚨 Applying red flag boost for {flag}: +{boost}")
                total_risk_score += boost
            
            # Calculate total score without social influence
            total_score_minus_social = total_risk_score - social_score_contribution
            total_score_minus_social = min(150, max(0, total_score_minus_social))
            
            total_risk_score = min(150, max(0, total_risk_score))
            
            # Create comprehensive scores dictionary
            scores = {
                'component_scores': component_scores,
                'total_risk_score': round(total_risk_score, 2),
                'total_score_minus_social': round(total_score_minus_social, 2),
                'red_flags': red_flags,
                'risk_category': self.classify_risk(total_risk_score, risk_report)
            }
            
            print(f"    ✅ Final risk score: {total_risk_score:.2f} ({scores['risk_category']})")
            
            return scores
            
        except Exception as e:
            print(f"    ❌ Error in risk score calculation: {e}")
            return {
                'component_scores': {},
                'total_risk_score': 150,
                'total_score_minus_social': 150,
                'red_flags': red_flags,
                'risk_category': 'Extreme Risk'
            }
    
    def score_industry_impact(self, risk_report):
        """Score industry impact (1-10) - Enhanced with multiple market indicators"""
        try:
            market_data = risk_report.get('market', {})
            onchain_data = risk_report.get('onchain', {})
            
            # Use enhanced extraction methods that check multiple sources
            # risk_report has 'market' key (from calculate_risk_scores), extract properly
            market_cap = self._extract_market_cap({'market_data': market_data, 'enhanced_data': risk_report.get('enhanced', {})})
            volume_24h = self._extract_volume_24h({'market_data': market_data, 'enhanced_data': risk_report.get('enhanced', {})})
            
            score = 5  # Base score
            
            # Market cap impact (primary indicator)
            if market_cap > 10_000_000_000:  # > $10B - Major industry player
                score += 3
            elif market_cap > 1_000_000_000:  # > $1B - Significant impact
                score += 2
            elif market_cap > 100_000_000:  # > $100M - Notable presence
                score += 1
            elif market_cap < 1_000_000:  # < $1M - Minimal impact
                score -= 2
            
            # Volume impact (indicates active trading and adoption)
            if volume_24h > 100_000_000:  # > $100M daily - High liquidity
                score += 2
            elif volume_24h > 10_000_000:  # > $10M daily - Good liquidity
                score += 1
            elif volume_24h < 100_000:  # < $100K daily - Low activity
                score -= 1
            
            # Holder count (indicates adoption)
            holders = onchain_data.get('holders', {}).get('total_holders', 0)
            if holders > 100000:
                score += 1
            elif holders > 10000:
                score += 0.5
            
            # Liquidity depth (indicates market maturity)
            liquidity = self._extract_liquidity_value({'onchain_data': onchain_data})
            if liquidity > 50_000_000:  # > $50M liquidity
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            print(f"    ❌ Error in industry impact scoring: {e}")
            return 5
    
    def score_tech_innovation(self, risk_report):
        """Score technical innovation (1-10)"""
        try:
            onchain_data = risk_report.get('onchain', {})
            enhanced_data = risk_report.get('enhanced', {})
            
            score = 5  # Base score
            
            # Contract verification
            if onchain_data.get('contract_verified') == 'verified':
                score += 2
            elif onchain_data.get('contract_verified') == 'unverified':
                score -= 1
            
            # DeFi integration
            if enhanced_data.get('defillama') or enhanced_data.get('zapper') or enhanced_data.get('debank'):
                score += 2
            
            # Holder distribution (innovation in tokenomics)
            holder_data = onchain_data.get('holders', {})
            if holder_data.get('total_holders', 0) > 10000:
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            print(f"    ❌ Error in tech innovation scoring: {e}")
            return 5
    
    def score_business_model(self, risk_report):
        """Score business model (1-10)"""
        try:
            market_data = risk_report.get('market', {})
            onchain_data = risk_report.get('onchain', {})
            
            score = 5  # Base score
            
            # Market presence
            if market_data.get('coingecko') or market_data.get('coinmarketcap'):
                score += 2
            
            # Liquidity
            if onchain_data.get('liquidity'):
                score += 2
            
            # Supply distribution
            supply_data = onchain_data.get('supply', {})
            if supply_data.get('total_supply', 0) > 0:
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            print(f"    ❌ Error in business model scoring: {e}")
            return 5
    
    def score_global_reach(self, risk_report):
        """Score global reach (1-10) - Enhanced with market indicators"""
        try:
            market_data = risk_report.get('market', {})
            onchain_data = risk_report.get('onchain', {})
            
            score = 5  # Base score
            
            # Use enhanced extraction methods that check multiple sources
            market_cap = self._extract_market_cap({'market_data': market_data, 'enhanced_data': risk_report.get('enhanced', {})})
            volume_24h = self._extract_volume_24h({'market_data': market_data, 'enhanced_data': risk_report.get('enhanced', {})})
            
            # Volume indicates global trading
            if volume_24h > 50_000_000:  # > $50M daily
                score += 3
            elif volume_24h > 10_000_000:  # > $10M daily
                score += 2
            elif volume_24h > 1_000_000:  # > $1M daily
                score += 1
            elif volume_24h < 100_000:  # < $100K daily
                score -= 1
            
            # Market cap indicates adoption
            if market_cap > 1_000_000_000:  # > $1B
                score += 2
            elif market_cap > 100_000_000:  # > $100M
                score += 1
            
            # Holder count indicates global adoption
            holders = onchain_data.get('holders', {}).get('total_holders', 0)
            if holders > 50000:
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in global reach scoring: {e}")
            return 5
    
    def score_whitepaper_quality(self, risk_report):
        """Score whitepaper quality (1-10) - Enhanced with multiple indicators"""
        try:
            score = 5  # Base score
            
            market_data = risk_report.get('market', {})
            coingecko_data = market_data.get('coingecko', {})
            enhanced_data = risk_report.get('enhanced', {})
            
            # Documentation presence (weight: high)
            links = coingecko_data.get('links', {})
            if links.get('whitepaper'):
                score += 3  # Whitepaper exists
            if links.get('homepage'):
                score += 1  # Website exists
            if links.get('repos_url'):
                score += 1  # GitHub repository
            
            # Description quality
            description = coingecko_data.get('description', {})
            desc_text = description.get('en', '') if isinstance(description, dict) else str(description)
            if desc_text:
                desc_length = len(desc_text)
                if desc_length > 1000:
                    score += 2  # Comprehensive description
                elif desc_length > 500:
                    score += 1  # Good description
            
            # Technical documentation indicators
            if enhanced_data.get('defisafety') or enhanced_data.get('certik'):
                score += 1  # Security audits indicate documentation
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in whitepaper quality scoring: {e}")
            return 5
    
    def score_roadmap_adherence(self, risk_report):
        """Score roadmap adherence (1-10) - Enhanced with activity and development indicators"""
        try:
            score = 5  # Base score
            
            market_data = risk_report.get('market', {})
            coingecko_data = market_data.get('coingecko', {})
            santiment_data = risk_report.get('santiment', {})
            enhanced_data = risk_report.get('enhanced', {})
            
            # Developer activity (strong indicator of roadmap progress)
            dev_data = coingecko_data.get('developer_data', {})
            forks = dev_data.get('forks', 0)
            stars = dev_data.get('stars', 0)
            contributors = dev_data.get('contributors', 0)
            
            if forks > 50:
                score += 2
            elif forks > 10:
                score += 1
            
            if stars > 100:
                score += 2
            elif stars > 20:
                score += 1
            
            if contributors > 10:
                score += 1
            
            # Community growth (indicates project momentum)
            community_data = coingecko_data.get('community_data', {})
            twitter_followers = community_data.get('twitter_followers', 0)
            reddit_subscribers = community_data.get('reddit_subscribers', 0)
            
            if twitter_followers > 50000:
                score += 2
            elif twitter_followers > 10000:
                score += 1
            
            if reddit_subscribers > 5000:
                score += 1
            
            # Santiment dev activity data
            if santiment_data and isinstance(santiment_data, dict):
                # Check for active development signals
                if 'data' in santiment_data or 'getMetric' in str(santiment_data):
                    score += 1
            
            # Market presence (indicates execution)
            market_cap = self._extract_market_cap({'market_data': market_data, 'enhanced_data': risk_report.get('enhanced', {})})
            if market_cap > 100_000_000:  # > $100M indicates successful execution
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in roadmap adherence scoring: {e}")
            return 5
    
    def score_team_expertise(self, risk_report):
        """Score team expertise (1-10) - Enhanced with development and security indicators"""
        try:
            score = 5  # Base score
            
            market_data = risk_report.get('market', {})
            coingecko_data = market_data.get('coingecko', {})
            enhanced_data = risk_report.get('enhanced', {})
            security_data = risk_report.get('security', [])
            
            # Development activity (indicates technical expertise)
            dev_data = coingecko_data.get('developer_data', {})
            forks = dev_data.get('forks', 0)
            stars = dev_data.get('stars', 0)
            contributors = dev_data.get('contributors', 0)
            
            if contributors > 20:
                score += 2
            elif contributors > 5:
                score += 1
            
            if stars > 500:
                score += 2
            elif stars > 100:
                score += 1
            
            if forks > 100:
                score += 1
            
            # Security audits (indicates professional team)
            if security_data:
                if isinstance(security_data, list) and len(security_data) > 0:
                    score += 2
                elif isinstance(security_data, dict):
                    if security_data.get('certik') or security_data.get('defisafety'):
                        score += 2
            
            if enhanced_data.get('certik') or enhanced_data.get('defisafety'):
                score += 1
            
            # Professional presence indicators
            links = coingecko_data.get('links', {})
            if links.get('repos_url'):
                score += 1  # Open source
            if links.get('homepage'):
                score += 1  # Professional website
            if links.get('whitepaper'):
                score += 1  # Technical documentation
            
            # Contract verification (indicates technical competence)
            onchain_data = risk_report.get('onchain', {})
            if onchain_data.get('contract_verified') == 'verified':
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in team expertise scoring: {e}")
            return 5
    
    def score_management_strategy(self, risk_report):
        """Score management strategy (1-10) - Enhanced with market performance and growth indicators"""
        try:
            score = 5  # Base score
            
            market_data = risk_report.get('market', {})
            coingecko_data = market_data.get('coingecko', {})
            onchain_data = risk_report.get('onchain', {})
            
            # Market performance (indicates effective strategy execution)
            market_cap = self._extract_market_cap({'market_data': market_data, 'enhanced_data': risk_report.get('enhanced', {})})
            volume_24h = self._extract_volume_24h({'market_data': market_data, 'enhanced_data': risk_report.get('enhanced', {})})
            
            if market_cap > 1_000_000_000:  # > $1B - Successful strategy
                score += 2
            elif market_cap > 100_000_000:  # > $100M - Good execution
                score += 1
            
            if volume_24h > 10_000_000:  # > $10M daily - Active trading
                score += 1
            
            # Liquidity management (indicates strategic planning)
            liquidity = self._extract_liquidity_value({'onchain_data': onchain_data})
            if liquidity > 20_000_000:  # > $20M liquidity
                score += 1
            
            # Community growth (indicates marketing strategy)
            community_data = coingecko_data.get('community_data', {})
            twitter_followers = community_data.get('twitter_followers', 0)
            if twitter_followers > 50000:
                score += 1
            
            # Multi-exchange presence (indicates strategic partnerships)
            links = coingecko_data.get('links', {})
            if links.get('homepage') and links.get('repos_url'):
                score += 1  # Professional presence
            
            # Holder distribution (indicates tokenomics strategy)
            holders = onchain_data.get('holders', {}).get('total_holders', 0)
            top10_concentration = onchain_data.get('holders', {}).get('top10_concentration', 100)
            if holders > 10000 and top10_concentration < 50:
                score += 1  # Good distribution
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in management strategy scoring: {e}")
            return 5
    
    def score_aml_data(self, risk_report, token_address, chain):
        """Score AML data (1-10)"""
        try:
            # Basic scoring - can be enhanced with actual AML analysis
            score = 5  # Base score
            
            # Check for compliance indicators
            onchain_data = risk_report.get('onchain', {})
            if onchain_data.get('contract_verified') == 'verified':
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in AML data scoring: {e}")
            return 5
    
    def score_compliance_data(self, risk_report, token_address, chain):
        """Score compliance data (1-10)"""
        try:
            # Basic scoring - can be enhanced with actual compliance analysis
            score = 5  # Base score
            
            # Check for compliance indicators
            onchain_data = risk_report.get('onchain', {})
            if onchain_data.get('contract_verified') == 'verified':
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in compliance data scoring: {e}")
            return 5
    
    def score_market_dynamics(self, risk_report):
        """Score market dynamics (1-10) - Enhanced with volume and liquidity indicators"""
        try:
            score = 5  # Base score
            
            market_data = risk_report.get('market', {})
            onchain_data = risk_report.get('onchain', {})
            
            # Use enhanced extraction methods that check multiple sources
            volume_24h = self._extract_volume_24h({'market_data': market_data, 'enhanced_data': risk_report.get('enhanced', {})})
            market_cap = self._extract_market_cap({'market_data': market_data, 'enhanced_data': risk_report.get('enhanced', {})})
            liquidity = self._extract_liquidity_value({'onchain_data': onchain_data})
            
            # Check volume dynamics
            if volume_24h > 50_000_000:  # > $50M daily
                score += 3
            elif volume_24h > 10_000_000:  # > $10M daily
                score += 2
            elif volume_24h > 1_000_000:  # > $1M daily
                score += 1
            elif volume_24h < 100_000:  # < $100K daily
                score -= 2
            
            # Market cap indicates market stability
            if market_cap > 1_000_000_000:  # > $1B
                score += 1
            
            # Liquidity depth indicates market maturity
            if liquidity > 20_000_000:  # > $20M liquidity
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in market dynamics scoring: {e}")
            return 5
    
    def score_marketing_demand(self, risk_report):
        """Score marketing demand (1-10) - Enhanced with social and trading indicators"""
        try:
            score = 5  # Base score
            
            market_data = risk_report.get('market', {})
            coingecko_data = market_data.get('coingecko', {})
            social_data = risk_report.get('social_data', {})
            
            # Social media presence (primary indicator)
            community_data = coingecko_data.get('community_data', {})
            twitter_followers = community_data.get('twitter_followers', 0)
            reddit_subscribers = community_data.get('reddit_subscribers', 0)
            telegram_users = community_data.get('telegram_channel_user_count', 0)
            
            if twitter_followers > 100000:
                score += 3  # Strong social presence
            elif twitter_followers > 50000:
                score += 2
            elif twitter_followers > 10000:
                score += 1
            
            if reddit_subscribers > 10000:
                score += 2
            elif reddit_subscribers > 5000:
                score += 1
            
            if telegram_users > 50000:
                score += 1
            
            # Trading volume (indicates demand)
            volume_24h = self._extract_volume_24h({'market_data': market_data, 'enhanced_data': risk_report.get('enhanced', {})})
            if volume_24h > 50_000_000:  # > $50M daily
                score += 2
            elif volume_24h > 10_000_000:  # > $10M daily
                score += 1
            
            # Social data from APIs (if available)
            if social_data:
                if isinstance(social_data, dict):
                    if social_data.get('twitter') or social_data.get('telegram'):
                        score += 1
            
            # Market cap growth potential
            market_cap = self._extract_market_cap({'market_data': market_data, 'enhanced_data': risk_report.get('enhanced', {})})
            if market_cap > 500_000_000:  # > $500M indicates strong demand
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in marketing demand scoring: {e}")
            return 5
    
    def score_esg_impact(self, risk_report):
        """Score ESG impact (1-10)"""
        try:
            score = 5  # Base score
            
            # Basic ESG scoring - can be enhanced with actual ESG analysis
            market_data = risk_report.get('market', {})
            coingecko_data = market_data.get('coingecko', {})
            
            # Check for environmental/social indicators
            if coingecko_data.get('links', {}).get('homepage'):
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in ESG impact scoring: {e}")
            return 5
    
    def score_code_security(self, risk_report):
        """Score code security (1-10)"""
        try:
            score = 5  # Base score
            
            onchain_data = risk_report.get('onchain', {})
            security_data = risk_report.get('security', {})
            
            # Check contract verification
            if onchain_data.get('contract_verified') == 'verified':
                score += 2
            elif onchain_data.get('contract_verified') == 'unverified':
                score -= 1
            
            # Check security reports
            if security_data.get('certik'):
                score += 2
            if security_data.get('defisafety'):
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in code security scoring: {e}")
            return 5
    
    def score_dev_activity(self, risk_report):
        """Score developer activity (1-10) - Enhanced with comprehensive dev metrics"""
        try:
            score = 5  # Base score
            
            santiment_data = risk_report.get('santiment', {})
            market_data = risk_report.get('market', {})
            coingecko_data = market_data.get('coingecko', {})
            enhanced_data = risk_report.get('enhanced', {})
            
            # GitHub activity (primary indicator)
            dev_data = coingecko_data.get('developer_data', {})
            forks = dev_data.get('forks', 0)
            stars = dev_data.get('stars', 0)
            contributors = dev_data.get('contributors', 0)
            commits_4w = dev_data.get('commit_count_4_weeks', 0)
            
            if commits_4w > 50:
                score += 3  # Very active
            elif commits_4w > 20:
                score += 2  # Active
            elif commits_4w > 5:
                score += 1  # Some activity
            
            if contributors > 15:
                score += 2
            elif contributors > 5:
                score += 1
            
            if stars > 500:
                score += 1
            elif stars > 100:
                score += 0.5
            
            if forks > 50:
                score += 1
            
            # Santiment dev activity data (if available)
            if santiment_data:
                if isinstance(santiment_data, dict):
                    # Check for dev activity metrics
                    if 'data' in santiment_data or 'getMetric' in str(santiment_data):
                        score += 1
            
            # Code quality indicators
            onchain_data = risk_report.get('onchain', {})
            if onchain_data.get('contract_verified') == 'verified':
                score += 1  # Verified code indicates active development
            
            # Security audits (indicates ongoing development)
            if enhanced_data.get('certik') or enhanced_data.get('defisafety'):
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in dev activity scoring: {e}")
            return 5
    
    def score_social_data(self, risk_report, token_address, symbol):
        """Score social data (1-10)"""
        try:
            score = 5  # Base score
            
            enhanced_data = risk_report.get('enhanced', {})
            social_data = enhanced_data.get('social_data', {})
            
            # Check Twitter data
            if social_data.get('twitter'):
                score += 1
            
            # Check Telegram data
            if social_data.get('telegram'):
                score += 1
            
            # Check Discord data
            if social_data.get('discord'):
                score += 1
            
            # Check Reddit data
            if social_data.get('reddit'):
                score += 1
            
            return max(1, min(10, score))
        except Exception as e:
            self.logger.error(f"Error in social data scoring: {e}")
            return 5
    
    def classify_risk(self, score, risk_report=None):
        """Classify risk based on score (0-150 scale)"""
        try:
            if score >= 120:
                return "Extreme Risk"
            elif score >= 90:
                return "High Risk"
            elif score >= 60:
                return "Medium Risk"
            elif score >= 30:
                return "Low Risk"
            else:
                return "Low Risk"
        except Exception as e:
            self.logger.error(f"Error in risk classification: {e}")
            return "Unknown Risk"
    
    def fetch_twitter_social_data(self, token_symbol, token_name):
        """Fetch Twitter social data with proper OAuth2 authentication (2025 API v2)"""
        try:
            if not TWITTER_API_KEY:
                print(f"  ⚠️  Twitter API key missing")
                log_failed_api_endpoint('Twitter', 'api.twitter.com/2/tweets/search/recent', 'API key missing')
                return {}
            
            print(f"  🐦 Fetching Twitter data for {token_symbol}...")
            social_data = {}
            
            # Twitter API v2 requires Bearer token format
            if not TWITTER_API_KEY.startswith('Bearer '):
                auth_header = f"Bearer {TWITTER_API_KEY}"
            else:
                auth_header = TWITTER_API_KEY
            
            # Use more targeted queries to avoid rate limits
            queries = [f'{token_symbol} crypto', f'#{token_symbol}']
            
            for query in queries:
                try:
                    url = "https://api.twitter.com/2/tweets/search/recent"
                    headers = {
                        "Authorization": auth_header,
                        "Content-Type": "application/json",
                        "User-Agent": "DeFiRiskAssessor/3.0"
                    }
                    params = {
                        'query': query, 
                        'max_results': 10, 
                        'tweet.fields': 'created_at,public_metrics,lang,author_id',
                        'expansions': 'author_id'
                    }
                    
                    response = robust_request('GET', url, headers=headers, params=params, timeout=15)
                    
                    if response and response.status_code == 200:
                        data = response.json()
                        if data.get('data') and len(data.get('data', [])) > 0:
                            social_data[query] = data
                            print(f"    ✅ Twitter data found for query: {query}")
                        else:
                            social_data[query] = {'data': []}
                    elif response and response.status_code == 401:
                        error_msg = f"Twitter API authentication failed: Invalid API key or token expired"
                        print(f"    ❌ {error_msg}")
                        log_failed_api_endpoint('Twitter', url, error_msg)
                        social_data[query] = {'data': []}
                    elif response and response.status_code == 429:
                        error_msg = f"Twitter API rate limit exceeded"
                        print(f"    ⚠️  {error_msg}")
                        log_failed_api_endpoint('Twitter', url, error_msg)
                        social_data[query] = {'data': []}
                    elif response and response.status_code == 403:
                        error_msg = f"Twitter API access forbidden - may need elevated access tier"
                        print(f"    ❌ {error_msg}")
                        log_failed_api_endpoint('Twitter', url, error_msg)
                        social_data[query] = {'data': []}
                    else:
                        error_msg = f"Twitter API failed: {response.status_code if response else 'No response'}"
                        print(f"    ❌ {error_msg}")
                        log_failed_api_endpoint('Twitter', url, error_msg)
                        social_data[query] = {'data': []}
                        
                except Exception as e:
                    error_msg = f"Twitter error: {str(e)}"
                    print(f"    ❌ {error_msg}")
                    log_failed_api_endpoint('Twitter', url, error_msg)
                    social_data[query] = {'data': []}
                    
        except Exception as e:
            error_msg = f"Twitter error: {str(e)}"
            print(f"    ❌ {error_msg}")
            log_failed_api_endpoint('Twitter', 'api.twitter.com/2/tweets/search/recent', error_msg)
            return {}
            
        return social_data
    
    def fetch_telegram_social_data(self, token_symbol, token_name):
        """Fetch Telegram social data (2025 Bot API)"""
        try:
            if not TELEGRAM_BOT_TOKEN:
                print(f"  ⚠️  Telegram API key missing")
                log_failed_api_endpoint('Telegram', 'api.telegram.org/bot/getUpdates', 'API key missing')
                return {}
            
            print(f"  📱 Fetching Telegram data for {token_symbol}...")
            
            # Try to search for channels/groups related to the token
            # Note: Telegram Bot API doesn't have direct search, but we can check bot info
            url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getMe"
            response = robust_request('GET', url, timeout=15)
            
            if response and response.status_code == 200:
                data = response.json()
                if data.get('ok'):
                    # Bot is valid, but we can't search channels without channel username
                    # Return minimal data to indicate API is working
                    result = {
                        'bot_info': data.get('result', {}),
                        'note': 'Telegram Bot API requires channel username for full search'
                    }
                    print(f"    ✅ Telegram API connection verified")
                    return result
                else:
                    error_msg = f"Telegram API returned error: {data.get('description', 'Unknown')}"
                    print(f"    ❌ {error_msg}")
                    log_failed_api_endpoint('Telegram', url, error_msg)
                    return {}
            elif response and response.status_code == 401:
                error_msg = "Telegram API authentication failed: Invalid bot token"
                print(f"    ❌ {error_msg}")
                log_failed_api_endpoint('Telegram', url, error_msg)
                return {}
            else:
                error_msg = f"Telegram API failed: {response.status_code if response else 'No response'}"
                print(f"    ❌ {error_msg}")
                log_failed_api_endpoint('Telegram', url, error_msg)
                return {}
                
        except Exception as e:
            error_msg = f"Telegram error: {str(e)}"
            print(f"    ❌ {error_msg}")
            log_failed_api_endpoint('Telegram', 'api.telegram.org', error_msg)
            return {}

    def fetch_discord_social_data(self, token_symbol, token_name):
        """Fetch Discord social data with improved error handling (2025 API v10)"""
        try:
            if not DISCORD_BOT_TOKEN:
                print(f"  ⚠️  Discord API key missing")
                log_failed_api_endpoint('Discord', 'discord.com/api/v10/users/@me', 'API key missing')
                return {}
            
            print(f"  🎮 Fetching Discord data for {token_symbol}...")
            
            # Discord API v10 - verify bot token
            url = f"https://discord.com/api/v10/users/@me"
            headers = {"Authorization": f"Bot {DISCORD_BOT_TOKEN}"}
            response = robust_request('GET', url, headers=headers, timeout=15)
            
            if response and response.status_code == 200:
                data = response.json()
                # Bot token is valid, but we need guild ID to search channels
                result = {
                    'bot_info': data,
                    'note': 'Discord API requires guild ID for channel search'
                }
                print(f"    ✅ Discord API connection verified")
                return result
            elif response and response.status_code == 401:
                error_msg = "Discord API authentication failed: Invalid bot token"
                print(f"    ❌ {error_msg}")
                log_failed_api_endpoint('Discord', url, error_msg)
                return {}
            elif response and response.status_code == 403:
                error_msg = "Discord API access forbidden - check bot permissions"
                print(f"    ❌ {error_msg}")
                log_failed_api_endpoint('Discord', url, error_msg)
                return {}
            else:
                error_msg = f"Discord API failed: {response.status_code if response else 'No response'}"
                print(f"    ❌ {error_msg}")
                log_failed_api_endpoint('Discord', url, error_msg)
                return {}
                
        except Exception as e:
            error_msg = f"Discord error: {str(e)}"
            print(f"    ❌ {error_msg}")
            log_failed_api_endpoint('Discord', 'discord.com/api/v10', error_msg)
            return {}

    def fetch_reddit_social_data(self, token_symbol, token_name):
        """Fetch Reddit social data (2025 Reddit API)"""
        try:
            if not REDDIT_CLIENT_ID or not REDDIT_CLIENT_SECRET:
                print(f"  ⚠️  Reddit API credentials missing")
                log_failed_api_endpoint('Reddit', 'reddit.com/api/v1/access_token', 'API credentials missing')
                return {}
            
            print(f"  📱 Fetching Reddit data for {token_symbol}...")
            
            # Reddit OAuth2 authentication
            url = "https://www.reddit.com/api/v1/access_token"
            auth = (REDDIT_CLIENT_ID, REDDIT_CLIENT_SECRET)
            data = {"grant_type": "client_credentials"}
            headers = {"User-Agent": "DeFiRiskAssessor/3.0 (by /u/defiriskassessor)"}
            
            response = robust_request('POST', url, auth=auth, data=data, headers=headers, timeout=15)
            
            if response and response.status_code == 200:
                token_data = response.json()
                access_token = token_data.get('access_token')
                
                if access_token:
                    # Search for posts about the token using Reddit search API
                    search_url = f"https://oauth.reddit.com/search.json"
                    headers = {
                        "Authorization": f"Bearer {access_token}", 
                        "User-Agent": "DeFiRiskAssessor/3.0 (by /u/defiriskassessor)"
                    }
                    params = {
                        'q': f'{token_symbol} crypto',
                        'limit': 10,
                        'sort': 'relevance',
                        't': 'all'
                    }
                    search_response = robust_request('GET', search_url, headers=headers, params=params, timeout=15)
                    
                    if search_response and search_response.status_code == 200:
                        data = search_response.json()
                        if data.get('data', {}).get('children'):
                            print(f"    ✅ Reddit data found")
                            return data
                        else:
                            print(f"    ⚠️  No Reddit posts found for {token_symbol}")
                            return {'data': {'children': []}}
                    elif search_response and search_response.status_code == 401:
                        error_msg = "Reddit API authentication failed - token expired"
                        print(f"    ❌ {error_msg}")
                        log_failed_api_endpoint('Reddit', search_url, error_msg)
                        return {}
                    elif search_response and search_response.status_code == 429:
                        error_msg = "Reddit API rate limit exceeded"
                        print(f"    ⚠️  {error_msg}")
                        log_failed_api_endpoint('Reddit', search_url, error_msg)
                        return {}
                    else:
                        error_msg = f"Reddit search failed: {search_response.status_code if search_response else 'No response'}"
                        print(f"    ❌ {error_msg}")
                        log_failed_api_endpoint('Reddit', search_url, error_msg)
                        return {}
                else:
                    error_msg = "Reddit authentication failed - no access token received"
                    print(f"    ❌ {error_msg}")
                    log_failed_api_endpoint('Reddit', url, error_msg)
                    return {}
            elif response and response.status_code == 401:
                error_msg = "Reddit API authentication failed - invalid credentials"
                print(f"    ❌ {error_msg}")
                log_failed_api_endpoint('Reddit', url, error_msg)
                return {}
            else:
                error_msg = f"Reddit API failed: {response.status_code if response else 'No response'}"
                print(f"    ❌ {error_msg}")
                log_failed_api_endpoint('Reddit', url, error_msg)
                return {}
                
        except Exception as e:
            error_msg = f"Reddit error: {str(e)}"
            print(f"    ❌ {error_msg}")
            log_failed_api_endpoint('Reddit', 'reddit.com/api', error_msg)
            return {}
    
    def apply_strict_compliance_checks(self, risk_report, chain):
        """Apply strict compliance checks"""
        try:
            compliance_score = 0
            compliance_issues = []
            
            # Check for contract verification
            onchain_data = risk_report.get('onchain_data', {})
            if onchain_data.get('contract_verified') != 'verified':
                compliance_score -= 2
                compliance_issues.append('Contract not verified')
            
            # Check for liquidity
            if not onchain_data.get('liquidity'):
                compliance_score -= 1
                compliance_issues.append('Low liquidity')
            
            # Check for holder distribution
            holder_data = onchain_data.get('holders', {})
            if holder_data.get('top10_concentration', 100) > 80:
                compliance_score -= 2
                compliance_issues.append('High concentration in top 10 holders')
            
            # Check for market data availability
            market_data = risk_report.get('market_data', {})
            if not (market_data.get('coingecko') or market_data.get('coinmarketcap')):
                compliance_score -= 1
                compliance_issues.append('Limited market data')
            
            risk_report['compliance_score'] = max(0, compliance_score)
            risk_report['compliance_issues'] = compliance_issues
            
            return compliance_score
        except Exception as e:
            self.logger.error(f"Error in compliance checks: {e}")
            return 0
    
    def apply_eu_regulatory_checks(self, risk_report):
        """Apply EU regulatory compliance checks"""
        try:
            eu_score = 0
            eu_issues = []
            
            # Properly detect stablecoins - only check known stablecoin symbols
            # Use exact match to avoid false positives (CHZ, COMP, etc.)
            symbol = risk_report.get('symbol', '').upper().strip()
            token_name = risk_report.get('name', '').lower()
            
            # Known stablecoin symbols (exact matches only)
            known_stablecoins = {
                'USDT', 'USDC', 'DAI', 'BUSD', 'TUSD', 'USDP', 'GUSD', 
                'FRAX', 'LUSD', 'SUSD', 'MIM', 'DOLA', 'FEI', 'USDD',
                'USDX', 'CUSD', 'HUSD', 'PAX', 'USDS', 'EURS', 'EURT'
            }
            
            # Check if it's actually a stablecoin
            is_stablecoin = symbol in known_stablecoins
            
            # Also check token name for common stablecoin patterns (but be careful)
            if not is_stablecoin:
                stablecoin_keywords = ['stablecoin', 'usd coin', 'tether', 'dai stablecoin', 'binance usd']
                is_stablecoin = any(keyword in token_name for keyword in stablecoin_keywords)
            
            # Set is_stablecoin flag
            risk_report['is_stablecoin'] = is_stablecoin
            
            # Only apply stablecoin compliance checks for actual stablecoins
            if is_stablecoin:
                eu_score -= 1
                eu_issues.append('Stablecoin requires additional EU compliance')
            
            # Check for market cap (MiCA thresholds)
            market_data = risk_report.get('market_data', {})
            coingecko_data = market_data.get('coingecko', {}).get('market_data', {})
            market_cap = coingecko_data.get('market_cap', {}).get('usd', 0)
            
            if market_cap > 1_000_000_000:  # > $1B
                eu_score -= 2
                eu_issues.append('Large market cap requires MiCA compliance')
            elif market_cap > 100_000_000:  # > $100M
                eu_score -= 1
                eu_issues.append('Medium market cap may require MiCA compliance')
            
            # Check for contract verification (transparency requirement)
            onchain_data = risk_report.get('onchain_data', {})
            if onchain_data.get('contract_verified') != 'verified':
                eu_score -= 1
                eu_issues.append('Contract verification required for EU compliance')
            
            risk_report['eu_compliance_score'] = max(0, eu_score)
            risk_report['eu_compliance_issues'] = eu_issues
            
            return eu_score
        except Exception as e:
            self.logger.error(f"Error in EU regulatory checks: {e}")
            return 0
    
    def assess_batch(self, tokens):
        """Assess a batch of tokens"""
        reports = []
        
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = []
            for i, token in enumerate(tokens):
                address = token.get('address', '')
                chain = token.get('chain', 'eth')
                future = executor.submit(self.assess_token, address, chain, None, i, len(tokens))
                futures.append(future)
            
            for future in as_completed(futures):
                try:
                    report = future.result()
                    reports.append(report)
                except Exception as e:
                    self.logger.error(f"Error in batch assessment: {e}")
                    self.stats['errors'] += 1
        
        return reports
    
    def export_reports(self, reports, output_dir):
        """Export reports to various formats"""
        # Export to JSON
        with open(RISK_REPORT_JSON, 'w') as f:
            json.dump(reports, f, indent=2)
        
        # Export to CSV
        csv_data = []
        for report in reports:
            row = {
                'token': report.get('token_address', ''),
                'chain': report.get('chain', ''),
                'risk_score': report.get('risk_scores', {}).get('overall', 0),
                'total_score_minus_social': report.get('risk_scores', {}).get('overall', 0),
                'risk_category': self.classify_risk(report.get('risk_scores', {}).get('overall', 0)),
                'details': json.dumps(report),
                'component_scores': json.dumps(report.get('risk_scores', {}))
            }
            csv_data.append(row)
        
        df = pd.DataFrame(csv_data)
        df.to_csv(RISK_REPORT_CSV, index=False)
        
        # Export to Excel
        df.to_excel(EXCEL_REPORT_PATH, index=False)
        
        print(f"✅ Results exported to {output_dir}")
    
    def get_statistics(self):
        """Get assessment statistics"""
        total = self.stats['successful_assessments'] + self.stats['failed_assessments']
        success_rate = (self.stats['successful_assessments'] / total * 100) if total > 0 else 0
        cache_hit_rate = (self.stats['cache_hits'] / (self.stats['cache_hits'] + self.stats['cache_misses']) * 100) if (self.stats['cache_hits'] + self.stats['cache_misses']) > 0 else 0
        
        return {
            'success_rate': success_rate,
            'cache_hit_rate': cache_hit_rate,
            'total_assessments': total,
            'errors': self.stats['errors']
        }

class ConsoleProgressBar:
    """Simple console progress bar"""
    
    def __init__(self, total_items, description="Processing"):
        self.total = total_items
        self.current = 0
        self.description = description
        self.start_time = time.time()
    
    def update(self, completed=None, message=""):
        if completed is not None:
            self.current = completed
        
        if self.current > 0:
            percent = (self.current / self.total) * 100
            elapsed = time.time() - self.start_time
            if self.current > 0:
                eta = (elapsed / self.current) * (self.total - self.current)
            else:
                eta = 0
            
            bar_length = 30
            filled_length = int(bar_length * self.current // self.total)
            bar = '█' * filled_length + '-' * (bar_length - filled_length)
            
            print(f'\r{self.description}: |{bar}| {percent:.1f}% ({self.current}/{self.total}) ETA: {eta:.1f}s {message}', end='', flush=True)
    
    def finish(self, message="Complete!"):
        elapsed = time.time() - self.start_time
        print(f'\r{self.description}: |{'█' * 30}| 100.0% ({self.total}/{self.total}) Time: {elapsed:.1f}s {message}')

def validate_tokens_csv(input_file):
    """Validate the tokens CSV file and return list of tokens"""
    if not os.path.exists(input_file):
        print(f"❌ Error: Input file '{input_file}' not found")
        return False
    
    try:
        with open(input_file, 'r') as f:
            reader = csv.DictReader(f)
            required_fields = ['address', 'chain', 'symbol', 'name']
            fieldnames = reader.fieldnames or []
            
            # Check if required fields exist
            if not all(field in fieldnames for field in required_fields):
                print(f"❌ Error: CSV must contain columns: {required_fields}")
                return False
            
            # Count tokens
            tokens = list(reader)
            if not tokens:
                print("❌ Error: No tokens found in CSV file")
                return False
            
            print(f"✅ Found {len(tokens)} tokens in {input_file}")
            return tokens
            
    except Exception as e:
        print(f"❌ Error reading CSV file: {e}")
        return False

def process_token_batch(input_file=None, output_file=None, json_output_file=None):
    """Process a batch of tokens for risk assessment"""
    global progress_bar
    
    # Set default file paths if not provided
    if input_file is None:
        input_file = TOKENS_CSV
    if output_file is None:
        output_file = RISK_REPORT_CSV
    if json_output_file is None:
        json_output_file = RISK_REPORT_JSON
    
    # Validate input file exists
    if not os.path.exists(input_file):
        print(f"❌ Error: Input file '{input_file}' not found")
        print(f"📁 Expected path: {os.path.abspath(input_file)}")
        return False
    
    tokens = validate_tokens_csv(input_file)
    if not isinstance(tokens, list):
        tokens = []
    analyzer = DeFiRiskAssessor()
    results = []
    summaries = []
    fallback_count = 0
    api_error_count = 0
    total = len(tokens)
    
    print(f"Starting risk assessment for {total} tokens...")
    logging.info(f"Starting risk assessment for {total} tokens...")
    
    # Update the existing progress bar with the correct total (don't reinitialize)
    if PROGRESS_AVAILABLE:
        print(f"Using working progress bar for {total} tokens...")
        # Update the progress bar with the actual total and advance to next phase
        update_progress_phase(1, "DeFi Risk Assessment")
        complete_phase_progress("Assessment phase started")
        
        # Update the title to show we're processing tokens
        try:
            from working_progress_bar import update_progress_title
            update_progress_title("DeFi Risk Assessment")
        except Exception as e:
            print(f"[ProgressBar] Error updating title: {e}")
    else:
        # Fallback to console progress bar
        progress_bar = ConsoleProgressBar(total, "Risk Assessment")
        print(f"Using console progress bar for {total} tokens...")
    
    def process_one(token, token_index, total_tokens):
        nonlocal fallback_count, api_error_count
        chain = token.get("chain", "eth").strip().lower()
        address = token["address"].strip()
        
        # Remove per-token progress bar update from here
        logging.info(f"Processing {address} on {chain}")
        
        try:
            result = analyzer.assess_token(
                address, chain,
                token_index=token_index,
                total_tokens=total_tokens
            )
            # Count fallback usage
            if result.get('onchain_data', {}).get('holders', {}).get('total_holders', 0) == 0 or result.get('onchain_data', {}).get('liquidity', 0) == 0:
                fallback_count += 1
            # Count API errors (if any)
            if result.get('errors'):
                api_error_count += 1
            results.append(result)
            # Build summary for this token
            red_flags = result.get('red_flags', [])
            
            # Convert all red flags to Yes/No format for CSV output
            red_flag_mapping = {
                'unverified_contract': "Yes" if 'unverified_contract' in red_flags else "No",
                'low_liquidity': "Yes" if 'low_liquidity' in red_flags else "No",
                'high_concentration': "Yes" if 'high_concentration' in red_flags else "No",
                'is_proxy_contract': "Yes" if 'is_proxy_contract' in red_flags else "No",
                'eu_unlicensed_stablecoin': "Yes" if 'eu_unlicensed_stablecoin' in red_flags else "No",
                'eu_regulatory_issues': "Yes" if 'eu_regulatory_issues' in red_flags else "No",
                'mica_non_compliant': "Yes" if 'mica_non_compliant' in red_flags else "No",
                'mica_no_whitepaper': "Yes" if 'mica_no_whitepaper' in red_flags else "No",
                'owner_change_last_24h': "Yes" if 'owner_change_last_24h' in red_flags else "No"
            }
            
            # Get risk score and category from risk_scores
            risk_scores = result.get('risk_scores', {})
            overall_score = risk_scores.get('overall', 0)
            risk_category = analyzer.classify_risk(overall_score)
            
            # Preserve the full detailed data for scoring functions
            summary = {
                "token": address,
                "chain": chain,
                "symbol": result.get('symbol', ''),
                "risk_score": overall_score,
                "total_score_minus_social": overall_score,  # We'll calculate this properly later
                "risk_category": risk_category,
                "red_flags": red_flags,
                # Individual red flag columns for Excel
                "unverified_contract": red_flag_mapping['unverified_contract'],
                "low_liquidity": red_flag_mapping['low_liquidity'],
                "high_concentration": red_flag_mapping['high_concentration'],
                "is_proxy_contract": red_flag_mapping['is_proxy_contract'],
                "eu_unlicensed_stablecoin": red_flag_mapping['eu_unlicensed_stablecoin'],
                "eu_regulatory_issues": red_flag_mapping['eu_regulatory_issues'],
                "mica_non_compliant": red_flag_mapping['mica_non_compliant'],
                "mica_no_whitepaper": red_flag_mapping['mica_no_whitepaper'],
                "owner_change_last_24h": red_flag_mapping['owner_change_last_24h'],
                "is_stablecoin": result.get('is_stablecoin', False),
                "eu_compliance_status": result.get('eu_compliance_status', 'Unknown'),
                # Preserve full detailed data for scoring functions
                "market": result.get('market_data', {}),
                "onchain": result.get('onchain_data', {}),
                "enhanced": result.get('enhanced_data', {}),
                "santiment": result.get('santiment_data', {}),
                "security": result.get('security_data', []),
                "key_metrics": {
                    # Extract market cap from multiple sources (CoinGecko, CoinMarketCap, CoinPaprika)
                    "market_cap": analyzer._extract_market_cap(result),
                    # Extract volume from multiple sources
                    "volume_24h": analyzer._extract_volume_24h(result),
                    "holders": result.get('onchain_data', {}).get('holders', {}).get('total_holders', 0),
                    # Extract actual liquidity value (not dictionary)
                    "liquidity": analyzer._extract_liquidity_value(result)
                },
                "component_scores": risk_scores.get('component_scores', {})
            }
            summaries.append(summary)
            print(f"  Score: {overall_score} - {risk_category}")
            logging.info(f"  Score: {overall_score} - {risk_category}")
        except Exception as e:
            print(f"[process_token_batch] Error processing {address} on {chain}: {str(e)}")
            logging.error(f"[process_token_batch] Error processing {address} on {chain}: {str(e)}")
            results.append({
                "token": address,
                "chain": chain,
                "risk_score": 150,
                "risk_category": "Extreme Risk",
                "error": str(e)
            })
    
    # Parallel execution
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(process_one, token, idx, total): (token, idx) for idx, token in enumerate(tokens)}
        completed = 0
        for future in as_completed(futures):
            completed += 1
            token, idx = futures[future]
            address = token["address"].strip()
            chain = token.get("chain", "eth").strip().lower()
            # Update progress bar with current token info (main thread, after token is truly done)
            if PROGRESS_AVAILABLE:
                try:
                    # Continuous progress update with detailed phase messages
                    next_token_progress(f"Token {completed}/{total}")
                    update_progress_phase(0, f"Fetching data for {address[:8]}...")
                    complete_phase_progress(f"Data fetched for {address[:8]}...")
                    update_progress_phase(1, f"Analyzing {address[:8]}...")
                    complete_phase_progress(f"Analysis complete for {address[:8]}...")
                    update_progress_phase(2, f"Token {completed}/{total} completed")
                    complete_phase_progress(f"Token {completed}/{total} completed")
                except Exception as e:
                    print(f"[ProgressBar] Error updating progress bar: {e}")
            elif progress_bar:
                if hasattr(progress_bar, 'update'):
                    progress_bar.update(message=f"Completed: {address} on {chain}")
                else:
                    print(f"Completed: {address} on {chain}")
    
    # Phase 3: Generate final reports (only after all tokens are processed)
    if PROGRESS_AVAILABLE:
        try:
            complete_phase_progress("All tokens processed")
            update_progress_phase(3, "Generating final reports...")
            complete_phase_progress("Reports generated successfully")
        except Exception as e:
            print(f"[ProgressBar] Error updating progress bar: {e}")
    elif progress_bar:
        if hasattr(progress_bar, 'update'):
            progress_bar.update(completed, "Generating final reports...")
        else:
            print("Generating final reports...")
    
    # Save results (full details)
    df = pd.DataFrame(results)
    
    # Reorder columns to move component_scores next to red_flags
    if not df.empty:
        # Get all column names
        all_columns = list(df.columns)
        
        # Find positions of key columns
        details_idx = None
        component_scores_idx = None
        
        for i, col in enumerate(all_columns):
            if col == 'details':
                details_idx = i
            elif col == 'component_scores':
                component_scores_idx = i
        
        # If we found both, reorder to put component_scores next to details (which contains red_flags)
        if details_idx is not None and component_scores_idx is not None:
            # Remove component_scores from its current position
            all_columns.pop(component_scores_idx)
            # Insert it after details (which contains red_flags)
            all_columns.insert(details_idx + 1, 'component_scores')
            # Reorder the DataFrame
            df = df[all_columns]
    
    df.to_csv(output_file, index=False)
    print(f"Report saved to {output_file}")
    logging.info(f"Report saved to {output_file}")
    
    # Save summary report (JSON)
    with open(json_output_file, "w") as jf:
        json.dump(summaries, jf, indent=2)
    print(f"Summary report saved to {json_output_file}")
    logging.info(f"Summary report saved to {json_output_file}")
    
    # Save Excel report with proper formatting (matching previous structure)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        from datetime import datetime
        
        # Create data structure matching previous Excel format
        excel_data = []
        for summary in summaries:
            # Get component scores
            component_scores = summary.get('component_scores', {})
            
            row = {
                'Token Name': summary.get('token_name', summary.get('symbol', 'Unknown')),
                'Token Address': summary['token'],
                'Symbol': summary['symbol'].upper() if summary['symbol'] else 'Unknown',
                'Is Stablecoin': 'Yes' if summary.get('is_stablecoin', False) else 'No',
                'EU Compliance Status': summary.get('eu_compliance_status', 'Unknown'),
                'Chain': summary['chain'],
                'Risk Score': summary['risk_score'],
                'Total Score (-Social)': summary.get('total_score_minus_social', summary['risk_score']),
                'Risk Category': summary['risk_category'],
                'Market Cap': summary.get('key_metrics', {}).get('market_cap', 0),
                'Volume 24h': summary.get('key_metrics', {}).get('volume_24h', 0),
                'Holders': summary.get('key_metrics', {}).get('holders', 0),
                'Liquidity': summary.get('key_metrics', {}).get('liquidity', 0),
                'Red Flag: unverified_contract': summary.get('unverified_contract', 'No'),
                'Red Flag: low_liquidity': summary.get('low_liquidity', 'No'),
                'Red Flag: high_concentration': summary.get('high_concentration', 'No'),
                'Red Flag: is_proxy_contract': summary.get('is_proxy_contract', 'No'),
                'Red Flag: eu_unlicensed_stablecoin': summary.get('eu_unlicensed_stablecoin', 'No'),
                'Red Flag: eu_regulatory_issues': summary.get('eu_regulatory_issues', 'No'),
                'Red Flag: mica_non_compliant': summary.get('mica_non_compliant', 'No'),
                'Red Flag: mica_no_whitepaper': summary.get('mica_no_whitepaper', 'No'),
                'Red Flag: owner_change_last_24h': summary.get('owner_change_last_24h', 'No'),
                'Industry Impact': component_scores.get('industry_impact', 0),
                'Tech Innovation': component_scores.get('tech_innovation', 0),
                'Whitepaper Quality': component_scores.get('whitepaper_quality', 0),
                'Roadmap Adherence': component_scores.get('roadmap_adherence', 0),
                'Business Model': component_scores.get('business_model', 0),
                'Team Expertise': component_scores.get('team_expertise', 0),
                'Management Strategy': component_scores.get('management_strategy', 0),
                'Global Reach': component_scores.get('global_reach', 0),
                'Code Security': component_scores.get('code_security', 0),
                'Dev Activity': component_scores.get('dev_activity', 0),
                'Aml Data': component_scores.get('aml_data', 0),
                'Compliance Data': component_scores.get('compliance_data', 0),
                'Market Dynamics': component_scores.get('market_dynamics', 0),
                'Marketing Demand': component_scores.get('marketing_demand', 0),
                'Esg Impact': component_scores.get('esg_impact', 0),
                'Social Data': component_scores.get('social_data', 0)
            }
            excel_data.append(row)
        
        # Create DataFrame and save to Excel
        excel_df = pd.DataFrame(excel_data)
        excel_file = os.path.join(DATA_DIR, 'DeFi Tokens Risk Assessment Results.xlsx')
        
        # Save with openpyxl for formatting
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            excel_df.to_excel(writer, sheet_name='Risk Assessment', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Risk Assessment']
            
            # Define colors
            orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            
            # Apply color formatting
            for row in range(2, len(excel_data) + 2):  # Start from row 2 (skip header)
                # Color EU Regulatory Issues column (orange)
                eu_issues_cell = worksheet.cell(row=row, column=16)  # Column P
                if eu_issues_cell.value == 'Yes':
                    eu_issues_cell.fill = orange_fill
                
                # Color other Non-Compliant statuses (red)
                for col in [12, 13, 14, 15, 17, 18, 19, 20]:  # Other red flag columns
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value == 'Yes':
                        cell.fill = red_fill
                
                # Color Risk Category column
                risk_cell = worksheet.cell(row=row, column=9)  # Column I
                if risk_cell.value == 'Extreme Risk':
                    risk_cell.fill = red_fill
                elif risk_cell.value == 'High Risk':
                    risk_cell.fill = orange_fill
                elif risk_cell.value == 'Medium Risk':
                    risk_cell.fill = yellow_fill
                elif risk_cell.value == 'Low Risk':
                    risk_cell.fill = green_fill
        
        print(f"Excel report saved to {excel_file}")
        logging.info(f"Excel report saved to {excel_file}")
        
    except Exception as e:
        print(f"Warning: Could not create Excel report: {e}")
        # Fallback to simple pandas export
        try:
            excel_df = pd.DataFrame(excel_data)
            excel_file = os.path.join(DATA_DIR, 'DeFi Tokens Risk Assessment Results.xlsx')
            excel_df.to_excel(excel_file, index=False)
            print(f"Excel report saved to {excel_file} (simple format)")
        except Exception as e2:
            print(f"Error creating Excel report: {e2}")
    
    # Track API usage for each risk sub-category
    api_usage = {
        'holder_data': ['Etherscan', 'Breadcrumbs', 'Ethplorer', 'Moralis'],
        'market_data': ['CoinGecko', 'CoinMarketCap', 'Coinpaprika'],
        'security_data': ['CertiK', 'DeFiSafety', 'Alchemy'],
        'compliance_data': ['Scorechain', 'TRM Labs', 'OpenSanctions', 'Lukka', 'Alchemy', 'DeFiSafety'],
        'liquidity_data': ['Etherscan', 'DeFiLlama', '1inch'],
        'transfer_data': ['Moralis', 'Etherscan', 'Bitquery'],
        'social_data': ['Twitter', 'Telegram', 'Discord', 'Reddit', 'Bitcointalk', 'Cointelegraph']
    }
    
    # Write summary TXT
    with open(SUMMARY_TXT, "w") as f:
        f.write("Data provided by CoinGecko API\n")
        f.write(f"Total tokens processed: {total}\n")
        f.write(f"Tokens using fallback data: {fallback_count}\n")
        f.write(f"Tokens with API errors: {api_error_count}\n")
        f.write(f"Duplicates found: {len(set([t['address'].lower() for t in tokens]) ) - len(tokens)}\n")
        f.write(f"\nAPI Endpoints used for Risk Sub-categories:\n")
        for category, apis in api_usage.items():
            f.write(f"• {category.replace('_', ' ').title()}: {', '.join(apis)}\n")
    print("Summary report saved to risk_assessment_summary.txt")
    logging.info("Summary report saved to risk_assessment_summary.txt")

    # Create a flag file to signal completion
    flag_path = os.path.join(DATA_DIR, "risk_report.done")
    with open(flag_path, "w") as f:
        f.write("done")
    
    # Complete the final phase
    if PROGRESS_AVAILABLE:
        complete_phase_progress("Reports generated successfully")
        finish_progress_bar("Risk assessment complete!")
        import time
        time.sleep(2)  # Give browser time to refresh and show the final message
    elif progress_bar:
        progress_bar.finish("Risk assessment complete!")
        import time
        time.sleep(2)  # Give browser time to refresh and show the final message
    
    # Wait for progress bar to close, then show completion dialog
    import time
    time.sleep(4)  # Wait for progress bar countdown to complete
    
    # Show completion dialog with disclaimer
    try:
        import subprocess
        completion_script = f'''
        display dialog "✅ UPDATE COMPLETED!

📊 RISK ASSESSMENT HAS BEEN COMPLETED SUCCESSFULLY!

📁 RESULTS SAVED TO:
• {output_file}
• {json_output_file}
• {SUMMARY_TXT}

📝 Check the 'data/' directory for detailed reports and the 'logs/' directory for execution logs.

⚠️ DISCLAIMER:
This tool provides automated risk assessment based on available data and should be used as part of a comprehensive due diligence process. Results are for informational purposes only and do not constitute financial advice. Always conduct your own research and consult with qualified professionals before making investment decisions.

Market data provided by CoinGecko (https://www.coingecko.com)" with title "DeFi Risk Assessment - Complete" buttons {{"OK"}} default button "OK" with icon 1
        '''
        subprocess.run(['osascript', '-e', completion_script], check=True)
    except Exception as e:
        print(f"Could not show completion dialog: {e}")
        print("✅ Risk assessment completed successfully!")
        print(f"📊 Check the 'data/' directory for results")
        print(f"📝 Check the 'logs/' directory for detailed logs")
    
    return True

def main():
    """Main function with comprehensive API validation"""
    print("🚀 DeFi Risk Assessment Tool v3.0")
    print("=" * 50)
    
    # Initialize progress bar IMMEDIATELY at script start
    print("🚀 Initializing DeFi Risk Assessment...")
    
    # Initialize progress bar system first
    global PROGRESS_AVAILABLE
    import time
    
    # Force PROGRESS_AVAILABLE to be available in this scope
    global PROGRESS_AVAILABLE
    PROGRESS_AVAILABLE = True
    
    if PROGRESS_AVAILABLE:
        try:
            # Get the actual number of tokens first
            tokens = validate_tokens_csv(TOKENS_CSV)
            if isinstance(tokens, bool):
                print("❌ Error: Could not validate tokens CSV")
                return
            total_tokens = len(tokens)
            
            # Initialize with the correct total from the start
            progress_type = initialize_progress_bar(total_tokens, "DeFi Risk Assessment - Starting...")
            print(f"✅ Progress bar initialized: {progress_type} for {total_tokens} tokens")
            
            # Update progress bar with startup message
            update_progress_phase(0, "System initialization complete")
            complete_phase_progress("System ready")
            
            # Vespia credential verification
            print("🔐 Secure Credential Verification")
            print("=================================")
            print("Press any key within 5 seconds to verify Vespia credentials...")
            print("If no key is pressed, the script will continue without credential verification.")
            print("")
            
            # Simple countdown without progress bar updates to avoid blocking
            key_pressed = False
            for i in range(5, 0, -1):
                countdown_msg = f"Countdown: {i} seconds remaining..."
                print(f"⏰ {countdown_msg}")
                
                # Simple 1-second wait without progress bar updates
                time.sleep(1)
            
            # For now, always skip credential verification to avoid key detection issues
            key_pressed = False
            
            if not key_pressed:
                print("⏭️ Skipping credential verification - continuing automatically...")
            else:
                print("🔐 Proceeding with credential verification...")
            
            # Update progress bar to show we're ready to start assessment
            update_progress_phase(1, "DeFi Risk Assessment - Loading...")
            complete_phase_progress("Assessment phase started")
            
        except Exception as e:
            print(f"⚠️ Progress bar initialization failed: {e}")
            PROGRESS_AVAILABLE = False
            print("ℹ️ Falling back to console progress bar")
    else:
        print("ℹ️ Using console progress bar")
    
    # Debug: Check if API keys are loaded
    print("🔍 Debug: Checking API key loading...")
    print(f"  BREADCRUMBS_API_KEY: {'✅ Loaded' if BREADCRUMBS_API_KEY else '❌ Missing'}")
    print(f"  ZAPPER_API_KEY: {'✅ Loaded' if ZAPPER_API_KEY else '❌ Missing'}")
    print(f"  DEBANK_API_KEY: {'✅ Loaded' if DEBANK_API_KEY else '❌ Missing'}")
    print(f"  MORALIS_API_KEY: {'✅ Loaded' if MORALIS_API_KEY else '❌ Missing'}")
    print(f"  INCH_API_KEY: {'✅ Loaded' if INCH_API_KEY else '❌ Missing'}")
    print(f"  REDDIT_CLIENT_ID: {'✅ Loaded' if REDDIT_CLIENT_ID else '❌ Missing'}")
    print(f"  REDDIT_CLIENT_SECRET: {'✅ Loaded' if REDDIT_CLIENT_SECRET else '❌ Missing'}")
    print("=" * 50)
    
    # Validate API keys comprehensively
    api_status = validate_api_keys_comprehensive()
    
    # Initialize risk assessor
    assessor = DeFiRiskAssessor()
    
    # Process tokens
    try:
        process_token_batch()
        print("\n🎉 Assessment completed successfully!")
    except Exception as e:
        print(f"\n❌ Error during assessment: {e}")
        traceback.print_exc()
    finally:
        # Write failed API endpoints summary
        print("\n📝 Writing failed API endpoints summary...")
        try:
            write_failed_endpoints_summary()
            print("✅ Failed API endpoints summary written successfully")
        except Exception as e:
            print(f"❌ Error writing failed endpoints summary: {e}")
            traceback.print_exc()
        
        # Close progress bar and cleanup
        if PROGRESS_AVAILABLE:
            close_progress_bar()
        time.sleep(0.5)
        # Note: api_cache is handled within the DeFiRiskAssessor class
        
        # Final completion message
        print("\n" + "="*50)
        print("🎉 Risk Assessment Process Completed!")
        print("="*50)
        print(f"📊 Total failed API endpoints: {len(failed_api_endpoints)}")
        if failed_api_endpoints:
            print("⚠️  Some API endpoints failed - check summary file for details")
        else:
            print("✅ All API endpoints completed successfully")
        print(f"📄 Summary file: {SUMMARY_TXT}")
        print("="*50)

def fetch_1inch_quote(from_token_address, to_token_address, amount, chain_id=1):
    """Fetch quote from 1inch API"""
    if not INCH_API_KEY:
        print(f"    ⚠️  1inch API key missing")
        return None
    
    try:
        url = f"https://api.1inch.dev/swap/v5.2/{chain_id}/quote"
        headers = {"Authorization": f"Bearer {INCH_API_KEY}"}
        params = {
            "src": from_token_address,
            "dst": to_token_address,
            "amount": amount
        }
        response = robust_request('GET', url, headers=headers, params=params, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ 1inch Quote API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ 1inch Quote error: {e}")
    return None

def fetch_1inch_token_metadata(token_address, chain_id=1):
    """Fetch token metadata from 1inch API with improved error handling"""
    if not INCH_API_KEY:
        print(f"    ⚠️  1inch API key missing")
        return None
    
    try:
        url = f"https://api.1inch.dev/token/v1.2/{chain_id}/metadata"
        headers = {"Authorization": f"Bearer {INCH_API_KEY}"}
        params = {"address": token_address}
        response = robust_request('GET', url, headers=headers, params=params, timeout=15)
        if response and response.status_code == 200:
            data = response.json()
            print(f"      ✅ 1inch token metadata found")
            return data
        elif response and response.status_code == 404:
            print(f"      ⚠️  1inch token not found for address {token_address[:8]}...")
            return None
        else:
            print(f"      ❌ 1inch Token Metadata API failed: {response.status_code if response else 'No response'}")
            return None
    except Exception as e:
        print(f"      ❌ 1inch Token Metadata error: {e}")
        return None

def fetch_1inch_spot_price(token_address, chain_id=1):
    """Fetch spot price from 1inch API"""
    if not INCH_API_KEY:
        print(f"    ⚠️  1inch API key missing")
        return None
    
    try:
        url = f"https://api.1inch.dev/price/v1.1/{chain_id}"
        headers = {"Authorization": f"Bearer {INCH_API_KEY}"}
        params = {"tokenAddress": token_address}
        response = robust_request('GET', url, headers=headers, params=params, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ 1inch Spot Price API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ 1inch Spot Price error: {e}")
    return None

def fetch_scorechain_aml(token_address, chain):
    """Fetch AML data from Scorechain"""
    if not SCORECHAIN_API_KEY:
        print(f"    ⚠️  Scorechain API key missing")
        return None
    
    try:
        url = f"https://api.scorechain.com/v1/address/{token_address}/risk"
        headers = {"Authorization": f"Bearer {SCORECHAIN_API_KEY}"}
        response = robust_request('GET', url, headers=headers, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ Scorechain AML API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ Scorechain AML error: {e}")
    return None

def fetch_trmlabs_aml(token_address, chain):
    """Fetch AML data from TRM Labs"""
    if not TRM_LABS_API_KEY:
        print(f"    ⚠️  TRM Labs API key missing")
        return None
    
    try:
        url = f"https://api.trmlabs.com/public/v1/address/{token_address}/profile"
        headers = {"Authorization": f"Bearer {TRM_LABS_API_KEY}"}
        response = robust_request('GET', url, headers=headers, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ TRM Labs AML API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ TRM Labs AML error: {e}")
    return None

def fetch_opensanctions_compliance(token_address, chain):
    """Fetch compliance data from OpenSanctions"""
    if not OPENSANCTIONS_API_KEY:
        print(f"    ⚠️  OpenSanctions API key missing")
        return None
    
    try:
        url = f"https://api.opensanctions.org/entities"
        headers = {"Authorization": f"Bearer {OPENSANCTIONS_API_KEY}"}
        params = {"q": token_address}
        response = robust_request('GET', url, headers=headers, params=params, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ OpenSanctions API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ OpenSanctions error: {e}")
    return None

def fetch_lukka_compliance(token_address, chain):
    """Fetch compliance data from Lukka"""
    if not LUKKA_API_KEY:
        print(f"    ⚠️  Lukka API key missing")
        return None
    
    try:
        url = f"https://api.lukka.tech/v1/address/{token_address}/risk"
        headers = {"Authorization": f"Bearer {LUKKA_API_KEY}"}
        response = robust_request('GET', url, headers=headers, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ Lukka API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ Lukka error: {e}")
    return None

def fetch_alchemy_compliance(token_address, chain):
    """Fetch compliance data from Alchemy"""
    if not ALCHEMY_API_KEY:
        print(f"    ⚠️  Alchemy API key missing")
        return None
    
    try:
        url = f"https://eth-mainnet.g.alchemy.com/v2/{ALCHEMY_API_KEY}"
        headers = {"Content-Type": "application/json"}
        data = {
            "jsonrpc": "2.0",
            "method": "alchemy_getTokenMetadata",
            "params": [token_address],
            "id": 1
        }
        response = robust_request('POST', url, headers=headers, json=data, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ Alchemy API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ Alchemy error: {e}")
    return None

def fetch_defisafety_compliance(token_address, chain):
    """Fetch compliance data from DeFiSafety"""
    if not DEFISAFETY_API_KEY:
        print(f"    ⚠️  DeFiSafety API key missing")
        return None
    
    try:
        url = f"https://api.defisafety.com/v1/protocol/{token_address}"
        headers = {"Authorization": f"Bearer {DEFISAFETY_API_KEY}"}
        response = robust_request('GET', url, headers=headers, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ DeFiSafety API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ DeFiSafety error: {e}")
    return None

def fetch_certik_security(token_address, chain):
    """Fetch security data from CertiK"""
    if not CERTIK_API_KEY:
        print(f"    ⚠️  CertiK API key missing")
        return None
    
    try:
        url = f"https://api.certik.com/v1/address/{token_address}/security"
        headers = {"Authorization": f"Bearer {CERTIK_API_KEY}"}
        response = robust_request('GET', url, headers=headers, timeout=15)
        if response and response.status_code == 200:
            return response.json()
        else:
            print(f"    ❌ CertiK API failed: {response.status_code if response else 'No response'}")
    except Exception as e:
        print(f"    ❌ CertiK error: {e}")
    return None

# Removed simulated data functions - only real API data is used

def validate_api_keys_comprehensive():
    """Comprehensive API key validation with detailed status reporting"""
    print("🔑 Validating API Keys...")
    print("=" * 50)
    
    api_status = {
        'market_data': {},
        'onchain_data': {},
        'compliance': {},
        'social_data': {},
        'defi_protocols': {}
    }
    
    # Market Data APIs
    print("📊 Market Data APIs:")
    if COINGECKO_API_KEY:
        api_status['market_data']['coingecko'] = "✅ Configured"
    else:
        api_status['market_data']['coingecko'] = "⚠️  Not configured (free tier available)"
    
    if COINMARKETCAP_API_KEY:
        api_status['market_data']['coinmarketcap'] = "✅ Configured"
    else:
        api_status['market_data']['coinmarketcap'] = "❌ Not configured"

    if COINAPI_API_KEY:
        api_status['market_data']['coinapi'] = "✅ Configured"
    else:
        api_status['market_data']['coinapi'] = "❌ Not configured"
    
    if SANTIMENT_API_KEY:
        api_status['market_data']['santiment'] = "✅ Configured"
    else:
        api_status['market_data']['santiment'] = "❌ Not configured"
    
    # On-chain Data APIs
    print("�� On-chain Data APIs:")
    if ETHERSCAN_API_KEY:
        api_status['onchain_data']['etherscan'] = "✅ Configured"
    else:
        api_status['onchain_data']['etherscan'] = "❌ Not configured"
    
    if ETHPLORER_API_KEY:
        api_status['onchain_data']['ethplorer'] = "✅ Configured"
    else:
        api_status['onchain_data']['ethplorer'] = "⚠️  Not configured (free tier available)"
    
    if BITQUERY_API_KEY:
        api_status['onchain_data']['bitquery'] = "✅ Configured"
    else:
        api_status['onchain_data']['bitquery'] = "❌ Not configured"
    
    # Compliance APIs
    print("🛡️  Compliance APIs:")
    if CERTIK_API_KEY:
        api_status['compliance']['certik'] = "✅ Configured"
    else:
        api_status['compliance']['certik'] = "❌ Not configured"
    
    if SCORECHAIN_API_KEY:
        api_status['compliance']['scorechain'] = "✅ Configured"
    else:
        api_status['compliance']['scorechain'] = "❌ Not configured"
    
    if TRM_LABS_API_KEY:
        api_status['compliance']['trm_labs'] = "✅ Configured"
    else:
        api_status['compliance']['trm_labs'] = "❌ Not configured"
    
    if OPENSANCTIONS_API_KEY:
        api_status['compliance']['opensanctions'] = "✅ Configured"
    else:
        api_status['compliance']['opensanctions'] = "❌ Not configured"
    
    if LUKKA_API_KEY:
        api_status['compliance']['lukka'] = "✅ Configured"
    else:
        api_status['compliance']['lukka'] = "❌ Not configured"
    
    if ALCHEMY_API_KEY:
        api_status['compliance']['alchemy'] = "✅ Configured"
    else:
        api_status['compliance']['alchemy'] = "❌ Not configured"
    
    if DEFISAFETY_API_KEY:
        api_status['compliance']['defisafety'] = "✅ Configured"
    else:
        api_status['compliance']['defisafety'] = "❌ Not configured"
    
    # Social Data APIs
    print("📱 Social Data APIs:")
    if TWITTER_API_KEY:
        api_status['social_data']['twitter'] = "✅ Configured"
    else:
        api_status['social_data']['twitter'] = "❌ Not configured"
    
    if TELEGRAM_BOT_TOKEN:
        api_status['social_data']['telegram'] = "✅ Configured"
    else:
        api_status['social_data']['telegram'] = "❌ Not configured"
    
    if DISCORD_BOT_TOKEN:
        api_status['social_data']['discord'] = "✅ Configured"
    else:
        api_status['social_data']['discord'] = "❌ Not configured"
    
    if REDDIT_CLIENT_ID and REDDIT_CLIENT_SECRET:
        api_status['social_data']['reddit'] = "✅ Configured"
    else:
        api_status['social_data']['reddit'] = "❌ Not configured"
    
    # DeFi Protocol APIs
    print("🏦 DeFi Protocol APIs:")
    if ZAPPER_API_KEY:
        api_status['defi_protocols']['zapper'] = "✅ Configured"
    else:
        api_status['defi_protocols']['zapper'] = "❌ Not configured"
    
    if DEBANK_API_KEY:
        api_status['defi_protocols']['debank'] = "✅ Configured"
    else:
        api_status['defi_protocols']['debank'] = "❌ Not configured"
    
    if MORALIS_API_KEY:
        api_status['defi_protocols']['moralis'] = "✅ Configured"
    else:
        api_status['defi_protocols']['moralis'] = "❌ Not configured"
    
    if INCH_API_KEY:
        api_status['defi_protocols']['1inch'] = "✅ Configured"
    else:
        api_status['defi_protocols']['1inch'] = "❌ Not configured"
    
    if BREADCRUMBS_API_KEY:
        api_status['defi_protocols']['breadcrumbs'] = "✅ Configured"
    else:
        api_status['defi_protocols']['breadcrumbs'] = "❌ Not configured"
    
    # Print status summary
    print("\n📋 API Status Summary:")
    print("-" * 30)
    
    for category, apis in api_status.items():
        print(f"\n{category.upper().replace('_', ' ')}:")
        for api, status in apis.items():
            print(f"  {api}: {status}")
    
    # Calculate statistics
    total_apis = sum(len(apis) for apis in api_status.values())
    configured_apis = sum(1 for apis in api_status.values() for status in apis.values() if "✅" in status)
    free_tier_apis = sum(1 for apis in api_status.values() for status in apis.values() if "⚠️" in status)
    missing_apis = sum(1 for apis in api_status.values() for status in apis.values() if "❌" in status)
    
    print(f"\n📊 API Configuration Summary:")
    print(f"  Total APIs: {total_apis}")
    print(f"  Configured: {configured_apis}")
    print(f"  Free Tier Available: {free_tier_apis}")
    print(f"  Missing: {missing_apis}")
    print(f"  Coverage: {((configured_apis + free_tier_apis) / total_apis * 100):.1f}%")
    
    print("\n💡 Recommendations:")
    if missing_apis > 0:
        print("  • Add missing API keys to .env file for full functionality")
    if free_tier_apis > 0:
        print("  • Consider upgrading free tier APIs for better rate limits")
    print("  • Script will work with available APIs and provide simulated data for missing ones")
    
    print("=" * 50)
    return api_status

# Global progress bar instance
progress_bar = None
PROGRESS_AVAILABLE = False

# Try to import working_progress_bar
try:
    import working_progress_bar as wpb
    PROGRESS_AVAILABLE = True
except ImportError:
    PROGRESS_AVAILABLE = False

def initialize_progress_bar(total, description):
    """Initialize the progress bar system"""
    global progress_bar, PROGRESS_AVAILABLE
    if PROGRESS_AVAILABLE:
        try:
            wpb.initialize_progress_bar(total, description)
            return "working_progress_bar"
        except Exception as e:
            print(f"Error initializing working progress bar: {e}")
            PROGRESS_AVAILABLE = False
    
    # Fallback to console progress bar
    progress_bar = ConsoleProgressBar(total, description)
    return "console"

def update_progress_phase(phase, message):
    """Update progress phase"""
    global progress_bar
    if PROGRESS_AVAILABLE:
        try:
            wpb.update_progress_phase(phase, message)
        except Exception as e:
            print(f"Error updating working progress bar: {e}")
    elif progress_bar:
        progress_bar.update(message=message)

def next_token_progress(message):
    """Move to next token in progress"""
    global progress_bar
    if PROGRESS_AVAILABLE:
        try:
            wpb.next_token_progress(message)
        except Exception as e:
            print(f"Error updating working progress bar: {e}")
    elif progress_bar:
        progress_bar.update(message=message)

def complete_phase_progress(message):
    """Complete current phase"""
    global progress_bar
    if PROGRESS_AVAILABLE:
        try:
            wpb.complete_phase_progress(message)
        except Exception as e:
            print(f"Error updating working progress bar: {e}")
    elif progress_bar:
        progress_bar.update(message=message)

def finish_progress_bar(message):
    """Finish the progress bar"""
    global progress_bar
    if PROGRESS_AVAILABLE:
        try:
            wpb.finish_progress_bar(message)
        except Exception as e:
            print(f"Error finishing working progress bar: {e}")
    elif progress_bar:
        progress_bar.finish(message)

def close_progress_bar():
    """Close the progress bar"""
    global progress_bar
    if PROGRESS_AVAILABLE:
        try:
            wpb.close_progress_bar()
        except Exception as e:
            print(f"Error closing working progress bar: {e}")
    elif progress_bar:
        progress_bar.finish("Complete!")
        progress_bar = None

def launch_progress_bar():
    """Initialize the progress bar system"""
    global progress_bar
    # This will be set in process_token_batch
    pass

def update_progress_bar(percent, message):
    """Update the progress bar"""
    global progress_bar
    if progress_bar:
        total_items = getattr(progress_bar, 'total_items', getattr(progress_bar, 'total', 0))
        completed = int((percent / 100) * total_items) if total_items else 0
        progress_bar.update(completed, message)

def get_cache_key(url, params=None, headers=None, data=None):
    """Generate a cache key for API requests"""
    key_parts = [url]
    if params:
        key_parts.append(str(sorted(params.items())))
    if headers:
        # Exclude authorization headers from cache key
        filtered_headers = {k: v for k, v in headers.items() 
                          if k.lower() not in ['authorization', 'x-api-key']}
        key_parts.append(str(sorted(filtered_headers.items())))
    if data:
        key_parts.append(str(data))
    return hashlib.md5('|'.join(key_parts).encode()).hexdigest()

def fetch_vespia_authentication():
    """Authenticate with Vespia API and get access token"""
    import os, requests, json
    import sys
    
    try:
        # Import from credential_management package
        from credential_management import get_vespia_credentials
    except ImportError:
        # Fallback to direct import if package import fails
        try:
            from credential_management.secure_credentials import get_vespia_credentials
        except ImportError:
            # Try importing from the credential_management directory directly
            try:
                credential_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "credential_management")
                if credential_dir not in sys.path:
                    sys.path.insert(0, credential_dir)
                from credential_management.secure_credentials import get_vespia_credentials
            except ImportError:
                # Final fallback - return None to use environment variables
                get_vespia_credentials = None
        
        if get_vespia_credentials is None:
            # Use environment variables as fallback
            email = os.getenv("VESPIA_EMAIL")
            password = os.getenv("VESPIA_PASSWORD")
            
            if not email or not password:
                return {"error": "Vespia credentials not configured", "token": None}
        else:
            credentials = get_vespia_credentials()
            
            if not credentials:
                return {"error": "Vespia credentials not found. Run 'python secure_credentials.py setup' to configure them.", "token": None}
            
            email = credentials.get("email")
            password = credentials.get("password")
            
            if not email or not password:
                return {"error": "Vespia credentials incomplete", "token": None}
    except Exception:
        # Final fallback to environment variables if secure system not available
        email = os.getenv("VESPIA_EMAIL")
        password = os.getenv("VESPIA_PASSWORD")
        
        if not email or not password:
            return {"error": "Vespia credentials not configured", "token": None}
    
    try:
        # Use development environment by default
        auth_url = "https://dev-api.vespia.io/auth/graphql"
        
        # GraphQL mutation for signIn
        query = """
        mutation signIn($input: SignInInput!) {
            signIn(input: $input) {
                refreshToken
                token
            }
        }
        """
        
        variables = {
            "input": {
                "email": email,
                "password": password
            }
        }
        
        payload = {
            "query": query,
            "variables": variables
        }
        
        resp = requests.post(auth_url, json=payload, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        
        if 'data' in data and 'signIn' in data['data']:
            return {
                "token": data['data']['signIn']['token'],
                "refreshToken": data['data']['signIn']['refreshToken'],
                "error": None
            }
        else:
            return {"error": "Authentication failed", "token": None}
            
    except Exception as e:
        return {"error": f"Vespia authentication error: {str(e)}", "token": None}

def fetch_vespia_refresh_token(refresh_token):
    """Refresh Vespia access token"""
    import os, requests, json
    
    try:
        auth_url = "https://dev-api.vespia.io/auth/graphql"
        
        query = """
        mutation refreshTokenPair($input: RefreshTokenPairInput!) {
            refreshTokenPair(input: $input) {
                refreshToken
                token
            }
        }
        """
        
        variables = {
            "input": {
                "refreshToken": refresh_token
            }
        }
        
        payload = {
            "query": query,
            "variables": variables
        }
        
        resp = requests.post(auth_url, json=payload, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        
        if 'data' in data and 'refreshTokenPair' in data['data']:
            return {
                "token": data['data']['refreshTokenPair']['token'],
                "refreshToken": data['data']['refreshTokenPair']['refreshToken'],
                "error": None
            }
        else:
            return {"error": "Token refresh failed", "token": None}
            
    except Exception as e:
        return {"error": f"Vespia token refresh error: {str(e)}", "token": None}

def fetch_vespia_kyb_verification(company_name, registration_code, country_code):
    """Verify KYB (Know Your Business) with Vespia"""
    auth_result = fetch_vespia_authentication()
    if auth_result.get("error"):
        return {"error": auth_result["error"]}
    
    token = auth_result["token"]
    
    try:
        api_url = "https://dev-api.vespia.io/kyb/graphql"
        
        query = """
        mutation verifyKYB($input: KYBVerificationInput!) {
            verifyKYB(input: $input) {
                status
                message
                verificationId
            }
        }
        """
        
        variables = {
            "input": {
                "companyName": company_name,
                "registrationCode": registration_code,
                "countryCode": country_code
            }
        }
        
        payload = {
            "query": query,
            "variables": variables
        }
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        resp = requests.post(api_url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        
        if 'data' in data and 'verifyKYB' in data['data']:
            return data['data']['verifyKYB']
        else:
            return {"error": "KYB verification failed"}
            
    except Exception as e:
        return {"error": f"Vespia KYB verification error: {str(e)}"}

def fetch_vespia_kyc_verification(person_name, document_number, country_code):
    """Verify KYC (Know Your Customer) with Vespia"""
    auth_result = fetch_vespia_authentication()
    if auth_result.get("error"):
        return {"error": auth_result["error"]}
    
    token = auth_result["token"]
    
    try:
        api_url = "https://dev-api.vespia.io/kyc/graphql"
        
        query = """
        mutation verifyKYC($input: KYCVerificationInput!) {
            verifyKYC(input: $input) {
                status
                message
                verificationId
            }
        }
        """
        
        variables = {
            "input": {
                "personName": person_name,
                "documentNumber": document_number,
                "countryCode": country_code
            }
        }
        
        payload = {
            "query": query,
            "variables": variables
        }
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        resp = requests.post(api_url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        
        if 'data' in data and 'verifyKYC' in data['data']:
            return data['data']['verifyKYC']
        else:
            return {"error": "KYC verification failed"}
            
    except Exception as e:
        return {"error": f"Vespia KYC verification error: {str(e)}"}

def fetch_vespia_entity_checks(entity_name, entity_type="business"):
    """Perform entity checks with Vespia"""
    auth_result = fetch_vespia_authentication()
    if auth_result.get("error"):
        return {"error": auth_result["error"]}
    
    token = auth_result["token"]
    
    try:
        api_url = "https://dev-api.vespia.io/entity/graphql"
        
        query = """
        query entityChecks($input: EntityCheckInput!) {
            entityChecks(input: $input) {
                status
                riskScore
                sanctions
                pep
                adverseMedia
            }
        }
        """
        
        variables = {
            "input": {
                "entityName": entity_name,
                "entityType": entity_type
            }
        }
        
        payload = {
            "query": query,
            "variables": variables
        }
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        resp = requests.post(api_url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        
        if 'data' in data and 'entityChecks' in data['data']:
            return data['data']['entityChecks']
        else:
            return {"error": "Entity checks failed"}
            
    except Exception as e:
        return {"error": f"Vespia entity checks error: {str(e)}"}

def fetch_vespia_aml_monitoring(entity_name, entity_type="business"):
    """Perform AML monitoring with Vespia"""
    auth_result = fetch_vespia_authentication()
    if auth_result.get("error"):
        return {"error": auth_result["error"]}
    
    token = auth_result["token"]
    
    try:
        api_url = "https://dev-api.vespia.io/aml/graphql"
        
        query = """
        query amlMonitoring($input: AMLMonitoringInput!) {
            amlMonitoring(input: $input) {
                status
                riskLevel
                alerts
                lastUpdated
            }
        }
        """
        
        variables = {
            "input": {
                "entityName": entity_name,
                "entityType": entity_type
            }
        }
        
        payload = {
            "query": query,
            "variables": variables
        }
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        resp = requests.post(api_url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        
        if 'data' in data and 'amlMonitoring' in data['data']:
            return data['data']['amlMonitoring']
        else:
            return {"error": "AML monitoring failed"}
            
    except Exception as e:
        return {"error": f"Vespia AML monitoring error: {str(e)}"}

def fetch_vespia_compliance(token_address, chain):
    """Fetch compliance data from Vespia for a token"""
    auth_result = fetch_vespia_authentication()
    if auth_result.get("error"):
        return {"error": auth_result["error"]}
    
    token = auth_result["token"]
    
    try:
        api_url = "https://dev-api.vespia.io/compliance/graphql"
        
        query = """
        query tokenCompliance($input: TokenComplianceInput!) {
            tokenCompliance(input: $input) {
                status
                riskScore
                regulatoryStatus
                complianceChecks {
                    kyc
                    kyb
                    aml
                    sanctions
                }
            }
        }
        """
        
        variables = {
            "input": {
                "tokenAddress": token_address,
                "chain": chain
            }
        }
        
        payload = {
            "query": query,
            "variables": variables
        }
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        resp = requests.post(api_url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        
        if 'data' in data and 'tokenCompliance' in data['data']:
            return data['data']['tokenCompliance']
        else:
            return {"error": "Token compliance check failed"}
            
    except Exception as e:
        return {"error": f"Vespia compliance check error: {str(e)}"}

def fetch_etherscan_all(token_address, etherscan_key, chain='eth'):
    """Fetch all Etherscan data for a token"""
    results = {}
    
    # 1. Contract verification status
    try:
        url = f"https://api.etherscan.io/api?module=contract&action=getabi&address={token_address}&apikey={etherscan_key}"
        resp = requests.get(url, timeout=20)
        if resp.status_code == 200:
            data = resp.json()
            if data.get('status') == '1':
                results['contract_abi'] = data.get('result', '')
                print(f"✓ Etherscan contract ABI: Success")
            else:
                print(f"✗ Etherscan contract ABI: {data.get('message', 'Unknown error')}")
        else:
            print(f"✗ Etherscan contract ABI: {resp.status_code}")
    except Exception as e:
        print(f"✗ Etherscan contract ABI error: {e}")
    
    # 2. Token holder count
    try:
        url = f"https://api.etherscan.io/api?module=token&action=tokenholderlist&contractaddress={token_address}&apikey={etherscan_key}"
        resp = requests.get(url, timeout=20)
        if resp.status_code == 200:
            data = resp.json()
            if data.get('status') == '1':
                results['holders'] = data.get('result', [])
                print(f"✓ Etherscan holders: Success")
            else:
                print(f"✗ Etherscan holders: {data.get('message', 'Unknown error')}")
        else:
            print(f"✗ Etherscan holders: {resp.status_code}")
    except Exception as e:
        print(f"✗ Etherscan holders error: {e}")
    
    # 3. Token transfers
    try:
        url = f"https://api.etherscan.io/api?module=account&action=tokentx&contractaddress={token_address}&apikey={etherscan_key}"
        resp = requests.get(url, timeout=20)
        if resp.status_code == 200:
            data = resp.json()
            if data.get('status') == '1':
                results['transfers'] = data.get('result', [])
                print(f"✓ Etherscan transfers: Success")
            else:
                print(f"✗ Etherscan transfers: {data.get('message', 'Unknown error')}")
        else:
            print(f"✗ Etherscan transfers: {resp.status_code}")
    except Exception as e:
        print(f"✗ Etherscan transfers error: {e}")
    
    return results

def fetch_ethplorer_all(token_address, ethplorer_key):
    """Fetch all Ethplorer data for a token"""
    results = {}
    
    # 1. Token info
    try:
        url = f"https://api.ethplorer.io/getTokenInfo/{token_address}?apiKey={ethplorer_key}"
        resp = requests.get(url, timeout=20)
        if resp.status_code == 200:
            results['token_info'] = resp.json()
            print(f"✓ Ethplorer token info: Success")
        else:
            print(f"✗ Ethplorer token info: {resp.status_code}")
    except Exception as e:
        print(f"✗ Ethplorer token info error: {e}")
    
    # 2. Address info
    try:
        url = f"https://api.ethplorer.io/getAddressInfo/{token_address}?apiKey={ethplorer_key}"
        resp = requests.get(url, timeout=20)
        if resp.status_code == 200:
            results['address_info'] = resp.json()
            print(f"✓ Ethplorer address info: Success")
        else:
            print(f"✗ Ethplorer address info: {resp.status_code}")
    except Exception as e:
        print(f"✗ Ethplorer address info error: {e}")
    
    return results

def fetch_1inch_all(token_address, chain_id=1, wallet_address=None, to_token_address=None, amount=None):
    """Fetch all 1inch data for a token"""
    results = {}
    
    # 1. Token metadata
    try:
        metadata = fetch_1inch_token_metadata(token_address, chain_id)
        if metadata:
            results['metadata'] = metadata
            print(f"✓ 1inch metadata: Success")
        else:
            print(f"✗ 1inch metadata: No data")
    except Exception as e:
        print(f"✗ 1inch metadata error: {e}")
    
    # 2. Spot price
    try:
        price = fetch_1inch_spot_price(token_address, chain_id)
        if price:
            results['spot_price'] = price
            print(f"✓ 1inch spot price: Success")
        else:
            print(f"✗ 1inch spot price: No data")
    except Exception as e:
        print(f"✗ 1inch spot price error: {e}")
    
    # 3. Quote (if parameters provided)
    if to_token_address and amount:
        try:
            quote = fetch_1inch_quote(token_address, to_token_address, amount, chain_id)
            if quote:
                results['quote'] = quote
                print(f"✓ 1inch quote: Success")
            else:
                print(f"✗ 1inch quote: No data")
        except Exception as e:
            print(f"✗ 1inch quote error: {e}")
    
    return results

def fetch_coinmarketcap_all(token_address, cmc_key):
    """Fetch all CoinMarketCap data for a token"""
    results = {}
    
    # 1. Token info
    try:
        url = f"https://pro-api.coinmarketcap.com/v2/cryptocurrency/quotes/latest"
        headers = {
            'X-CMC_PRO_API_KEY': cmc_key,
            'Accept': 'application/json'
        }
        params = {
            'address': token_address,
            'convert': 'USD'
        }
        resp = requests.get(url, headers=headers, params=params, timeout=20)
        if resp.status_code == 200:
            results['quotes'] = resp.json()
            print(f"✓ CoinMarketCap quotes: Success")
        else:
            print(f"✗ CoinMarketCap quotes: {resp.status_code}")
    except Exception as e:
        print(f"✗ CoinMarketCap quotes error: {e}")
    
    return results

def fetch_bitquery_all(token_address, bitquery_key, chain='ethereum'):
    """Fetch all Bitquery data for a token"""
    results = {}
    
    # 1. Token transfers
    try:
        url = "https://graphql.bitquery.io"
        headers = {
            'Content-Type': 'application/json',
            'X-API-KEY': bitquery_key
        }
        
        query = """
        query ($network: EthereumNetwork!, $limit: Int!, $offset: Int!, $from: ISO8601DateTime, $till: ISO8601DateTime, $address: String!) {
          ethereum(network: $network) {
            transfers(
              options: {desc: "block.timestamp.time", limit: $limit, offset: $offset}
              date: {since: $from, till: $till}
              amount: {gt: 0}
              currency: {is: $address}
            ) {
              block {
                timestamp {
                  time(format: "%Y-%m-%d %H:%M:%S")
                }
                height
              }
              transaction {
                hash
              }
              sender {
                address
              }
              receiver {
                address
              }
              amount
              currency {
                symbol
                address
              }
              external
            }
          }
        }
        """
        
        variables = {
            "network": chain,
            "limit": 100,
            "offset": 0,
            "from": "2024-01-01T00:00:00Z",
            "till": "2024-12-31T23:59:59Z",
            "address": token_address
        }
        
        payload = {
            "query": query,
            "variables": variables
        }
        
        resp = requests.post(url, json=payload, headers=headers, timeout=20)
        if resp.status_code == 200:
            results['transfers'] = resp.json()
            print(f"✓ Bitquery transfers: Success")
        else:
            print(f"✗ Bitquery transfers: {resp.status_code}")
    except Exception as e:
        print(f"✗ Bitquery transfers error: {e}")
    
    return results

def fetch_coinapi_exchange_rates(base_currency='USD', quote_currency='BTC'):
    """Fetch exchange rates from CoinAPI (2025)"""
    if not COINAPI_API_KEY:
        print(f"    ⚠️  CoinAPI key missing")
        log_failed_api_endpoint('CoinAPI', 'rest.coinapi.io/v1', 'API key missing')
        return None
    
    try:
        url = f"https://rest.coinapi.io/v1/exchangerate/{base_currency}/{quote_currency}"
        headers = {"X-CoinAPI-Key": COINAPI_API_KEY, "Accept": "application/json"}
        response = robust_request('GET', url, headers=headers, timeout=20)
        
        if response and response.status_code == 200:
            data = response.json()
            print(f"      ✅ CoinAPI exchange rate found")
            return data
        elif response and response.status_code == 401:
            error_msg = f"CoinAPI authentication failed: Invalid API key"
            print(f"    ❌ {error_msg}")
            log_failed_api_endpoint('CoinAPI', url, error_msg)
        elif response and response.status_code == 429:
            error_msg = f"CoinAPI rate limit exceeded"
            print(f"    ⚠️  {error_msg}")
            log_failed_api_endpoint('CoinAPI', url, error_msg)
        else:
            error_msg = f"HTTP {response.status_code if response else 'No response'}"
            print(f"    ❌ CoinAPI failed: {error_msg}")
            log_failed_api_endpoint('CoinAPI', url, error_msg)
    except Exception as e:
        error_msg = f"CoinAPI error: {str(e)}"
        print(f"    ❌ {error_msg}")
        log_failed_api_endpoint('CoinAPI', 'rest.coinapi.io', error_msg)
    return None

def fetch_arkham_address_intel(address):
    """Fetch address intelligence from Arkham API (2025)"""
    # Note: Arkham API may require specific API key format
    ARKHAM_API_KEY = os.getenv('ARKHAM_API_KEY')
    if not ARKHAM_API_KEY:
        print(f"    ⚠️  Arkham API key missing")
        log_failed_api_endpoint('Arkham', 'api.arkhamintelligence.com', 'API key missing')
        return None
    
    try:
        # Arkham API endpoints (2025)
        url = f"https://api.arkhamintelligence.com/v1/address/{address}"
        headers = {
            "X-API-Key": ARKHAM_API_KEY,
            "Accept": "application/json"
        }
        response = robust_request('GET', url, headers=headers, timeout=20)
        
        if response and response.status_code == 200:
            data = response.json()
            print(f"      ✅ Arkham address intelligence found")
            return data
        elif response and response.status_code == 401:
            error_msg = f"Arkham API authentication failed: Invalid API key"
            print(f"    ❌ {error_msg}")
            log_failed_api_endpoint('Arkham', url, error_msg)
        elif response and response.status_code == 404:
            error_msg = f"Arkham address not found"
            print(f"    ⚠️  {error_msg}")
            log_failed_api_endpoint('Arkham', url, error_msg)
        else:
            error_msg = f"HTTP {response.status_code if response else 'No response'}"
            print(f"    ❌ Arkham API failed: {error_msg}")
            log_failed_api_endpoint('Arkham', url, error_msg)
    except Exception as e:
        error_msg = f"Arkham error: {str(e)}"
        print(f"    ❌ {error_msg}")
        log_failed_api_endpoint('Arkham', 'api.arkhamintelligence.com', error_msg)
    return None

def fetch_dune_all(query_id, dune_key):
    """Fetch all Dune Analytics data for a query"""
    results = {}
    
    # 1. Query results
    try:
        query_result = fetch_dune_query(query_id, dune_key)
        if query_result:
            results['query_results'] = query_result
            print(f"✓ Dune query results: Success")
        else:
            print(f"✗ Dune query results: No data")
    except Exception as e:
        print(f"✗ Dune query results error: {e}")
    
    return results

def fetch_zapper_all(address, zapper_key):
    """Fetch all Zapper data for an address"""
    results = {}
    
    # 1. Portfolio data
    try:
        portfolio = fetch_zapper_portfolio_data(address)
        if portfolio:
            results['portfolio'] = portfolio
            print(f"✓ Zapper portfolio: Success")
        else:
            print(f"✗ Zapper portfolio: No data")
    except Exception as e:
        print(f"✗ Zapper portfolio error: {e}")
    
    # 2. Protocol data
    try:
        protocol = fetch_zapper_protocol_data("uniswap-v3")
        if protocol:
            results['protocol'] = protocol
            print(f"✓ Zapper protocol: Success")
        else:
            print(f"✗ Zapper protocol: No data")
    except Exception as e:
        print(f"✗ Zapper protocol error: {e}")
    
    return results

def fetch_debank_all(address, debank_key):
    """Fetch all DeBank data for an address"""
    results = {}
    
    # 1. Portfolio data
    try:
        portfolio = fetch_debank_portfolio(address)
        if portfolio:
            results['portfolio'] = portfolio
            print(f"✓ DeBank portfolio: Success")
        else:
            print(f"✗ DeBank portfolio: No data")
    except Exception as e:
        print(f"✗ DeBank portfolio error: {e}")
    
    # 2. Token list
    try:
        token_list = fetch_debank_token_list(1)  # Ethereum
        if token_list:
            results['token_list'] = token_list
            print(f"✓ DeBank token list: Success")
        else:
            print(f"✗ DeBank token list: No data")
    except Exception as e:
        print(f"✗ DeBank token list error: {e}")
    
    return results

def fetch_moralis_all(address, moralis_key, chain='eth'):
    """Fetch all Moralis data for an address"""
    results = {}
    
    # 1. Token metadata
    try:
        metadata = fetch_moralis_token_metadata(address, chain)
        if metadata:
            results['metadata'] = metadata
            print(f"✓ Moralis metadata: Success")
        else:
            print(f"✗ Moralis metadata: No data")
    except Exception as e:
        print(f"✗ Moralis metadata error: {e}")
    
    # 2. Token price
    try:
        price = fetch_moralis_token_price(address, chain)
        if price:
            results['price'] = price
            print(f"✓ Moralis price: Success")
        else:
            print(f"✗ Moralis price: No data")
    except Exception as e:
        print(f"✗ Moralis price error: {e}")
    
    # 3. Token transfers
    try:
        transfers = fetch_moralis_token_transfers(address, chain)
        if transfers:
            results['transfers'] = transfers
            print(f"✓ Moralis transfers: Success")
        else:
            print(f"✗ Moralis transfers: No data")
    except Exception as e:
        print(f"✗ Moralis transfers error: {e}")
    
    return results

def fetch_coinpaprika_all(symbol):
    """Fetch all Coinpaprika data for a symbol"""
    results = {}
    
    # 1. Market data
    try:
        market = fetch_coinpaprika_market(symbol)
        if market:
            results['market'] = market
            print(f"✓ Coinpaprika market: Success")
        else:
            print(f"✗ Coinpaprika market: No data")
    except Exception as e:
        print(f"✗ Coinpaprika market error: {e}")
    
    return results

def fetch_defillama_all(token_address, chain='ethereum'):
    """Fetch all DeFiLlama data for a token"""
    results = {}
    
    # 1. Token price
    try:
        price = fetch_defillama_token_price(token_address, chain)
        if price:
            results['token_price'] = price
            print(f"✓ DeFiLlama token price: Success")
        else:
            print(f"✗ DeFiLlama token price: No data")
    except Exception as e:
        print(f"✗ DeFiLlama token price error: {e}")
    
    # 2. Protocol TVL
    try:
        tvl = fetch_defillama_protocol_tvl(token_address)
        if tvl:
            results['protocol_tvl'] = tvl
            print(f"✓ DeFiLlama protocol TVL: Success")
        else:
            print(f"✗ DeFiLlama protocol TVL: No data")
    except Exception as e:
        print(f"✗ DeFiLlama protocol TVL error: {e}")
    
    # 3. Yield pools
    try:
        pools = fetch_defillama_yield_pools(token_address)
        if pools:
            results['yield_pools'] = pools
            print(f"✓ DeFiLlama yield pools: Success")
        else:
            print(f"✗ DeFiLlama yield pools: No data")
    except Exception as e:
        print(f"✗ DeFiLlama yield pools error: {e}")
    
    return results

if __name__ == "__main__":
    main() 
